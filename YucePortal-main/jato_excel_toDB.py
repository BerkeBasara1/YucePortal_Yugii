import pandas as pd
import openpyxl
import datetime
from db_funcs import *
from config import *
import os
import glob


def Jato_excel_toDB(excel_path):
    today = datetime.datetime.today()

    xl = pd.read_excel(excel_path, sheet_name='Value Summary', header=None)

    query = "SELECT MAX(Process_ID) AS Highest_Process_ID FROM [Raporlar].[dbo].[Jato_Data];"
    highest_process_id_in_db = AssignDBContenttoListWithQuery(YuceDB, query)
    try:
        process_id = highest_process_id_in_db[0] + 1
    except:
        process_id = 1001

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['Value Summary']


    sheet1 = workbook['JATO Worksheet - Sheet 1'] # The Sheet which has the MY info inside

    model_year_values = []

    # Loop through the rows in the sheet and find the row where column A contains "Model year"
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "Model year":
                model_year_row = cell.row  # The Cell where "Model year" is located
                break

    try:
        for cell in sheet1[model_year_row]:
            model_year_values.append(cell.value)
    except:pass
    
    start_idx = None
    for index, row in xl.iterrows():
        if '    Unique Identity' in row.values:
            start_idx = index
            break

    new_df = xl.loc[start_idx:].copy()
    new_df.reset_index(drop=True, inplace=True)

    data_columns = new_df.applymap(pd.notna).sum(axis=0) > 0
    selected_columns = new_df.loc[:, data_columns].columns.tolist()

    headerz = []
    header_count = 1
    for cell in sheet[2]:
        if cell.value != None:
            headerz.append(cell.value.replace("\n", " "))
        else:
            headerz.append(f"header{header_count}")
            header_count += 1

    # Removes the headerz' element which is after  "Average Excluding Selected Vehicles" ("None")
    value_to_find = "Average Excluding Selected Vehicles"

    try:
        index = headerz.index(value_to_find)
        if index < len(headerz) - 1:
            headerz.pop(index + 1)
    except:
        pass

    data = []
    for index, row in new_df.iterrows():
        # Sadece veri içeren sütunları al
        selected_data = row[selected_columns].tolist()
        data.append(selected_data)

    structured_df = pd.DataFrame(data, columns=headerz)

    yesil_liste_temp = [] # Temporary, we replace the elements later and append to "yesil_liste"
    yesil_liste = [] # Yeşil liste (Analiz Dökümanındaki)
    mor_liste = [] # Mor liste (Analiz Dökümanındaki)
    nested_kirmizi2_list = [] # Kırmızı (2) liste (Analiz Dökümanındaki) Her kırmızı-2 kolonu liste olarak bu listenin içine atılıyor
    nested_kirmizi1_list = [] # Kırmızı liste (Analiz Dökümanındaki) Her kırmızı kolonu liste olarak bu listenin içine atılıyor

    for column_name in structured_df.columns:
        kirmizi2_list = []
        kirmizi1_list = []
        column_data = structured_df[column_name]
        if column_name == "header1":
            yesil_liste_temp.extend([value for value in column_data])
        else:
            if "header" in column_name:
                kirmizi2_list.extend([value for value in column_data])
                nested_kirmizi2_list.append(kirmizi2_list)
            elif "Average Excluding Selected Vehicles" in column_name:
                mor_liste.extend([value for value in column_data])
            else:
                kirmizi1_list.extend([value for value in column_data])
                nested_kirmizi1_list.append(kirmizi1_list)

    for el in yesil_liste_temp:
        yesil_liste.append(el.replace("  ",""))

    #Creates the Queries and inserts
    for i in range(len(yesil_liste)):
        # Finds - Unique Identity - Make - Model - Version name -
        j = 0
        for yesil_row in yesil_liste:
            if "Unique" in yesil_row and "Id" in yesil_row:
                index_of_UniqueID = j
                break
            j += 1
        j = 0
        for yesil_row in yesil_liste:
            if "Make" in yesil_row:
                index_of_Make = j
                break
            j += 1
        j = 0
        for yesil_row in yesil_liste:
            if "Model" in yesil_row:
                index_of_Model = j
                break
            j += 1
        j = 0
        for yesil_row in yesil_liste:
            if "Version name" in yesil_row:
                index_of_Version_name = j
                break
            j += 1

        try:
            for k in range(len(yesil_liste)):
                if k != index_of_Make and k != index_of_Model and k != index_of_Version_name and k != index_of_UniqueID:
                    make = nested_kirmizi1_list[i][index_of_Make]
                    model = nested_kirmizi1_list[i][index_of_Model]
                    version = nested_kirmizi1_list[i][index_of_Version_name]
                    uniqueID = nested_kirmizi1_list[i][index_of_UniqueID]

                    property = yesil_liste[k]
                    value = nested_kirmizi1_list[i][k]
                    result = nested_kirmizi2_list[i][k]

                    query = f"INSERT INTO [Raporlar].[dbo].[Jato_Data] ([Make], [Model], [Version], [Property], [Value], [Result], [Created_Date], [Process_ID], [CarID]) VALUES ('{make}', '{model}', '{version}', '{property}', '{value}', '{result}', '{today}', {process_id}, {uniqueID});"
                    QueryToDB(YuceDB, query)
        except:pass

    # The Model year row is being inserted here (What Leyla needed in 2024 august)
    queries = []
    try:
        for i in range(len(nested_kirmizi1_list)):
            MY_query = f"INSERT INTO [Raporlar].[dbo].[Jato_Data] ([Make], [Model], [Version], [Property], [Value], [Created_Date], [Process_ID], [CarID]) VALUES ('{nested_kirmizi1_list[i][index_of_Make]}', '{nested_kirmizi1_list[i][index_of_Model]}', '{nested_kirmizi1_list[i][index_of_Version_name]}', 'Model year', '{model_year_values[i + 1]}', '{today}', {process_id}, {nested_kirmizi1_list[i][index_of_UniqueID]});"
            if MY_query not in MY_query:
                MY_query.append(MY_query)
                QueryToDB(YuceDB,MY_query)
    except:pass

    for e in range(len(mor_liste)):
        carID = "Average Excluding Selected Vehicles"
        property = yesil_liste[e]
        result = mor_liste[e]
        AESV_query = f"INSERT INTO [Raporlar].[dbo].[Jato_Data] ([Property], [Result], [Created_Date], [Process_ID], [CarID]) VALUES ('{property}', '{result}', '{today}', {process_id}, '{carID}');"
        QueryToDB(YuceDB, AESV_query)

    correction_query = "UPDATE [Raporlar].[dbo].[Jato_Data] SET Result = NULL WHERE Result = 'nan';"
    QueryToDB(YuceDB, correction_query)
    correction_query = "UPDATE [Raporlar].[dbo].[Jato_Data] SET Value = NULL WHERE Value = 'nan';"
    QueryToDB(YuceDB, correction_query)

def Jato_excel_toDB_Jato_New_Table(excel_path):
    today = datetime.datetime.today()
    xl = pd.read_excel(excel_path, sheet_name='Value Summary', header=None)
    query = "SELECT MAX(Process_ID) AS Highest_Process_ID FROM [Raporlar].[dbo].[Jato_Yeni_Tablo];"
    highest_process_id_in_db = AssignDBContenttoListWithQuery(YuceDB, query)
    try:
        process_id = highest_process_id_in_db[0] + 1
    except:
        process_id = 1001

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['Value Summary']
    sheet1 = workbook['JATO Worksheet - Sheet 1'] # The Sheet which has the MY info inside
    model_year_values = []
    # Loop through the rows in the sheet and find the row where column A contains "Model year"
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == "Model year":
                model_year_row = cell.row  # The Cell where "Model year" is located
                break
    try:
        for cell in sheet1[model_year_row]:
            model_year_values.append(cell.value)
    except:pass
    start_idx = None
    for index, row in xl.iterrows():
        if '    Unique Identity' in row.values:
            start_idx = index
            break

    new_df = xl.loc[start_idx:].copy()
    new_df.reset_index(drop=True, inplace=True)
    data_columns = new_df.applymap(pd.notna).sum(axis=0) > 0
    selected_columns = new_df.loc[:, data_columns].columns.tolist()
    headerz = []
    header_count = 1
    for cell in sheet[2]:
        if cell.value != None:
            headerz.append(cell.value.replace("\n", " "))
        else:
            headerz.append(f"header{header_count}")
            header_count += 1

    # Removes the headerz' element which is after  "Average Excluding Selected Vehicles" ("None")
    value_to_find = "Average Excluding Selected Vehicles"
    try:
        index = headerz.index(value_to_find)
        if index < len(headerz) - 1:
            headerz.pop(index + 1)
    except:
        pass
    data = []
    for index, row in new_df.iterrows():
        # Sadece veri içeren sütunları al
        selected_data = row[selected_columns].tolist()
        data.append(selected_data)
    structured_df = pd.DataFrame(data, columns=headerz)
    yesil_liste_temp = [] # Temporary, we replace the elements later and append to "yesil_liste"
    yesil_liste = [] # Yeşil liste (Analiz Dökümanındaki)
    mor_liste = [] # Mor liste (Analiz Dökümanındaki)
    nested_kirmizi2_list = [] # Kırmızı (2) liste (Analiz Dökümanındaki) Her kırmızı-2 kolonu liste olarak bu listenin içine atılıyor
    nested_kirmizi1_list = [] # Kırmızı liste (Analiz Dökümanındaki) Her kırmızı kolonu liste olarak bu listenin içine atılıyor
    for column_name in structured_df.columns:
        kirmizi2_list = []
        kirmizi1_list = []
        column_data = structured_df[column_name]
        if column_name == "header1":
            yesil_liste_temp.extend([value for value in column_data])
        else:
            if "header" in column_name:
                kirmizi2_list.extend([value for value in column_data])
                nested_kirmizi2_list.append(kirmizi2_list)
            elif "Average Excluding Selected Vehicles" in column_name:
                mor_liste.extend([value for value in column_data])
            else:
                kirmizi1_list.extend([value for value in column_data])
                nested_kirmizi1_list.append(kirmizi1_list)

    for el in yesil_liste_temp:
        yesil_liste.append(el.replace("  ",""))

    #Creates the Queries and inserts
    for i in range(len(yesil_liste)):
        # Finds - Unique Identity - Make - Model - Version name -
        j = 0
        for yesil_row in yesil_liste:
            if "Unique" in yesil_row and "Id" in yesil_row:
                index_of_UniqueID = j
                break
            j += 1
        j = 0
        for yesil_row in yesil_liste:
            if "Make" in yesil_row:
                index_of_Make = j
                break
            j += 1
        j = 0
        for yesil_row in yesil_liste:
            if "Model" in yesil_row:
                index_of_Model = j
                break
            j += 1
        j = 0
        for yesil_row in yesil_liste:
            if "Version name" in yesil_row:
                index_of_Version_name = j
                break
            j += 1
        try:
            for k in range(len(yesil_liste)):
                if k != index_of_Make and k != index_of_Model and k != index_of_Version_name and k != index_of_UniqueID:
                    make = nested_kirmizi1_list[i][index_of_Make]
                    model = nested_kirmizi1_list[i][index_of_Model]
                    version = nested_kirmizi1_list[i][index_of_Version_name]
                    uniqueID = nested_kirmizi1_list[i][index_of_UniqueID]
                    property = yesil_liste[k]
                    value = nested_kirmizi1_list[i][k]
                    result = nested_kirmizi2_list[i][k]
                    query = f"INSERT INTO [Raporlar].[dbo].[Jato_Yeni_Tablo] ([Make], [Model], [Version], [Property], [Value], [Result], [Created_Date], [Process_ID], [CarID]) VALUES ('{make}', '{model}', '{version}', '{property}', '{value}', '{result}', '{today}', {process_id}, {uniqueID});"
                    QueryToDB(YuceDB, query)
        except:pass

    # The Model year row is being inserted here (What Leyla needed in 2024 august)
    queries = []
    try:
        for i in range(len(nested_kirmizi1_list)):
            MY_query = f"INSERT INTO [Raporlar].[dbo].[Jato_Yeni_Tablo] ([Make], [Model], [Version], [Property], [Value], [Created_Date], [Process_ID], [CarID]) VALUES ('{nested_kirmizi1_list[i][index_of_Make]}', '{nested_kirmizi1_list[i][index_of_Model]}', '{nested_kirmizi1_list[i][index_of_Version_name]}', 'Model year', '{model_year_values[i + 1]}', '{today}', {process_id}, {nested_kirmizi1_list[i][index_of_UniqueID]});"
            if MY_query not in MY_query:
                MY_query.append(MY_query)
                QueryToDB(YuceDB,MY_query)
    except:pass

    for e in range(len(mor_liste)):
        carID = "Average Excluding Selected Vehicles"
        property = yesil_liste[e]
        result = mor_liste[e]
        AESV_query = f"INSERT INTO [Raporlar].[dbo].[Jato_Yeni_Tablo] ([Property], [Result], [Created_Date], [Process_ID], [CarID]) VALUES ('{property}', '{result}', '{today}', {process_id}, '{carID}');"
        QueryToDB(YuceDB, AESV_query)

    correction_query = "UPDATE [Raporlar].[dbo].[Jato_Yeni_Tablo] SET Result = NULL WHERE Result = 'nan';"
    QueryToDB(YuceDB, correction_query)
    correction_query = "UPDATE [Raporlar].[dbo].[Jato_Yeni_Tablo] SET Value = NULL WHERE Value = 'nan';"
    QueryToDB(YuceDB, correction_query)

# Finds and returns the list of excel files in the given directory
def find_excel_files(directory):
    excel_files = []
    os.chdir(directory)
    for file in glob.glob("*.xlsx"):  # Search for all .xlsx files in the directory
        excel_files.append(file)
    for file in glob.glob("*.xls"):  # Search for all .xls files in the directory
        excel_files.append(file)
    return excel_files


