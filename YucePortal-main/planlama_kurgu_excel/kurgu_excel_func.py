import openpyxl
from openpyxl.styles import PatternFill
import datetime


def kurgu_excel_func(excel_path, toptan_from_OraDB, perakende_from_OraDB, YSstok_from_OraDB, StokFiktif_from_OraDB, Toplam_Ay_Gelis_from_OraDB, excel_path_to_be_saved):
    excel_file = openpyxl.load_workbook(excel_path, data_only=False)
    sheet = excel_file["Kurgu"]


    def month_name_to_number(month_no):
        months = {
            1 : 'Ocak',
            2 : 'Şubat',
            3 : 'Mart',
            4 : 'Nisan',
            5 : 'Mayıs',
            6 : 'Haziran',
            7 : 'Temmuz',
            8 : 'Ağustos',
            9 : 'Eylül',
            10 : 'Ekim',
            11 : 'Kasım',
            12 : 'Aralık'
        }
        # Convert month name to number
        return months.get(month_no)

    today = datetime.date.today()
    bu_ay = today.month
    sonraki_ay = today + datetime.timedelta(days=32)
    sonraki_ay = sonraki_ay.month
    search_text1 = "{BU_AY}"
    replacement_text1 = month_name_to_number(bu_ay)
    search_text2 = "{SONRAKI_AY}"
    replacement_text2 = month_name_to_number(sonraki_ay)

    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=sheet.max_row):
        for cell in col:
            if isinstance(cell.value, str) and search_text1 in cell.value:
                cell.value = cell.value.replace(search_text1, replacement_text1)
            elif isinstance(cell.value, str) and search_text2 in cell.value:
                cell.value = cell.value.replace(search_text2, replacement_text2)

    class Car:
        def __init__(self, top_model, base_model, toptan=None, perakende=None, ys_stok=None, stok_fiktif=None):
            self.TopModel = top_model
            self.BaseModel = base_model
            self.Toptan = toptan
            self.Perakende = perakende
            self.YSstok = ys_stok
            self.StokFiktif = stok_fiktif

    top_models_list = []
    models_list = []
    # Tüm Paket ve Motor kırılımlı modellerin olduğu bir liste yaratma (Herhangi bir kolonda adet olduğu anda listeye eklenir)
    for model in toptan_from_OraDB:
        if model[1] not in models_list:
            top_models_list.append(model[0].replace("Yeni ",""))
            models_list.append(model[1])
    for model in perakende_from_OraDB:
        if model[1] not in models_list:
            top_models_list.append(model[0].replace("Yeni ",""))
            models_list.append(model[1])
    for model in YSstok_from_OraDB:
        if model[1] not in models_list:
            top_models_list.append(model[0].replace("Yeni ",""))
            models_list.append(model[1])
    for model in StokFiktif_from_OraDB:
        if model[1] not in models_list:
            top_models_list.append(model[0].replace("Yeni ",""))
            models_list.append(model[1])
    for model in Toplam_Ay_Gelis_from_OraDB:
        if model[1] not in models_list:
            top_models_list.append(model[0].replace("Yeni ",""))
            models_list.append(model[1])
    
    # Excel üzerindeki rowlar: model kırılımı
    fabia_index = 2
    scala_index = 21
    octavia_index = 37
    superb_index = 55
    kamiq_index = 71
    karoq_index = 83
    kodiaq_index = 99
    for i in range(len(top_models_list)):
        if top_models_list[i] == "Fabia":
            cell_A = sheet["A" + str(fabia_index)]
            cell_B = sheet["B" + str(fabia_index)]
            fabia_index += 1
        elif top_models_list[i] == "Scala":
            cell_A = sheet["A" + str(scala_index)]
            cell_B = sheet["B" + str(scala_index)]
            scala_index += 1
        elif top_models_list[i] == "Octavia":
            cell_A = sheet["A" + str(octavia_index)]
            cell_B = sheet["B" + str(octavia_index)]
            octavia_index += 1
        elif top_models_list[i] == "Superb":
            cell_A = sheet["A" + str(superb_index)]
            cell_B = sheet["B" + str(superb_index)]
            superb_index += 1
        elif top_models_list[i] == "Kamiq":
            cell_A = sheet["A" + str(kamiq_index)]
            cell_B = sheet["B" + str(kamiq_index)]
            kamiq_index += 1
        elif top_models_list[i] == "Karoq":
            cell_A = sheet["A" + str(karoq_index)]
            cell_B = sheet["B" + str(karoq_index)]
            karoq_index += 1
        elif top_models_list[i] == "Kodiaq":
            cell_A = sheet["A" + str(kodiaq_index)]
            cell_B = sheet["B" + str(kodiaq_index)]
            kodiaq_index += 1
        cell_A.value = top_models_list[i]
        cell_B.value = models_list[i]

    excel_aracmodel_asdict = {}
    for cell in sheet['B']:
        if  cell.value != None:
            excel_aracmodel_asdict[cell.value] = cell.row

    for aracmodel_toptan in toptan_from_OraDB:
        if aracmodel_toptan[1] in excel_aracmodel_asdict.keys():
            loc = "D" + str(excel_aracmodel_asdict[aracmodel_toptan[1]])
            sheet[loc] = aracmodel_toptan[2]
    
    for aracmodel_toptan in perakende_from_OraDB:
        if aracmodel_toptan[1] in excel_aracmodel_asdict.keys():
            loc = "E" + str(excel_aracmodel_asdict[aracmodel_toptan[1]])
            sheet[loc] = aracmodel_toptan[2]
    
    for aracmodel_toptan in YSstok_from_OraDB:
        if aracmodel_toptan[1] in excel_aracmodel_asdict.keys():
            loc = "F" + str(excel_aracmodel_asdict[aracmodel_toptan[1]])
            sheet[loc] = aracmodel_toptan[2]
    
    for aracmodel_toptan in StokFiktif_from_OraDB:
        if aracmodel_toptan[1] in excel_aracmodel_asdict.keys():
            loc = "G" + str(excel_aracmodel_asdict[aracmodel_toptan[1]])
            sheet[loc] = aracmodel_toptan[2]
    
    for aracmodel_Gelis in Toplam_Ay_Gelis_from_OraDB:
        if aracmodel_Gelis[1] in excel_aracmodel_asdict.keys():
            loc = "I" + str(excel_aracmodel_asdict[aracmodel_Gelis[1]])
            sheet[loc] = aracmodel_Gelis[2]

        
    for row in range(sheet.max_row, 0, -1):

        # Deletes the rows where A column value is none
        if sheet.cell(row=row, column=1).value is None: 
            sheet.delete_rows(row)

        # Deletes the rows where A and B columns have value but D, E, F, G is None
        elif sheet.cell(row=row, column=1).value is not None and sheet.cell(row=row, column=1).value != "-":
            if sheet.cell(row=row, column=2).value is not None:
                if sheet.cell(row=row, column=4).value is None:
                    if sheet.cell(row=row, column=5).value is None:
                        if sheet.cell(row=row, column=6).value is None:
                            if sheet.cell(row=row, column=7).value is None:
                                if sheet.cell(row=row, column=9).value is None:
                                    sheet.delete_rows(row)

    sum_columns_aslist = ["D","E","F","G","I","J","K", "L"]
    sum_rows_aslist = [1]
    # Emerald Green lines
    for row in range(2, sheet.max_row + 1):
        cell = sheet[f'B{row}']
        if cell.fill.fgColor.rgb == 'FF78FAAE':
            sum_rows_aslist.append(row)
    
    max_toplam_row = max(sum_rows_aslist) + 1
    
    # Toplam satırlarına Toplam formulu yazıyor (Electric Green satırlar)
    for column in range(len(sum_columns_aslist)):
        temp_list = []
        for row in range(len(sum_rows_aslist)):
            if row != 0:
                cell = sum_columns_aslist[column] + str(sum_rows_aslist[row])
                start_index = sum_columns_aslist[column] + str(sum_rows_aslist[row-1]+1)
                end_index = sum_columns_aslist[column] + str(sum_rows_aslist[row]-1)
                formula_ToBeWritten = f'=SUM({start_index}:{end_index})'
                sheet[cell] = formula_ToBeWritten
                temp_list.append(cell)
        formula_string = ""
        for cell in temp_list:
            formula_string = formula_string + "+" + cell
        formula_string = '=SUM(' + formula_string[1:] + ')'
        toplam_bottom_cell = sum_columns_aslist[column] + str(max_toplam_row)
        sheet[toplam_bottom_cell] = formula_string


    x = 2
    # J K L kolonlarına formulleri yazar
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value is None:
            break
        elif cell_value == "-":
            pass
        else:
            formula_J = f"=E{x}+F{x}+G{x}+I{x}" # E F G I
            formula_K = f"=E{x}"
            formula_L = f"=J{x}-K{x}"
            
            sheet[f'J{x}'] = formula_J
            sheet[f'K{x}'] = formula_K
            sheet[f'L{x}'] = formula_L
        x+=1


    excel_file.save(excel_path_to_be_saved)

            
