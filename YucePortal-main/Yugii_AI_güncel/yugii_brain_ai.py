# yugii_brain_ai.py
import os
os.environ.setdefault("TRANSFORMERS_NO_TF", "1")
os.environ.setdefault("USE_TF", "0")
import pyodbc   
from openai import OpenAI
from config import OPENAI_API_KEY, YA_2El_AracSatis,YA_RPA_MYSQL
from datetime import datetime, timedelta ,time
from flask import session
import re
import dateutil.parser
import random
import mysql.connector
import calendar
import difflib
import pandas as pd
from sentence_transformers import SentenceTransformer, util
from rapidfuzz import fuzz
from datetime import date
import requests
import oracledb
from ldap3 import Server, Connection, NTLM, ALL, SUBTREE

from Yugii_AI_güncel.yugii_general_assistant_fonk import GeneralAssistant
from Yugii_AI_güncel.NLP.yugii_brain_lists import *
from Yugii_AI_güncel.NLP import normalize as nrm
from Yugii_AI_güncel.NLP.people_schema import *
from pathlib import Path
from Yugii_AI_güncel.logger.yugii_logger import log_process_message
from Yugii_AI_güncel.logger.yugii_logger import YugiiLogger


general_assistant = GeneralAssistant()

BASE_DIR = Path(__file__).resolve().parent
PROMPT_PATH = BASE_DIR / "instructions" / "yugii_personality.txt"


def _mysql_conn():
    """MySQL bağlantısını (config üzerinden) oluşturur ve döndürür."""
    try:
        conn = mysql.connector.connect(
            host=YA_RPA_MYSQL["host"],
            user=YA_RPA_MYSQL["user"],
            password=YA_RPA_MYSQL["password"],
            database=YA_RPA_MYSQL["database"],
            charset="utf8mb4",
            use_unicode=True,
        )
        return conn
    except Exception as e:
        print(f"❌ MySQL bağlantı hatası: {e}")
        return None

client = OpenAI(api_key=OPENAI_API_KEY)

WEEKDAY_MAP = {
    "pazartesi": 0, "salı": 1, "çarşamba": 2,
    "perşembe": 3, "cuma": 4, "cumartesi": 5, "pazar": 6
}

LDAP_SERVER_HOST = "10.0.0.5"
LDAP_SERVER_PORT = 389
LDAP_SEARCH_BASE = "DC=YUCEAUTO,DC=com,DC=tr"

DEPARTMENT_DISPLAY_ALIASES = {
    "Yönetim": [
        "Yönetim",
    ],
    "Yönetim Kurulu": [
        "Yönetim Kurulu",
    ],
    "İç Denetim": [
        "İç Denetim",
        "İç Denetim Birimi",
        "Denetim",
    ],
    "Stratejik Planlama ve Mali İşler": [
        "Stratejik Planlama ve Mali İşler",
    ],
    "Stratejik Planlama ve Proje Yönetimi": [
        "Stratejik Planlama ve Proje Yönetimi",
    ],
    "Dijital Dönüşüm ve Müşteri Deneyim Yönetimi": [
        "Dijital Dönüşüm ve Müşteri Deneyim Yönetimi",
    ],
    "Satış": [
        "Satış",
    ],
    "Filo Satış": [
        "Filo Satış",
    ],
    "Ürün": [
        "Ürün",
    ],
    "Ürün ve Planlama": [
        "Ürün ve Planlama",
    ],
    "Pazarlama": [
        "Pazarlama",
    ],
    "SSH": [
        "SSH",
        "Satış Sonrası Hizmetler",
    ],
    "Bayi Geliştirme": [
        "Bayi Geliştirme",
    ],
    "İnsan Kaynakları": [
        "İnsan Kaynakları",
    ],
    "Bilgi Teknolojileri": [
        "Bilgi Teknolojileri",
    ],
    "Yeni İş Modelleri": [
        "Yeni İş Modelleri",
    ],
}

def contains_next_week_relative(text: str, rel_key: str, max_gap: int = 2) -> bool:
    tokens = text.split()

    next_idxs = [
        i for i, t in enumerate(tokens)
        if any(p in t for p in NEXT_WEEK_PREFIXES)   
    ] 
    rel_idxs = [
        i for i, t in enumerate(tokens)
        if any(r in t for r in RELATIVE_DAY_TRIGGERS[rel_key])
    ]

    for i in next_idxs:
        for j in rel_idxs:
            if abs(i - j) <= max_gap:
                return True
    return False

def gpt_fix_text(text):
    try:
        response = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Sadece yazım hatalarını düzelt.\n"
                        "AY ve GÜN ifadelerini düzeltebilirme sadece anlamaya çalış kullanıcının yazdığı tarih veya tarihleri yazım yanlışları olsa bile doğruya çevirmeye çalış.\n"
                        "❗YILI ASLA değiştirme.\n"
                        "❗Tarihi yeniden yorumlama, sadece yazım/karakter hatalarını düzelt.\n"
                        "ÖRNEK: '17 kasim 2025' → '17 kasım 2025'.\n"
                        "ÖRNEK: '18 kasım için' → '18 kasım için'.\n"
                    )
                },
                {"role": "user", "content": text}
            ],
            max_tokens=40
        )
        return response.choices[0].message.content.strip()
    except:
        return text



#otopark Önümüzdeki 7 gün  kuralı  sabah 08:00
def get_effective_today():
    now = datetime.now()
    if now.time() < time(8, 0):
        return now.date() - timedelta(days=1)
    return now.date()

class BrainAIYugii:

    _INSTANCE = None

    #ilk çağtıda class yarat, diğer çağrılarda aynı nesneyi döndür
    @classmethod
    def get_instance(cls):
        if cls._INSTANCE is None:
            cls._INSTANCE = cls()
        return cls._INSTANCE

    def __init__(self):
        # Yeni OpenAI client (v1.x uyumlu))
        self.client = OpenAI(api_key=OPENAI_API_KEY)

        # Prompt'ı yükle
        self.base_prompt = self._load_prompt()
        
        self.cities = cities 
        self.general_assistant = GeneralAssistant()
        self.people_field_roots = self._build_people_field_roots()
        self._dealer_cache = None
        self._dealer_cache_time = None

    def _canonical_department_label(self, department: str) -> str:
        if not department:
            return ""

        raw = str(department).strip()
        if not raw:
            return ""

        alias_map = {}
        for canonical, aliases in DEPARTMENT_DISPLAY_ALIASES.items():
            for alias in aliases:
                alias_map[self.normalize_people_text(alias)] = canonical

        cleaned = re.sub(r"[?�]+", "", raw)
        normalized = self.normalize_people_text(cleaned)

        if normalized in alias_map:
            return alias_map[normalized]

        matches = difflib.get_close_matches(normalized, alias_map.keys(), n=1, cutoff=0.55)
        if matches:
            return alias_map[matches[0]]

        return raw

    def _fetch_directory_people_overrides(self, usernames):
        usernames = sorted({
            str(u).strip().lower()
            for u in (usernames or [])
            if str(u).strip()
        })
        if not usernames:
            return {}

        viewer_username = str(session.get("username") or "").strip()
        viewer_password = session.get("pw")
        if not viewer_username or not viewer_password:
            return {}

        safe_usernames = [
            u for u in usernames
            if re.fullmatch(r"[a-zA-Z0-9._-]+", u)
        ]
        if not safe_usernames:
            return {}

        conn = None
        try:
            server = Server(LDAP_SERVER_HOST, port=LDAP_SERVER_PORT, get_info=ALL)
            conn = Connection(
                server,
                user=f"yuceauto\\{viewer_username}",
                password=viewer_password,
                authentication=NTLM
            )
            if not conn.bind():
                return {}

            overrides = {}

            for i in range(0, len(safe_usernames), 50):
                batch = safe_usernames[i:i + 50]
                search_filter = "(&(objectClass=user)(|{}))".format(
                    "".join(f"(sAMAccountName={u})" for u in batch)
                )
                conn.search(
                    LDAP_SEARCH_BASE,
                    search_filter,
                    search_scope=SUBTREE,
                    attributes=["sAMAccountName", "cn", "department", "userAccountControl"]
                )

                for entry in conn.entries:
                    attrs = entry.entry_attributes_as_dict

                    def first_value(key):
                        value = attrs.get(key)
                        if isinstance(value, list):
                            return value[0] if value else None
                        return value

                    username = str(first_value("sAMAccountName") or "").strip().lower()
                    if not username:
                        continue

                    user_account_control = int(first_value("userAccountControl") or 0)
                    if user_account_control & 0x0002:
                        continue

                    full_name = str(first_value("cn") or "").strip()
                    department = str(first_value("department") or "").strip()

                    overrides[username] = {
                        "full_name": full_name,
                        "department": self._canonical_department_label(department),
                    }

            return overrides

        except Exception as e:
            print("⚠️ LDAP override alınamadı:", e)
            return {}
        finally:
            try:
                if conn is not None:
                    conn.unbind()
            except Exception:
                pass

    def _apply_directory_overrides_to_rows(self, rows):
        if not rows:
            return rows

        overrides = self._fetch_directory_people_overrides([
            row.get("username")
            for row in rows
            if isinstance(row, dict)
        ])

        enriched_rows = []

        for row in rows:
            if not isinstance(row, dict):
                enriched_rows.append(row)
                continue

            item = dict(row)
            username = str(item.get("username") or "").strip().lower()
            override = overrides.get(username, {})

            full_name = str(override.get("full_name") or "").strip()
            if not full_name:
                full_name = " ".join(
                    part for part in [
                        str(item.get("name") or "").strip(),
                        str(item.get("surname") or "").strip(),
                    ]
                    if part
                ).strip()

            if full_name:
                parts = full_name.rsplit(" ", 1)
                if len(parts) == 2:
                    item["name"], item["surname"] = parts[0], parts[1]
                else:
                    item["name"] = full_name
                    item["surname"] = ""
                item["display_name"] = full_name

            department = str(override.get("department") or "").strip()
            if not department:
                department = self._canonical_department_label(item.get("departman"))
            if department:
                item["departman"] = department
                item["display_department"] = department

            enriched_rows.append(item)

        return enriched_rows

    def _oracle_conn(self):
        try:
            username = "USR_YUCE_PWR"
            password = "vJ9Q3BpH01"
            host = "10.113.197.4"
            port = 1521
            service_name = "DOTODWHDBP1"

            dsn = f"{host}:{port}/{service_name}"

            conn = oracledb.connect(
                user=username,
                password=password,
                dsn=dsn
            )

            return conn

        except Exception as e:
            print(f"❌ Oracle bağlantı hatası: {e}")
            return None
        
    def _load_prompt(self):
        """Asistanın kişiliğini içeren dosyayı okur."""
        try:
            with open(PROMPT_PATH, "r", encoding="utf-8") as f:
                return f.read()
        except Exception as e:
            print("Prompt dosyası okunamadı:", e)
            return "Sen bir şirkete ait chatbot'sun."
        
    # NORMALIZE PROXY METHODS- normalize.py den çağırmak için
    def normalize_people_text(self, text: str) -> str:
        return nrm.normalize_people_text(text)

    def normalize_calendar_text(self, text: str) -> str:
        return nrm.normalize_calendar_text(text)

    def normalize_intent_text(self, text: str) -> str:
        return nrm.normalize_intent_text(text)

    def normalize_field_name(self, field: str) -> str:
        return nrm.normalize_field_name(field)

    def normalize_people_name_token(self, token: str) -> str:
        return nrm.normalize_people_name_token(token)

    def normalize_department_suffixes(self, text: str) -> str:
        return nrm.normalize_department_suffixes(text)
    
    def normalize_month_with_suffixes(self, text: str) -> str:
        return nrm.normalize_month_with_suffixes(text)

    def format_date_with_day(self, date_obj):
        return nrm.format_date_with_day(date_obj)
    
    def fuzzy_match(self, word, options, threshold=0.75):
        return nrm.fuzzy_match(word, options, threshold)

    def fuzzy_match_in_text(self, msg: str, keywords: list, threshold=0.75) -> bool:
        return nrm.fuzzy_match_in_text(msg, keywords, threshold)

    def auto_log_sql(self, sql: str, params=None):
        
        logger = YugiiLogger.get_instance()

        # buffer varsa oraya ekle
        if hasattr(self, "_db_query_buffer"):
            self._db_query_buffer.append({
                "sql": sql,
                "params": params
            })

        logger.log_db_query(sql, params)


    def ask_gpt(self, user_message: str) -> str:
        """
        SADECE smalltalk ve yönlendirme için kullanılır.
        İşlem yapmaz, uydurmaz, yönlendirme yapmaz.
        """
        SMALLTALK_GUARD_PROMPT = """
            ROLE:
            You are Yüce Auto’s corporate digital assistant.
            This call is ONLY for light conversation and soft responses.
            
            OUTPUT LANGUAGE:
            - ALL user-facing responses MUST be in Turkish.
            
            STRICT RULES (VERY IMPORTANT):
            - NEVER perform any action.
            - NEVER imply that an action was or will be performed.
            - NEVER describe any process or ongoing work.
            - NEVER use technical terms (backend, system, portal, database, API, vb.).
            - NEVER say “I don’t understand” or similar phrases.
            - NEVER invent, assume, or infer information.
            - NEVER use future tense or imply future actions.
            - Responses must be 1–2 short sentences only.

            IDENTITY / CAPABILITY RULE (MANDATORY):

            If the user asks about:
            - what you do
            - your purpose
            - your role
            - what you are
            - your capabilities
            - your mission

            You MUST respond ONLY with this exact sentence in Turkish:

            "Ben Yüce Auto’nun kurumsal dijital asistanı Yugii’yim. Otopark rezervasyonu, haftalık takvim, çalışan bilgileri ve portal işlemlerinde yardımcı olurum."

            - Do NOT add anything.
            - Do NOT rephrase.
            - Do NOT extend.
            - Do NOT invent features.

            GREETING-ONLY CASE (CRITICAL):
            - If the user message contains ONLY a greeting
            (e.g. “selam”, “merhaba”, “naber”, “nasılsın”):
                → Respond with a short greeting only.
                → NEVER ask follow-up questions.
                → NEVER say phrases like
                “detay verir misin” or
                “nasıl yardımcı olabilirim”.
            
            - If a greeting is combined WITH a request
            (e.g. “merhaba yarın park ayır”):
                → Ignore the greeting.
                → Handle the request normally.
            
            VAGUE OR GENERAL REQUESTS:
            - Respond politely and conversationally.
            - You MAY ask for clarification ONLY if the user actually made a request.
            - In that case, you MAY say:
            “Biraz daha detay verirsen daha net yardımcı olabilirim.”
            
            UNSUPPORTED REQUESTS:
            - Politely set a boundary.
            - Say:
            “Bu konuda şu an yardımcı olamıyorum ama başka bir şey sorabilirsin.”
            
            CASUAL CONVERSATION:
            - Keep responses natural, short, and corporate.
            - Light humor is allowed.
            - NEVER invent information.
            
            IMPORTANT – STRICT BEHAVIOR RULE (MANDATORY):
            - Asla aksiyon veya süreç anlatma.
            - “kontrol ediyorum”, “bakıyorum”, “inceleyip döneceğim”, “bekle” gibi ifadeler KULLANMA.
            - Yapılan veya yapılacak bir eylemi ima etme.
            - İç işleyiş, bekleme, kontrol, işlem ifadeleri kullanma.
            - Yalnızca:
            • NET bir SONUÇ cümlesi
            • veya NET bir AÇIKLAMA / EKSİK BİLGİ sorusu üret.
            
            EXAMPLE TONES (REFERENCE ONLY):
            - “Güzel bir soru 😊 Biraz daha açarsan daha net yardımcı olabilirim.”
            - “Bu konuda şu an yardımcı olamıyorum ama başka bir şey sorabilirsin.”
            - “Rica ederim 😊 Yardımcı olabildiysem ne mutlu.”
            """


        try:
            system_prompt = (
                self.base_prompt.strip()
                + "\n\n"
                + SMALLTALK_GUARD_PROMPT.strip()
            )

            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ],
                temperature=0.4,   
                max_tokens=120
            )

            reply = response.choices[0].message.content.strip()
            reply_lower = reply.lower()

            FORBIDDEN = [
                "yapılacak", "yapılacaktır",
                "iletilecek", "iletildi",
                "backend", "sistem", "portal",
                "ilgili birim", "yönlendiriyorum",
                "işlem yapılacak", "işleme alınacak"
            ]

            if any(x in reply_lower for x in FORBIDDEN):
                return "Ne yapmak istediğini tam anlayamadım. Biraz daha detay yazar mısın?"

            return reply

        except Exception as e:
            print("GPT hata:", e)
            return (
                "Şu anda bu isteğe yanıt veremiyorum. "
                "Biraz daha net ifade edebilir misin?"
            )
        
    #REPUTATION HELPERS
    def fuzzy_contains(self, text, phrase, threshold=80):
        return fuzz.partial_ratio(text, phrase) >= threshold

    def root_match(self, text, root):
        if self.fuzzy_contains(text, root):
            return True
        if self.fuzzy_contains(text.replace(" ", ""), root.replace(" ", "")):
            return True
        return False
    
    #kur bilgisi
    def get_currency_rates(self):
        import requests
        import xml.etree.ElementTree as ET

        url = "https://www.tcmb.gov.tr/kurlar/today.xml"
        r = requests.get(url, timeout=5)
        r.raise_for_status()

        root = ET.fromstring(r.content)
        tarih = root.attrib.get("Tarih", "")

        rates = {}
        for currency in root.findall("Currency"):
            code = currency.get("CurrencyCode")
            if code in ["USD", "EUR"]:
                value = currency.find("ForexSelling").text.replace(",", ".")
                rates[code] = float(value)

        return rates, tarih
    
    def _format_currency_result(self, amount, from_unit, total, to_unit, rate_used, tarih):
        return (
            "💱 Döviz Dönüşümü<br>"
            f"{amount:,.2f} {from_unit} karşılığı "
            f"<b>{total:,.2f} {to_unit}</b><br>"
            f"<small>(1 {from_unit} = {rate_used:,.2f} TL)</small><br>"
            f"<small>{tarih} tarihli TCMB satış kuru baz alınmıştır.</small>"
        )

    def handle_currency_calculation(self, user_message: str):

        msg = self.normalize_people_text(user_message)

        # --- Tutar ---
        match = re.search(r"\d+[.,]?\d*", msg)
        if not match:
            return "Hesaplama için tutar belirtmelisin."

        amount = float(match.group().replace(",", "."))

        # --- Para Birimleri ---
        found_units = re.findall(r"(dolar|usd|euro|eur|tl|try)", msg)

        if len(found_units) < 2:
            return "Dönüştürmek istediğin iki para birimini belirtmelisin."

        unit_map = {
            "dolar": "USD",
            "usd": "USD",
            "euro": "EUR",
            "eur": "EUR",
            "tl": "TL",
            "try": "TL"
        }

        source = unit_map.get(found_units[0])
        target = unit_map.get(found_units[1])

        if source == target:
            return "Aynı para birimi arasında dönüşüm yapılamaz."

        try:
            rates, tarih = self.get_currency_rates()

            usd = rates.get("USD")
            eur = rates.get("EUR")

            if not usd or not eur:
                return "Kur bilgisi alınamadı."

            # --- TL bazına çevir ---
            if source == "USD":
                tl_amount = amount * usd
                rate_used = usd
            elif source == "EUR":
                tl_amount = amount * eur
                rate_used = eur
            elif source == "TL":
                tl_amount = amount
                rate_used = 1
            else:
                return "Desteklenmeyen para birimi."

            # --- TL'den hedefe ---
            if target == "USD":
                total = tl_amount / usd
            elif target == "EUR":
                total = tl_amount / eur
            elif target == "TL":
                total = tl_amount
            else:
                return "Desteklenmeyen para birimi."

            return self._format_currency_result(
                amount, source, total, target, rate_used, tarih
            )

        except Exception as e:
            print("❌ Döviz hesaplama hata:", e)
            return "⚠️ Hesaplama şu an yapılamıyor."
    
    #kur modülü
    def handle_currency_rate(self):
        try:
            rates, tarih = self.get_currency_rates()

            dolar = rates.get("USD")
            euro = rates.get("EUR")

            if not dolar or not euro:
                return "⚠️ Döviz kuru bilgisi alınamıyor."

            return (
                "💱 Güncel Döviz Kurları<br>"
                f"💵 1 USD = <b>{dolar:,.2f} TL</b><br>"
                f"💶 1 EUR = <b>{euro:,.2f} TL</b><br>"
                f"<small>TCMB tarafından yayımlanan son geçerli kur "
                f"({tarih}) baz alınmıştır.</small>"
            )

        except Exception as e:
            print("❌ Döviz API hata:", e)
            return "⚠️ Döviz kuru bilgisi geçici olarak alınamıyor."
    
    def _build_people_field_roots(self):
        roots = set()

        for keywords in RAW_FIELD_KEYWORDS.values():
            for k in keywords:
                k_norm = self.normalize_people_text(k)
                if " " not in k_norm and len(k_norm) >= 3:   
                    roots.add(k_norm)

        return roots
    
    def has_first_person_possessive(self, text: str) -> bool:
        text = self.normalize_people_text(text)

        # 🔒 SADECE GERÇEK 1. TEKİL ŞAHIS
        FIRST_PERSON_KEYS = {
            "benim",
            "bana",
            "bende",
            "benden",
            "bizim",
            "bizi"
        }

        return any(k in text for k in FIRST_PERSON_KEYS)

    def detect_person_prescan(self, text: str) -> list:
        """
        INTENT ÖNCESİ PERSON TESPİTİ
        - users1 tablosuna bakar
        - normalize + suffix strip kullanır
        - SADECE tespit eder (karar vermez)
        - return: ["isimsoyisim]veya []
        """

        text_norm = self.normalize_people_text(text)
        tokens = text_norm.split() 

        CORPORATE_GUARD_BASE = [
            "yuce portal","yuceportal","yuce auto","şirketimiz",
            "yuce oto","yuceauto","yuce firma","sirketimiz","firma","marka","kurum","skoda","skodanın","skoda'da","skoda'nın"
        ]

        CORPORATE_GUARD_INFO = [
            "nedir",
            "bilgi",
            "hakkinda","nasıl bir yer","nasıl bir şirket",
            "ne ise yarar","ne is yapar","hakkinda", "anlat","bilgi","fikrin","fikirilerin","neler düşünüyorsun","düşüncelerin",
        ]

        if (
            any(b in text_norm for b in CORPORATE_GUARD_BASE)
            and any(i in text_norm for i in CORPORATE_GUARD_INFO)
        ):
            return []
        
        # SESSION USER NAME OVERRIDE (ADIYLA SORARSA)
        session_name = f"{session.get('name','')} {session.get('surname','')}".strip()
        session_first = self.normalize_people_text(session.get("name",""))

        if (
            session_first
            and session_first in text_norm.split()
            and not self.has_people_field_request(text_norm)
        ):
            return [session_name]

        #  BEN / 1. TEKİL ŞAHIS → SESSION USER
        if "benim" in text_norm:
            session_name = f"{session.get('name','')} {session.get('surname','')}".strip()
            if session_name:
                return [session_name]

        # ---- Türkçe ek temizleme ----
        def strip_tr_suffix(token: str) -> str:
            SUFFIXES = ["ninki","nin","nın","nun","nün","in","ın","un","ün",
                        "ye","ya","de","da","den","dan","te","ta","si","sı","su","sü"]
            changed = True
            while changed:
                changed = False
                for s in SUFFIXES:
                    if token.endswith(s) and len(token) > len(s) + 1:
                        token = token[:-len(s)]
                        changed = True
                        break
            return token

        tokens = text_norm.split()
        tokens = [strip_tr_suffix(t) for t in tokens]

        IGNORE = {
            "mail","maili","eposta","adres",
            "telefon","numara",
            "kim","nedir","ver"
        }

        tokens = [
            t for t in tokens
            if t not in IGNORE
            and t.isalpha()
            and len(t) >= 3
        ]

        # ---- DB'den kişi isimlerini al ----
        try:
            conn = _mysql_conn()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT username, name, surname
                FROM users1
                WHERE is_active = 1
            """)
            users = self._apply_directory_overrides_to_rows(cursor.fetchall())
            cursor.close()
            conn.close()
        except Exception:
            return []

        found = []
        match_type = "NONE"

        for u in users:
            name_norm = self.normalize_people_text(u["name"])
            surname_norm = self.normalize_people_text(u["surname"])

            # 1) isim + soyisim
            if name_norm in tokens and surname_norm in tokens:
                result = [f"{u['name']} {u['surname']}"]
                match_type = "FULL_NAME"

                return result
            
            # 2) sadece soyisim
            if surname_norm in tokens and match_type == "NONE":
                match_type = "SURNAME"
                found.append(f"{u['name']} {u['surname']}")
                continue

            # 3) sadece isim
            if name_norm in tokens and match_type == "NONE":
                match_type = "NAME"
                found.append(f"{u['name']} {u['surname']}")

        try:
            session_name = f"{session.get('name','')} {session.get('surname','')}".strip()
            if session_name:
                session_name_norm = self.normalize_people_text(session.get("name",""))
                if session_name_norm in tokens and session_name not in found:
                    found.insert(0, session_name)
        except Exception:
            pass
        final_found = list(dict.fromkeys(found))

        return final_found
    
    def has_corporate_context(self, msg: str, has_bulk_people: bool) -> bool:
        msg = self.normalize_people_text(msg)

        return any([
            bool(has_bulk_people),
            any(k in msg for k in ALL_PEOPLE_KEYWORDS),
            any(k in msg for k in DEPARTMENT_WORD_VARIANTS),
            self.has_people_field_request(msg),          # mail, tel, title
            bool(self.detect_person_prescan(msg)),       # kişi adı
            self.extract_department_name(msg) is not None,  # departman
        ])
    
    def has_people_field_request(self, msg: str) -> bool:
        text = self.normalize_people_text(msg)

        for keywords in RAW_FIELD_KEYWORDS.values():
            for k in keywords:
                if self.normalize_people_text(k) in text:
                    return True
            for root in self.people_field_roots:
                if root in text:
                    return True
        return False
    
    def has_all_people_reference(self, msg: str) -> bool:
        text = self.normalize_people_text(msg)

        STRONG_PEOPLE_ROOTS = [
            "calisan",
            "personel",
            "kisi",
            "ekip",
            "herkes"
        ]
        CONTEXTUAL_ROOTS = [
            "ofis",
            "sirket",
            "isyeri",
            "tum",
            "tamami",
            "hepsi"
        ]

        #  Güçlü kök varsa direkt true
        if any(r in text for r in STRONG_PEOPLE_ROOTS):
            return True

        #  Bağlamsal kök + people kökü birlikteyse true
        if any(r in text for r in CONTEXTUAL_ROOTS) and any(
            p in text for p in STRONG_PEOPLE_ROOTS
        ):
            return True

        return False
    
    SENSITIVE_PASSWORD_KEYWORDS = [
        # Türkçe
        "sifre", "şifre",
        "kullanici sifresi", "kullanıcı şifresi",
        "pc sifresi", "pc şifresi",
        "bilgisayar sifresi", "bilgisayar şifresi",
        "portal sifresi", "portal şifresi",
        "giris sifresi", "giriş şifresi","pc şifre","pc pasword"

        # İngilizce
        "password",
        "pass",
        "login password",
        "pc password",
        "user password",
        "account password",

        # Karışık
        "mail sifresi",
        "email sifresi",
        "outlook sifresi",
        "office sifresi"
        ]
    

    def has_password_request(self, msg: str) -> bool:
        PASSWORD_GUIDE_KEYWORDS = [
        "nasil",
        "nasıl",
        "degistir",
        "değiştir",
        "guncelle",
        "güncelle",
        "sifremi degistir",
        "sifremi değiştir",
        "sifre degistirme",
        "sifre degistirilir",
        "sifre nasil",
        "password change",
        "reset",
        "yenile"
    ]
        text = self.normalize_people_text(msg)

        # 🔓 Eğer kullanıcı "nasıl / değiştir / reset" diyorsa → ŞİFRE TALEBİ DEĞİL
        if any(k in text for k in PASSWORD_GUIDE_KEYWORDS):
            return False

        # 🔒 Aksi halde → gerçek şifre talebi mi?
        return any(k in text for k in self.SENSITIVE_PASSWORD_KEYWORDS)
    
    def handle_password_guide(self):
        return (
            "🔐 Bilgisayar şifreni değiştirmek için iki yol bulunuyor:<br><br>"
            "• IT ekibiyle iletişime geçerek destek isteyebilirsin.<br>"
            "• Bilgisayarında <b>Ctrl + Alt + Delete</b> tuşlarına basarak "
            "“Parola Değiştir” seçeneğini kullanabilirsin.<br><br>"
            "Her iki yöntem de güvenli ve yaygın olarak kullanılıyor."
        )
    
    def handle_cancel_without_context(self) -> str:
        return (
            "Ne için iptal işlemi yapmak istersin? 🤔<br>"
            "İptal etmek istediğin işlemi biraz daha net yazar mısın?<br>"
            "Örnek olarak:<br>"
            "• “Yarınki veya aktif tüm <b>otopark</b> rezervasyonumu iptal et”<br>"
        )
    def get_department_count(self):
        """
        ofis_gunleri tablosuna göre
        GERÇEKTE kullanılan departman sayısını döndürür
        """
        conn = _mysql_conn()
        if conn is None:
            return None

        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(DISTINCT TRIM(LOWER(Department)))
            FROM ofis_gunleri
            WHERE Department IS NOT NULL
            AND TRIM(Department) <> ''
        """)

        count = cursor.fetchone()[0]

        cursor.close()
        conn.close()

        return count
    
    def contains_department_word(self, msg: str) -> bool:
        text = self.normalize_people_text(msg)

        for root in DEPARTMENT_WORD_VARIANTS:
            # departman, departmanlar, departmanlarda, departmanların...
            if re.search(rf"\b{root}(ler|lar|in|ın|un|ün|de|da|den|dan)?\b", text):
                return True

        return False
    
    def has_work_history_count_signal(self, msg: str) -> bool:
        """
        Yazım hatalarına dayanıklı 'kaç / ne kadar' tespiti
        """

        msg_norm = self.normalize_people_text(msg)

        for key in WORK_HISTORY_COUNT_KEYS:
            key_norm = self.normalize_people_text(key)

            # birebir
            if key_norm in msg_norm:
                return True

            # fuzzy (özellikle bozuk yazım)
            if fuzz.partial_ratio(msg_norm, key_norm) >= 85:
                return True

        return False
    
    def detect_currency_calculation(self, msg: str) -> bool:

        # sayı var mı
        has_number = bool(re.search(r"\d+[.,]?\d*", msg))

        # para birimi (kök bazlı, ekli halleri yakalar)
        has_currency = bool(re.search(
            r"\b(dolar|usd|euro|eur|tl|try|lira)\b|\$|€",
            msg
        ))

        # dönüşüm / hesap kelimeleri
        has_calc_word = bool(re.search(
            r"\b(kac|eder|ne kadar|cevir|donustur|hesapla|karsiligi)\b",
            msg
        ))

        # direkt pattern: "100 dolar tl"
        has_direct_pair = bool(re.search(
            r"\d+[.,]?\d*\s+(dolar|usd|euro|eur|tl|try)",
            msg
        ))

        return has_number and has_currency and (has_calc_word or has_direct_pair)

    def normalize_people_name_token(self, token: str) -> str:
        token = token.lower()

        SUFFIXES = [
            "ninki","ndaki","daki",
            "nin","nın","nun","nün",
            "dan","den","tan","ten",
            "yla","yle",
            "yi","yı","yu","yü",
            "in","ın","un","ün",
            "ya","ye"
        ]

        for s in SUFFIXES:
            if token.endswith(s) and len(token) > len(s) + 1:
                return token[:-len(s)]

        return token
    
    TAKVİM_KEYS = [
        "isaretle","ekle","ayarla","planla","yap","gir","yaz","tikla","yap","yp","ekle","isaretlesene",
        "belirle","kayit et","kaydet","isaretle","girermisin","isaretlermisin","calisicam","calisicam",
        "calisacagim","calisacagim","komple","calisacagim","calisacagim",
        "calisacam","calisacam","olacagim","olucam","olacagim",
        "gelecegim","gelicem","clisicam","olacagim","gelecegim","gelicem",
        "isaretle","isaretle","ekle","kaydet","planla","ayarla","olacagim","olucam","calisacagim","calisacagim","caliscam",
        "calisicam","calisicam","gelicem","gelecegim","gelcem","glcm","isarle","işaretle","ısretle","isaretle"
    ]

    def detect_haftalik_takvim_intent(self, user_message: str) -> bool:
        """
        Haftalık takvim EMİR tespiti.
        Bu fonksiyon eski detect_main_intent içindeki
        haftalık takvim mantığının AYNI HALİDİR.
        """

        msg = self.normalize_people_text(user_message)
        has_calendar_action = any(f in msg for f in self.TAKVİM_KEYS)

        # 🚫 LİSTE / EXPORT VARSA ASLA TAKVİM OLMASIN
        if any(k in msg for k in ["liste", "listesi", "export", "excel", "ver","listele"]):
            return False
        
        #  FIX: Başka bir kişi soruluyorsa BU BİR EMİR DEĞİLDİR
        if self.detect_person_prescan(user_message):
            return False
        
        #  OTOPARK kelimesi VARSA → HAFTALIK TAKVİM HİÇ ÇALIŞMAZ
        has_park = any(k in msg for k in OTOPARK_KEY)
        if has_park:
            print("⛔ [BLOCK]: Otopark intent varken haftalık takvim atlandı")


        calisma_tipleri = ["ofis","home","izin","bayi"]

        gun_kelimeleri = [
            "pazartesi","salı","sali","çarşamba","perşembe","cuma","cumartesi","pazar",
            "bugün","bugun","yarın","yarin",
            "haftaya", "gelecek hafta", "önümüzdeki hafta",
            "birdahaki hafta", "next week"
        ]

        tarih_kelimeleri = [
            "ocak","şubat","mart","nisan","mayıs","haziran","temmuz","ağustos",
            "eylül","ekim","kasım","aralık"
        ]

        work_type = self.detect_work_type(user_message)
        msg_has_tip = work_type is not None

        msg_has_gun  = any(g in msg for g in gun_kelimeleri)
        msg_has_date = any(t in msg for t in tarih_kelimeleri)
        is_question = bool(re.search(r"\b(mi|mı|mu|mü)\b", msg))
        has_count_q = any(k in msg for k in ["kaç", "kac", "kim", "kimler", "ne zaman", "hangi"])

        # ⭐ 1) TARİH + ÇALIŞMA TİPİ
        if (
            not has_park
            and msg_has_tip
            and (msg_has_gun or msg_has_date)
            and has_calendar_action              
            and not is_question
            and not has_count_q
        ):
            return True

        # ⭐ 2) ÇALIŞMA TİPİ + EMİR
        if (
            not has_park
            and msg_has_tip
            and has_calendar_action
            and not is_question
            and not has_count_q
        ):
            print("🧠 [INTENT]: haftalik_takvim (çalışma tipi + emir)")
            return True

        # ⭐ 3) GÜN + EMİR
        if (
            not has_park
            and msg_has_gun
            and has_calendar_action
            and not is_question
            and not has_count_q
        ):
            print("🧠 [INTENT]: haftalik_takvim (gün + emir)")
            return True

        # ⭐ 4) SADECE EMİR + ZAMAN
        if (
            not has_park
            and has_calendar_action
            and (msg_has_gun or msg_has_date)
            and not is_question
            and not has_count_q
        ):
            print("🧠 [INTENT]: haftalik_takvim (fiil + zaman)")
            return True

        return False

    def detect_bulk_people_context(self, msg: str) -> bool:
        msg = self.normalize_people_text(msg)

        if any(k in msg for k in ALL_PEOPLE_KEYWORDS):
            return True

        has_scope = any(w in msg for w in SCOPE_WORDS)
        has_people = any(w in msg for w in PEOPLE_WORDS)

        return has_scope and has_people
    
    #toplu company ile ilgili intent yönlendirme fonk
    def detect_business_intent(self, user_message: str, signals: dict) -> str | None:
        """
        İş domain karar motoru (DETAYLI vce GÜVENLİ)

        """

        # SIGNALS (TEK KAYNAK)
        has_date = signals["has_date"]
        has_question = signals["has_question"]
        has_work_type = signals["has_work_type"]
        has_person = signals["has_person"]
        has_department = signals["has_department"]
        has_export = signals["has_export"]
        has_history = signals["has_history"]
        has_people_field = signals["has_people_field"]
        has_bulk_people = signals.get("has_bulk_people_context")

        msg = self.normalize_people_text(user_message)

        
        # OTOPARK VARSA ASLA WORKFORCE'A GİRME
        if any(k in self.normalize_people_text(user_message) for k in OTOPARK_KEY):
            return None
        #  PORTAL LINK INTENT (SADECE NET NAVİGASYON / SAYFA İSTEĞİ)
        is_link_action = (
            any(v in msg for v in PORTAL_LINK_VERBS)
            or "link" in msg
            or "sayfa" in msg
            or "sayfası" in msg
            or "sayfasını" in msg
            or "syfa" in msg
        )

        is_information_query = any(k in msg for k in [
            "kaç", "kac", "sayısı", "sayisi",
            "nelerdir", "neler", "çeşitleri",
            "departman", "departmanlar", "departmanı",
            "departmanındaki", "departmanındakiler",
            "departmanındakilerin",
            "kim", "kimler", 
        ])

        if (
            any(h in msg for h in PORTAL_LINK_STRONG_HINTS)
            and is_link_action
            and not is_information_query
        ):
            return "portal_link"
        has_calendar_action = any(f in msg for f in self.TAKVİM_KEYS)

        LOCATION_QUESTION_WORDS = [
            "nerede", "nerde","ofiste","şirkette","evde","dışardan",
            "nereden", "nerden",
            "nereye","ofiste","ofis", "şirkette","şirket", "uzaktan","home", "bayide","bayi",
            "nerelerde", "nerdeler",
            "nerdeydi", "nerdeymis", "nerdeymiş","kimler var","kmler var"
        ]

        has_location_context = any(k in msg for k in LOCATION_QUESTION_WORDS)
        WEATHER_KEYWORDS = [
            "hava", "hava durumu", "sicaklik",
            "yagmur", "kar", "ruzgar",
            "gunesli", "bulutlu", "soguk", "sicak","otopark"
        ]

        if any(k in msg for k in WEATHER_KEYWORDS):
            return None
        
        has_month_word = any(
            re.search(rf"\b{m}\b", msg)
            for m in AY_NORMALIZE.values()
        )
        
        has_export = (
            any(k in msg for k in EXPORT_HINTS)
            and any(k in msg for k in ["excel", "exel"])
        )
        has_count = any(k in msg for k in PEOPLE_COUNT_KEYS)
        has_people_word = any(
            f" {k} " in f" {msg} "
            for k in PERSONEL_KEY
        )
        
        ALL_RELATIVE_DAYS = (
            RELATIVE_DAY_TRIGGERS["today"]
            + RELATIVE_DAY_TRIGGERS["tomorrow"]
            + RELATIVE_DAY_TRIGGERS["yesterday"]
        )

        has_numeric_date2 = bool(
            re.search(r"\b\d{1,2}\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)\b", msg)
        )

        has_real_day = (
            any(re.search(rf"\b{d}\b", msg) for d in ALL_WEEKDAY_VARIANTS)
            or any(re.search(rf"\b{r}\b", msg) for r in ALL_RELATIVE_DAYS)
            or has_numeric_date2
            or has_month_word
        )

        has_department_context = any(
            re.search(rf"\b{root}(ler|lar)?\b", msg)
            for root in DEPARTMENT_WORD_VARIANTS
        )
        BAYI_ROOTS = [
            "bayiler",
            "servis", "servisler", "servisleri",
            "satış servisi", "satis servisi",
        ]

        COUNT_HINTS = [
            "kaç", "kac",
            "adet", "adeti",
            "sayısı", "sayisi",
            "rakam", "liste", "listesi",
            "goster", "göster",
            "var"
        ]

        if (
            (
                "bayi" in msg
                and any(h in msg for h in COUNT_HINTS)
            )
            or any(root in msg for root in BAYI_ROOTS)
        ):
            return "bayi_about"
                
        if (
            any(k in msg for k in ["kaç", "kac", "sayısı", "sayisi"])
            and any(p in msg for p in ["çalışan", "calisan", "personel", "kisi", "kişi"])
            and not has_export
            and not has_real_day
            and not has_department
        ):
            return "people_count_only"
        
        SATIS_ROOTS = [
            "satış",
            "satis","satısı yapılan araba sayısı","araba sayısı"
            "satilan","arac satıldı","satıldı",
            "satılan","satıs oldu","satıs yapıldı",
        ]

        COUNT_ROOTS = [
            "kaç",
            "kac",
            "adet","ne kadar",
            "adedi","adeti","rakamı","rakam",
            "sayısı",
            "sayisi",
            "miktar","toplam satış",
            "toplam",
            "tüm","toplam kaç adet"
        ]
        # SATIŞ INTENT (daha akıllı)
        if (
            any(s in msg for s in SATIS_ROOTS)
            and any(c in msg for c in COUNT_ROOTS)
        ):
            return "satis_sayisi"
        
        print("DEBUG has_count:", has_count)
        print("DEBUG has_people_word:", has_people_word)
        print("DEBUG has_export:", has_export)
        print("DEBUG has_bulk_people:", has_bulk_people)

        #  WORK HISTORY (BEN)
        if (
            signals.get("has_work_history_intent")
            and signals.get("has_history_date")   #  AY / YIL / ARALIK
            and not signals.get("has_person")
            and not signals.get("has_department")
        ):
            return "work_history"

        #  HAFTALIK TAKVİM (EMİR) – PEOPLE FIELD DEĞİLSE
        if (
            self.detect_haftalik_takvim_intent(user_message)
            and not signals.get("has_people_field")
            and not signals.get("has_export")
        ):
            return "haftalik_takvim"
        
        has_department_word = any(k in msg for k in DEPARTMENT_WORD_VARIANTS)
        has_department_count = any(k in msg for k in DEPARTMENT_COUNT_HINTS)
        has_department_list = any(k in msg for k in DEPARTMENT_LIST_HINTS)

        # DEPARTMAN – COUNT vs LIST NET AYRIMI
        has_strict_department_count = (
            has_department_word
            and any(k in msg for k in DEPARTMENT_COUNT_HINTS)
            and not any(k in msg for k in DEPARTMENT_LIST_HINTS)
        )

        has_strict_department_list = (
            has_department_word
            and (
                any(k in msg for k in DEPARTMENT_LIST_HINTS)
                or "departmanlar" in msg
                or "isimleri" in msg
                or "çeşitleri" in msg
            )
        )
        #  WORKFORCE STATUS (KİM / NEREDE / KAÇ)
        if (
            has_date
            and signals.get("date_info")
            and signals["date_info"].get("start")
            and signals["date_info"]["start"].year > datetime.now().year + 1
            and any(k in msg for k in [
                "nasil olacak",
                "ne olacak",
                "modeli",
                "gelecekte",
                "ileride","ileride"
            ])
        ):
            return None
        if any(k in msg for k in PARK_AVAILABILITY_KEYWORDS):
            return None
        

        
        # FORCE WORKFORCE (GEÇMİŞ / BUGÜN / GELECEK)
        if (
            signals.get("has_department")
            and signals.get("has_date")
            and signals.get("has_work_type")
            and not has_calendar_action
            and not any(k in msg for k in EXPORT_HINTS)
        ):
            return "workforce_status"
        if (
            has_department
            and any(k in msg for k in PERSONEL_KEY)
            and not has_export
            and not has_people_field
            and not has_strict_department_list
            and not has_strict_department_count
            and not any(k in msg for k in EXPORT_HINTS)
        ):
            return "workforce_status"
        if (
            has_department
            and has_date
            and has_question
            and not has_export
            and not has_people_field
            and not has_strict_department_list
            and not has_strict_department_count
            and not any(k in msg for k in EXPORT_HINTS)
        ):
            return "workforce_status"
        
        # WORKFORCE STATUS
        if (
            (has_date or has_location_context)
            and (
                has_person
                or has_department
                or has_location_context
                
            )
            and (
                has_work_type
                or has_location_context
                or has_question
                or has_count
            )
            and not has_strict_department_list
            and not has_strict_department_count
            and not has_export
            and not has_calendar_action
            and not any(k in msg for k in EXPORT_HINTS)
        ):
            return "workforce_status"
        
        if (
            any(k in msg for k in EXPORT_HINTS)
            and not self.has_corporate_context(msg, has_bulk_people)
        ):
            return None   
        
        has_export_language = (
            any(e in msg for e in EXPORT_HINTS)
            or has_people_field   # mail / telefon / title
        )
        has_list_language = (
            any(p in msg for p in PERSONEL_KEY)
            or any(w in msg for w in WORKFORCE_LIST_KEYS)
        )
        #  PERSON / LIST → DIRECT EXPORT (SINGLE SOURCE OF TRUTH)
        if (
            # Kapsam: departman veya şirket geneli
            (
                has_department
                or has_bulk_people
                or signals.get("has_bulk_people_context")
                or signals.get("has_bulk_list_request")
            )

            # Liste / export dili
            and (
                has_list_language
                or has_export_language
            )

            and not any(k in msg for k in OTOPARK_KEY)
            and not has_strict_department_count
            and not has_real_day
        ):
            return "people_export"

        print("INTENT TEST MSG:", msg)


        #  COMPANY PEOPLE (KİŞİ / DEPARTMAN / ŞİRKET)
        has_people_intent = (
            has_person
            or (
                has_department
                and not signals.get("has_bulk_people_context")
            )
        )
        # TEK KİŞİ BİLGİSİ
        if (
            has_people_intent
            and not has_date
            and not has_work_type
            and not any(p in msg for p in PERSONEL_KEY)
        ):
            return "company_people"

        #  SADECE SAYI SORUSU
        if (
            has_strict_department_count
            and self.contains_department_word(user_message)
            and not has_date
            and not has_people_intent
        ):
            return "department_count"

        #  SADECE LİSTE / ÇEŞİT SORUSU
        if (
            has_strict_department_list
            and self.contains_department_word(user_message)
            and not has_date
            and not has_work_type
            and not has_people_intent
        ):
            return "department_list"
        
        return None
    
    def handle_people_count_only(self):
        conn = _mysql_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*)
            FROM users1
            WHERE is_active = 1
        """)
        self.auto_log_sql("""
            SELECT COUNT(*)
            FROM users1
            WHERE is_active = 1
        """, None)
        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

        return f"Şirkette toplam mevcut <b>{count}</b> aktif çalışan bulunuyor."

    #yeni ek detect main intetn fonksiyonu
    def extract_message_signals(self, user_message: str) -> dict:
        """
        Kullanıcı mesajından SADECE sinyal çıkarır.
        Intent, handler veya return kararı VERMEZ.
        """

        msg = self.normalize_people_text(user_message)
        msg_stripped = self._strip_tr_suffixes(msg)

        QUESTION_KEYWORDS = [
            "kim", "kimler","hangi kişiler","kç adt","sayısı","sayı ver","sayı",
            "kaç", "kac","kaç kişi","kaç tane","kaç adet","kaç olucak","kaç kişi",
            "var","nerede","nerde","nasıl","nasil",          
            "çalışıyor", "calisiyor","çalışıcak","nerde çalışıyor",
            "çalışıcak", "calisacak",
             "nasıl", "nasil","hangi", "hangileri", "nelerdir", "neler","kimler var"
        ]

        # TEMEL SİNYALLER
        has_calendar_action = any(f in msg for f in self.TAKVİM_KEYS)
        
        MONTH_KEYS = list(AY_NORMALIZE.values())
        
        has_numeric_date = bool(
            re.search(r"\b\d{1,2}\b", msg) and any(m in msg for m in MONTH_KEYS)
        ) or bool(
            re.search(r"\b20\d{2}\b", msg)   # yıl (2024, 2025…)
        )

        # Tarih sinyali (tek otorite)
        
        has_date = (
            any(
                re.search(rf"\b{re.escape(k)}\b", msg)
                for k in (
                    RELATIVE_DAY_TRIGGERS["today"]
                    + RELATIVE_DAY_TRIGGERS["tomorrow"]
                    + ALL_WEEKDAY_VARIANTS
                    + THIS_WEEK_KEYS
                    + NEXT_WEEK_KEYS
                    + THIS_MONTH_KEYS
                )
            )
            or any(re.search(rf"\b{re.escape(m)}\b", msg) for m in MONTH_KEYS)
            or has_numeric_date
        )

        has_explicit_date = (
            (
                any(
                    re.search(rf"\b{re.escape(k)}\b", msg)
                    for k in (
                        RELATIVE_DAY_TRIGGERS["today"]
                        + RELATIVE_DAY_TRIGGERS["tomorrow"]
                        + ALL_WEEKDAY_VARIANTS
                        + THIS_WEEK_KEYS
                        + NEXT_WEEK_KEYS
                        + THIS_MONTH_KEYS
                    )
                )
                or any(re.search(rf"\b{re.escape(m)}\b", msg) for m in MONTH_KEYS)
                or has_numeric_date
            )
            and not any(
                re.search(rf"\b{re.escape(m)}\b", msg)
                for m in WORKFORCE_COUNT_KEYS
            )
        )

        # Çalışma tipi (ofis / home / bayi / izin)
        work_type = self.detect_work_type(user_message)
        has_work_type = work_type is not None

        # Kişi (isim / soyisim)
        has_person = bool(
            self.detect_person_prescan(user_message)
            or self.has_person_in_db(user_message)
        )

        is_portal_link_candidate = (
            any(h in msg for h in PORTAL_LINK_STRONG_HINTS)
            and any(v in msg for v in PORTAL_LINK_VERBS)
        )

        # Departman
        department = self.extract_department_name(user_message)
        has_department = department is not None

        #  PORTAL LINK BAĞLAMI VARSA → DEPARTMAN SUSTURULUR
        if is_portal_link_candidate:
            has_department = False
            department = None

        # FIX: workforce bağlamında "bayi" work_type iken departman sayılmasın
        if work_type == "Bayi":
            has_department = False
            department = None

        has_question = (
            self.has_work_history_count_signal(msg)
            or any(k in msg for k in QUESTION_KEYWORDS)
            or bool(re.search(r"(mi|mı|mu|mü)\b", msg))
            or bool(re.search(r"(mi|mı|mu|mü)\b", msg_stripped))  # şirkettemi, uzaktanmı
        )
        
        has_bulk_list_request = (
            self.detect_bulk_people_context(msg)
            and not has_person          # kişi adı yok
            and not has_question        # soru değil
            and not any(k in msg for k in [
                "departman", "departmanlar", "departmanı",
                "birim", "birimler",
                "organizasyon"
            ])
        )

        #  EXPORT SİNYALİ
        EXPORT_KEYWORDS = [
            "excel", "exel",
            "liste", "listesi",
            "döküman", "dokuman",
            "indir", "çıkar",
            "export","döküman"
        ]
        has_export = (
            any(k in msg for k in EXPORT_HINTS)
            and any(k in msg for k in ["excel", "exel"])
        )

        #  WORK HISTORY SİNYALİ (BEN)
        HISTORY_KEYWORDS = [
            "geldim", "çalıştım", "calistim",
            "ofisteydim", "şirketteydim",
            "evdeydim", "izin aldım",
            "uzaktan çalıştım"
        ]
        has_history = any(k in msg for k in HISTORY_KEYWORDS)

        LOCATION_KEYWORDS = [
            
            "nerede","nerdede","nerde","neredeler","nerdeler"
        ]
        has_location_context = any(k in msg_stripped for k in LOCATION_KEYWORDS)

        #  ALAN (email, tel vs.)
        raw_has_people_field = (
            self.has_people_field_request(user_message)
            and not any(k in msg for k in [
                "kim", "kimler",
                "kac", "kaç", "kaç kişi", "kaç tane"
            ])
        )
        #  people_field SADECE GERÇEK KİŞİSEL ALANLAR İÇİN TRUE
        raw_has_people_field = self.has_people_field_request(user_message)

        #mail telefon gibi 
        has_people_field = (
            raw_has_people_field
            and not has_date
            and not has_work_type
        )

        #  EXPORT / PEOPLE FIELD VARSA → work_type İPTAL
        if has_export or has_people_field:
            has_work_type = False
            work_type = None

        #  TOPLU KİŞİ BAĞLAMI (şirket / departman / herkes) tek kişi değil
        has_bulk_people_context = (
            self.detect_bulk_people_context(msg)
            and not any(k in msg for k in [
                "departman", "departmanlar",
                "sayisi", "sayısı",
                "nelerdir", "neler"
            ])
        )

        DEPARTMENT_LIST_QUESTION_HINTS = [
            "hangi", "hangileri", "nelerdir", "neler",
            "liste", "listesi",
            "var", "mevcut",
            "şirketteki", "sirketteki"
        ]

        has_department_list_question = (
            any(k in msg for k in DEPARTMENT_WORD_VARIANTS)
            and (
                any(k in msg for k in ["hangi", "hangileri", "nelerdir", "neler","çeşitleri"])
                or "departmanlar" in msg
            )
            and not has_people_field   #  mail/tel varsa KAPAT
            and not has_export 
        )

        # ---- WORK HISTORY INTENT (FINAL) -----
        history_date_info = self.extract_history_dates(user_message)
        has_history_date = bool(history_date_info and history_date_info.get("start"))
        
        has_work_history_intent = (
            any(k in msg for k in [
                "kaç gün",
                "kac gun","kç gün",
                "ne kadar gün",
                "kç gün",
                "gün sayısı",
                "gun sayisi"
            ])
            and has_history_date
            and not any(k in msg for k in [
                "bugün", "yarın", "yarin",
                "şu an", "simdi"
            ])
        )

        has_work_history_count = has_work_history_intent

        #  TOPLU SONUÇ

        return {
            # temel
            "has_location_context": has_location_context,
            "has_date": has_date, 
            "has_work_type": has_work_type,
            "work_type": work_type,

            "has_person": has_person,
            "has_department": has_department,
            "department": department,

            # niyet
            "has_question": has_question,
            "has_calendar_action": has_calendar_action,
            "has_history": has_history,

            # export / directory
            "has_export": has_export,
            "has_people_field": has_people_field,
            "has_bulk_people_context": has_bulk_people_context,

            "has_department_list_question": has_department_list_question,
            "has_bulk_list_request": has_bulk_list_request,
            "has_work_history_count": has_work_history_count,
            "has_history_date": has_history_date,
            "has_work_history_intent": has_work_history_intent,

        }


    def detect_main_intent(self, user_message: str, signals: dict) -> str:

        print("\n🧪 [TEST] detect_main_intent INPUT:", user_message)

        msg = self.normalize_people_text(user_message)
        print("INTENT TEST MSG:", msg)

        #sinyal tanımlama
        has_date = signals.get("has_date")


        
        COMPANY_TERMS = [
            "yuce auto","yuce oto","skoda","sirket","firma","marka","yuceauto","skoda yüce auto","yüce auto","şirketimiz","şirketttimiz","yuce autonun","yuce autodaki","yuce auto'daki",
        ]

        is_company_question = any(t in msg for t in COMPANY_TERMS)

        has_person = bool(
            not is_company_question   #  KRİTİK
            and (
                session.get("detected_persons")
                or self.has_person_in_db(user_message)
            )
        )
        msg_norm = self.normalize_people_text(user_message)

        if (
            any(k in msg_norm for k in KIM_YAPTI_SORUSU)
            and not any(x in msg_norm for x in [   "komik", "saka", "espri", "eglendir", "guldur", "moral"
            ])
        ):
            return "Beni Yüce Auto’nun IT ekibinden Sude Bayhan geliştirdi 🙂"

        CAPABILITY_PATTERNS = [
            r"\bne\s+is\s+yap",
            r"\bne\s+yap",
            r"\bneler\s+yap",
            r"\bne\s+ise\s+yar",
            r"\bamac",
            r"\bgorev",
            r"\byetenek",
            r"\bkabiliyet",
            r"\bmisyon",
            r"\brol",
            r"\bne\s+icin\s+var",
            r"\bneden\s+burda",
            r"\bneden\s+buradasin",
            r"\bsen\s+nesin",
            r"\bsenin\s+is",
            r"\bbu\s+sistem\s+ne\s+yap",
        ]

        if any(re.search(p, msg) for p in CAPABILITY_PATTERNS):
            return "assistant_capabilities"

        # --- ROOT FALLBACK (yazım hatalı yakalama) ---
        tokens = msg.split()

        if any(t.startswith(("amac", "gorev", "yetenek", "kabiliyet")) for t in tokens):
            return "assistant_capabilities"

        if (
            "ne" in tokens and
            (
                any(t.startswith("yap") for t in tokens) or
                "ise" in tokens and any(t.startswith("yar") for t in tokens)
            )
        ):
            return "assistant_capabilities"

        if (
            "bu ve gecen ay analiz raporu hazirla" in msg_norm
        ):
            return "monthly_trend_analysis"

        for k in FUN_REQUEST_KEYWORDS:
            score = fuzz.partial_ratio(msg, k)
            if score >= 90:
                print(f"🧠 [INTENT]: fun_request (fuzzy={score})")
                return "fun_request"
            

        # BUGÜN / GÜN SORGUSU (SADECE NET KALIPLAR)
        for k in TODAY_DAY_KEYS:
            if k in msg:
                print("🧠 [INTENT]: today_day")
                return "today_day"

        #  YÜCE AUTO – REPUTATION (HELP’TEN önce)
        if (
            not any(h in msg for h in EXPORT_HINTS)
            and not self.contains_department_word(msg)
            and not self.detect_work_type(user_message)   
            and not self.has_work_history_count_signal(msg)  
            and any(self.root_match(msg, root) for root in YUCE_AUTO_REPUTATION_ROOTS)
            and any(self.fuzzy_contains(msg, q) for q in YUCE_AUTO_REPUTATION_QUESTIONS)
        ):
            return "yuce_auto_reputation"
        
        #  PORTAL LINK INTENT (SADECE NET SAYFA / LİNK İSTEĞİ)
        if (
            any(h in msg for h in PORTAL_LINK_STRONG_HINTS)
            and (
                any(v in msg for v in PORTAL_LINK_VERBS)
                or "link" in msg
                or "sayfa" in msg
            )
            #  BİLGİ SORULARINI ELEYELİM
            and not any(k in msg for k in [
                "kaç", "kac", "sayısı", "sayisi","departmnlar",
                "nelerdir", "neler", "çeşitleri",
                "departman", "departmanlar"
            ])
        ):
            return "portal_link"
        

        
        #şarj rezervasyon için
        norm = self.normalize_people_text(user_message)

        if (
            self.fuzzy_match_in_text(norm, CHARGE_ROOTS, 0.75)
            and self.fuzzy_match_in_text(norm, CHARGE_ACTIONS, 0.75)
        ):
            return "charge_reservation"
        
        #  CURRENCY CALCULATION (currency_rate'ten önce!)
        if self.detect_currency_calculation(user_message):
            print("🧠 [INTENT]: currency_calculation")
            return "currency_calculation"
        
        
        #  DÖVİZ KURU INTENT (HELP'ten sonra olmalı)
        if any(k in msg for k in CURRENCY_KEYWORDS):
            print("🧠 [INTENT]: currency_rate")
            return "currency_rate"

        #  GENEL SOHBET / MUHABBET (PERSON OLMASIN)
        GENERAL_CHAT_KEYS = [
            "onerin var mi",
            "oneri var mi",
            "tavsiyen var mi",
            "bir fikrin var mi",
            "ne dersin",
            "sence",
            "ne dusunuyorsun"
        ]

        if (
            any(k in msg for k in GENERAL_CHAT_KEYS)
            and not any(c in msg for c in YUCE_AUTO_REPUTATION_ROOTS)
        ):
            print("🧠 [INTENT]: smalltalk (general)")
            return None
        
        #  YÜCE PORTAL BİLGİ INTENT

        PORTAL_KEYWORDS = [
            "yuceportal", "yüceportal",
            "yuce portal", "yüce portal",
            "portal"
        ]

        PORTAL_INFO_TRIGGERS = [
            "nedir","ne","nedır","ne için var"
            "ne ise yarar","nedir","ne","nedır","ne için var",
            "ne ise yarar","ne","ndr",
            "ne işe yarar",
            "neler yapilir",
            "neler yapılir",
            "neler yapilir",
            "ne yapilir",
            "ne yapilir",
            "ne var",
            "bilgi"
        ]
        if (
            not self.contains_department_word(msg)   #
            and any(p in msg for p in PORTAL_KEYWORDS)
            and any(t in msg for t in PORTAL_INFO_TRIGGERS)
        ):
            return "portal_info"

        #  OTOPARK CONTEXT FLAG  
        has_park = any(k in msg for k in OTOPARK_KEY)
        has_availability = any(k in msg for k in PARK_AVAILABILITY_KEYWORDS)
        has_user_status = any(k in msg for k in USER_PARK_STATUS_KEYS)


        if has_availability and has_date and not has_park:
            print("🧠 [INTENT]: otopark_status (implicit)")
            return "otopark_status"
        
        if (
            has_park
            and has_availability
            and not has_user_status
        ):
            print("🧠 [INTENT]: otopark_status (availability)")
            return "otopark_status"
        
        #  OTOPARK INTENT (TEK VE NET BLOK)
        if has_park and any(k in msg for k in CANCEL_KEYS):
            print("🧠 [INTENT]: otopark_cancel")
            return "otopark_cancel"
        
        if (
            has_park
            and any(k in msg for k in USER_PARK_STATUS_KEYS)
            and not any(k in msg for k in PARK_CREATE_ACTION_VERBS)
        ):
            print("🧠 [INTENT]: otopark_status_user")
            return "otopark_status_user"

        
        #  CREATE otopark
        if (
            has_park
            and any(k in msg for k in PARK_CREATE_ACTION_VERBS)
            and not any(k in msg for k in USER_PARK_STATUS_KEYS)
        ):
            print("🧠 [INTENT]: otopark_create")
            return "otopark_create"

        #  WEATHER INTENT 
        if (
            (
                any(k in msg for k in WEATHER_KEYWORDS)
                or self.fuzzy_match_in_text(msg, WEATHER_KEYWORDS, 0.80)
            )
            and not any(k in msg for k in OTOPARK_KEY)
            and not self.detect_work_type(user_message)
        ):
            print("🧠 [INTENT]: weather")
            return "weather"
        
        HELP_KEYS_NORM = [self.normalize_people_text(h) for h in HELP_ACTION_KEYS]

        has_help_keyword = any(h in msg for h in HELP_KEYS_NORM)

        has_module_signal = any([
            any(k in msg for k in OTOPARK_KEY),
            "takvim" in msg,
            self.detect_work_type(user_message),
            self.extract_department_name(user_message),
            self.has_people_field_request(user_message),
        ])

        if has_help_keyword and has_module_signal:
            return "help"
        
        has_department = self.extract_department_name(user_message) is not None

        def _has_cancel_intent_local(msg: str) -> bool:
            msg = self.normalize_people_text(msg)
            for w in CANCEL_KEYS:
                # 🔒 SADECE TAM KELİME
                if re.search(rf"\b{re.escape(w)}\b", msg):
                    return True
            return False
        
        has_cancel =_has_cancel_intent_local(msg)

        has_any_context = any(k in msg for k in (
            OTOPARK_KEY
            + ["takvim", "haftalik", "haftalık"]
        ))

        if has_department :
            has_cancel = False

        if has_cancel and not has_any_context and not has_date:
            print("🧠 [INTENT]: cancel_without_context")
            return "cancel_without_context"
        
        # FALLBACK: OTOPARK GPT (SADECE HİÇBİR RULE TUTMADIYSA)

        if any(p in msg for p in ["park", "otopark", "rezervasyon","park yeri","prk","otoprk","rez","rezervsyn","yer ayır"]):
            print("🟡 [INTENT]: fallback → GPT otopark intent")

            prompt = f"""
            Aşağıdaki kullanıcı mesajından otopark niyetini tespit et.

            Kurallar:
            - Kullanıcı KENDİ rezervasyonunu soruyorsa → otopark_status_user
            - Genel müsaitlik soruyorsa → otopark_status
            - Ayırma / rezerve → otopark_create
            - İptal → otopark_cancel
            - Emin değilsen → none

            Kullanıcı mesajı:
            "{user_message}"

            Sadece şu seçeneklerden biriyle cevap ver:
            - otopark_create
            - otopark_cancel
            - otopark_status
            - otopark_status_user
            - none
            """

            try:
                resp = self.client.chat.completions.create(
                    model="gpt-4.1-mini",
                    messages=[{"role": "system", "content": prompt}],
                    max_tokens=5,
                    temperature=0
                )
                intent = resp.choices[0].message.content.strip().lower()
                print("🧠 [GPT FALLBACK INTENT]:", intent)
                return intent
            except Exception as e:
                print("❌ GPT fallback hata:", e)

    def is_company_only(self, text: str) -> bool:
        text = self.normalize_people_text(text)
        COMPANY_ROOTS = [
            "yuce auto",
            "yuce oto",
            "skoda",
            "škoda"
        ]
        if any(k in text for k in COMPANY_ROOTS) and not any(p in text for p in PERSONEL_KEY):
            return True

        return False

    def route_company_reputation(self, text: str):
        text_norm = self.normalize_people_text(text)

        # YÜCE AUTO
        if any(k in text_norm for k in [
            "yuce auto","yuce oto","yuce"
        ]):
            return self.handle_yuce_auto_reputation()

        # SKODA
        if any(k in text_norm for k in [
            "skoda","škoda"
        ]):
            return self.handle_skoda_reputation()

        return None
    def has_person_in_db(self, text: str) -> bool:
        text_norm = self.normalize_people_text(text)
        BLOCKED_SURNAME_ONLY = {"yuce"}
        if (
            any(t in text_norm for t in YUCE_AUTO_REPUTATION_ROOTS)
            and any(q in text_norm for q in YUCE_AUTO_REPUTATION_QUESTIONS)
        ):
            return False
        
        tokens = text_norm.split()

        conn = _mysql_conn()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT LOWER(name) AS name, LOWER(surname) AS surname
            FROM users1
            WHERE is_active = 1
        """)
        users = cursor.fetchall()
        cursor.close()
        conn.close()

        # 1️⃣ AD + SOYAD (en net durum)
        for u in users:
            full = f"{u['name']} {u['surname']}"
            if full in text_norm:
                return True

        #  TEK İSİM (DB’de GERÇEKTEN unique mi?)
        name_hits = [u for u in users if u["name"] in tokens]

        if len(name_hits) == 1:
            return True          

        if len(name_hits) > 1:
            return False         

        # 3 TEK SOYİSİM (DB’de GERÇEKTEN unique mi?)
        surname_hits = [u for u in users if u["surname"] in tokens]

        if any(s in BLOCKED_SURNAME_ONLY for s in tokens):
            return False

        if len(surname_hits) == 1:
            return True

        if len(surname_hits) > 1:
            return False

        return False

    def looks_like_company_people(self, msg: str) -> bool:

        if not msg:
            return False
        
        
        text = self.normalize_people_text(msg)
        
        #  asistan / smalltalk (hard stop)
        if text in {
            "naber", "naberr", "naberrr",
            "selam", "selamlar",
            "merhaba", "mrb", "slm", "hey", "hi", "hello",
            "nasilsin", "nasılsın", "nasil", "nası",
            "iyi misin", "ne haber", "ne var ne yok",
            "napıyorsun", "napiyorsun", "ne yapıyorsun",
             "nbr"
        }:
            return False

        #diğer modüller (hard stop)
        blacklist = [
            "otopark", "park", "rezervasyon", "takvim", "yemek", "menu",
            "onerin", "oneri", "tavsiye", "fikrin","hakkında"
        ]        
        if any(b in text for b in blacklist):
            return False

        #  DB'DEN İSİM BULUNUYORSA → DİREKT GİR
        try:
            persons = self.has_person_in_db(msg)
            if persons:
                return True
        except Exception:
            pass

        # Alan kelimeleri (RAW_FIELD_KEYWORDS'tan)
        field_keywords = []
        for kws in RAW_FIELD_KEYWORDS.values():
            field_keywords.extend(kws)

        field_keywords = [self.normalize_people_text(k) for k in field_keywords]
        has_field = any(k in text for k in field_keywords)

        # Kişi niyeti (çok sınırlı, kontrollü)
        person_intent_keywords = [
            "kim", "kimin", "çalışan", "personel", "kişi","kişiler","çalışanlar","kimler","kişi", "kişiler","çalışan",   
            "ekip", "ekibi"
        ]
        has_person_intent = any(p in text for p in person_intent_keywords)

        if has_field and has_person_intent:
            return True

        #  HİÇBİRİ YOKSA → GİRME

        return False


    #akıcı şekilde akıllı gözükmesi için başına ve sonuna metin ekleme
    def naturalize_text(self, base_message: str, category: str) -> str:
        """
        Sabit backend mesajlarını daha doğal, çeşitli ve insan gibi hale getirir.
        AI küçük bir ek cümle/ek ifade üretir.
        """
        try:
            prompt = f"""
            Aşağıdaki mesaj {category} kategorisinde bir sistem yanıtıdır.
            Görevin:
            - Anlamı BOZMAMAK.
            - Backend tarafından üretilen mesajı asla değiştirme.
            - Sadece sonuna 1 kısa doğal, samimi, profesyonel cümle ekle.
            - Emoji çok abartma (maks 1 kullan).
            - Cümle Türkçe olmalı.

            Mesaj:
            "{base_message}"

            Sadece geliştirilmiş nihai metni döndür.
            """
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role":"system", "content":prompt}],
                max_tokens=50,
                temperature=0.6
            )
            return resp.choices[0].message.content.strip()
        except:
            return base_message

    def naturalize_intro(self, category: str) -> str:
        """
        Yugii'nin çeşitli modüllerde kullanıcıya verdiği
        kısa, samimi, profesyonel ve işlem belirtmeyen intro cümlesi.
        """
        try:
            prompt = f"""
            Sen Yüce Auto'nun dijital asistanı Yugii'sin.
            Kullanıcıya işlemden ÖNCE gösterilecek çok kısa bir karşılama cümlesi üret.

            Kurallar:
            - En fazla 1 cümle.
            - İşlem belirtme: "kontrol ediyorum", "bakıyorum", "sorguluyorum", "inceleyim", "hallediyorum" gibi fiilleri ASLA kullanma.
            - Cümle sadece karşılayıcı ve olumlu olsun.
            - Profesyonel + samimi bir ton.
            - Emoji 0 veya 1 tane olabilir.
            - Her üretimde farklı ve çeşitli olsun.
            - Asistanın işlevini açıklama, sadece sıcak bir karşılayış olsun.

            Kategori: {category}

            Örnek doğru format kısa cümleler:
            - “Tabii ki 😊”
            - “Elbette, memnuniyetle.”
            - “Buradayım, yardımcı olmaya hazırım 🌿”
            - “Tamamdır 🙂”
            - “Memnuniyetle 💚”

            Sadece cümleyi döndür.
            """

            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=25,
                temperature=0.8
            )
            return resp.choices[0].message.content.strip()

        except:
            return "Tabii ki 😊"
    
    #yemek menüsü cevapları başına doğan cümle ekleme
    def naturalize_food_intro(self):
        try:
            prompt = """
            Yemek menüsü isteği için kullanıcıya gösterilecek
            çok kısa bir karşılama cümlesi üret.

            Kurallar:
            - İşlem belirtme YOK (bakıyorum, kontrol ediyorum vs. yasak).
            - En fazla 1 cümle.
            - 0–1 emoji kullanılabilir.
            - Samimi + profesyonel ton.
            - Her seferinde çeşitli olsun.

            Örnekler:
            - "Elbette 😊"
            - "Seve seve 🌿"
            - "Memnuniyetle 💚"
            - "Tabii ki 😊"
            """
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=20,
                temperature=0.9
            )
            return resp.choices[0].message.content.strip()
        except:
            return "Elbette 😊"

    #Yemek menüsü mesajının SONUNA doğal bir kapanış cümlesi eklemek
    def naturalize_food_outro(self):
        try:
            prompt = """
            Yemek menüsü mesajının sonuna eklenecek,
            kısa, samimi ve yemek temasına uygun bir kapanış cümlesi üret.

            Kurallar:
            - Tek cümle.
            - 0–1 emoji.
            - Afiyet temalı olabilir ama aşırı değil.

            Örnek:
            - "Afiyetle keyifli bir gün dilerim 😊"
            - "Şimdiden afiyet olsun 🌿"
            - "Umarım bugün menü tam damak tadına göredir 💚"
            """
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=20,
                temperature=0.8
            )
            return resp.choices[0].message.content.strip()
        except:
            return "Afiyet olsun 😊"
     
    # çalışma şekli belirleme haftalık takvim
    def _norm(self, s: str) -> str:
        return (
            s.lower()
            .replace("ı", "i")
            .replace("ş", "s")
            .replace("ğ", "g")
            .replace("ü", "u")
            .replace("ö", "o")
            .replace("ç", "c")
        )

    def _strip_tr_suffixes(self, text: str) -> str:
        """
        Türkçe fiil ve isim çekim eklerini kaba şekilde temizler.
        Amaç: geldiğim → geldim → gel
        """
        SUFFIXES = [
            "digim", "tigim",
            "dim", "tim", "dum", "tum",
            "mis", "mus", "mıs", "müş",
            "yor", "iyor",
            "deki", "daki", "teki", "taki",
            "ken", "iken",
            "im", "um", "üm",
            "mi", "mı", "mu", "mü",
        ]

        for suf in SUFFIXES:
            text = re.sub(rf"\b(\w+){suf}\b", r"\1", text)

        return text
    
    def _detect_by_root(self, text: str):

        if not any(x in text for x in ["calis", "ofis", "home", "izin", "bayi"]):
            return None
        
        ROOT_PATTERNS = [
            ("Izin",  [r"\b(izin|tatil|rapor|off)\w*"]),
            ("Home",  [r"\b(ev|home|uzaktan|remote)\w*"]),
            ("Ofis",  [r"\b(ofis|şirket|sirket)\w*"]),
            ("Bayi",  [r"\b(bayi|müşteri|musteri|saha|ziyaret)\w*"]),
        ]

        for work_type, patterns in ROOT_PATTERNS:
            for p in patterns:
                if re.search(p, text, re.IGNORECASE):
                    return work_type

        return None
       
    def detect_work_type(self, msg: str):
            """
            Kullanıcının yazdığı cümleden çalışma tipini tespit eder.
            Ultra dayanıklı, Türkçe karakter uyumlu, tüm varyasyonları kapsar.
            """

            if not msg:
                return None

            # METNİ HAZIRLA (ZORUNLU NORMALIZE)
            text = self._norm(msg)
            text = self._strip_tr_suffixes(text)
            text = re.sub(r"\s+", " ", text).strip()

            ignore_words = {"komple", "boyunca", "tamami", "tamamı","full","tamammına","tamamı","kmple"}
            tokens = text.split()
            tokens = [t for t in tokens if t not in ignore_words]
            text = " ".join(tokens)

            # ⭐ 1) ROOT + REGEX (ANA YOL)
            root_type = self._detect_by_root(text)
            if root_type:
                return root_type
            
            # 2) TÜM ANAhtar kelime listeleri

            # departman / şirket bağlamında work_type çıkmasın

            if any(k in msg for k in ["departman", "departmanlar","şirketteki departmanlar","ofisteki departmanlar","ofiste olan departmanlar"]):
                return None
            izin_words = [
                "izin", "izinli", "izinliyim", "izne", "izindeyim", "izinde",
                "tatil", "tatilde", "tatile", "yıllık izin", "yillik izin",
                "off", "offday", "off-day", "off day", "off gün", "offgun",
                "izinliydim", "izindeydim", "izindeyken", "izin yaptım",
                "izindeyim", "izindeydim","izin almısım","izinlydım","izin isaretlemisim"
            ]

            home_words = [
                # temel
                "home", "homeoffice", "home office",
                "ev", "evde", "evden",
                "evdi", "evdeydi", "evdiydi", "evdeydim", "evdiler", "evdeydiler",
                # çalışma fiilleri
                "evden calis", "evden çalış",
                "evde calis", "evde çalış",

                # uzaktan
                "uzaktan", "remote", "remotely", "wfh",

                # konuşma dili
                "evden baglan", "evden bağlan",
                "evden calisiyorum", "evden çalışıyorum",

                # yazım hataları
                "evdn", "hom"
            ]

            ofis_words = [
                # =====================
                # TEMEL OFİS
                "ofis", "ofiste", "ofise", "ofisten",
                "ofisde", "ofisden",

                # =====================
                # ŞİRKET (ANA)
                # =====================
                "şirket", "şirkete", "şirkette", "şirketten",
                "sirket", "sirkete", "sirkette", "sirketten",

                # =====================
                # İŞ YERİ / BİNA
                # =====================
                "iş yeri", "isyeri", "işyerinde",
                "iş yerinde", "iş yerinden",
                "is yeri", "is yerinde", "is yerinden",
                "merkez", "bina",

                # =====================
                # KONUŞMA DİLİ
                # =====================
                "ofise gel", "ofise geliyorum",
                "ofise geleceğim", "ofise gelecegim",
                "şirkete gel", "şirkete geliyorum",
                "şirkete geleceğim", "şirkete gelecegim","ise gel", "ise geldim", "ise geliyorum",
                "ise git", "ise gittim", "ise gidiyorum",
                "ise var", "ise vardim", "ise vardım",

                # =====================
                # YAYGIN YAZIM HATALARI
                # =====================
                "şirktette", "sirkette","şirketten","şirketde","şirktte","sirketden","şrktten"
            ]

            bayi_words = [
                "bayi", "bayide", "bayiye","bayii",
                "müşteri", "musteri", "ziyaret", "ziyarete", "ziyarette",
                "sahada", "saha", "aktif satış", "sales visit"
            ]


            # -----------------------------
            # 3) DİREKT KELİME İÇERİSİNDE ARA
            # -----------------------------
            for w in izin_words:
                if w in text:
                    return "Izin"

            for w in home_words:
                if w in text:
                    return "Home"

            for w in ofis_words:
                if w in text:
                    return "Ofis"

            for w in bayi_words:
                if w in text:
                    return "Bayi"

            # -----------------------------
            # 4) TOKENS & FUZZY MATCH (%80 üzeri)
            # -----------------------------

            tokens = text.split()

            def fuzzy_match(token, keywords):
                for w in keywords:
                    ratio = difflib.SequenceMatcher(None, token, w).ratio()
                    if ratio >= 0.80:   # güçlü eşleşme
                        return True
                return False

            for t in tokens:
                if fuzzy_match(t, izin_words):
                    return "Izin"
                if fuzzy_match(t, home_words):
                    return "Home"
                if fuzzy_match(t, ofis_words):
                    return "Ofis"
                if fuzzy_match(t, bayi_words):
                    return "Bayi"

            # -----------------------------
            # 5) SEMANTIC HINTS
            # -----------------------------
            if any(h in text for h in ["çalışmıyorum","çalışmadım","rahatsız", "hasta", "yokum", "istirahat","izinli","izin"]):
                return "Izin"

            if re.search(r"\bev(de|den)?\b", text):
                return "Home"

            if any(h in text for h in ["geliyorum", "geleceğim", "ofiste olacağım","geliyor","gelicek","gelecek","gelecek ofise",
            "ofise geliyor", "ofise gelicek","ofise gelecek"
            ]):
                return "Ofis"

            if any(h in text for h in ["bayi", "sahada", "ziyaret","bayide"]):
                return "Bayi"

            # -----------------------------
            # 6) BULUNAMADI
            # -----------------------------
            return None
    
    #Cümleyi “çalışma tipi”ne göre parçalara ayırır
    def extract_worktype_segments(self, text: str):
        """
        Metni work-type bazlı parçalara ayırır.
        Work type TESPİT ETMEZ, detect_work_type kullanır.
        """

        if not text:
            return []

        # normalize + suffix kırma (AYNI AKIŞ)
        t = self._norm(text)
        t = self._strip_tr_suffixes(t)

        tokens = t.split()

        segments = []
        current_tokens = []
        current_type = None

        for i, tok in enumerate(tokens):

            # 🔑 2 kelimelik pencere
            window2 = " ".join(tokens[i:i+2])

            # 🔑 BU HAFTA anchor
            if any(k in window2 for k in BU_WEEK_DAY_PREFIXES):
                current_tokens.append("bu hafta")
                continue

            # 🔑 HAFTAYA / GELECEK HAFTA / ÖNÜMÜZDEKİ HAFTA
            if any(k in window2 for k in NEXT_WEEK_KEYS):
                current_tokens.append("haftaya")
                continue

            # 🔥 VE bağlacı özel kontrolü
            if tok == "ve":
                next_window = " ".join(tokens[i+1:i+3])
                next_type = self.detect_work_type(next_window)

                if next_type and current_tokens and current_type:
                    segments.append({
                        "text": " ".join(current_tokens),
                        "type": current_type
                    })
                    current_tokens = []
                    current_type = None
                continue

            # 🔍 çalışma tipi algılama
            window3 = " ".join(tokens[max(0, i-2):i+1])
            wt = self.detect_work_type(window3)

            if wt:
                if current_tokens and current_type:
                    segments.append({
                        "text": " ".join(current_tokens),
                        "type": current_type
                    })
                    current_tokens = []
                current_type = wt
            else:
                current_tokens.append(tok)

        # son segment
        if current_tokens and current_type:
            segments.append({
                "text": " ".join(current_tokens),
                "type": current_type
            })

        return segments
    
    def is_plate_add_question(self, msg: str) -> bool:
        text = self.normalize_people_text(msg)

        PLATE_KEYWORDS = [
            "plaka ekle",
            "arac plakasi ekle",
            "araç plakası ekle",
            "plaka nasil eklenir",
            "plaka nasil girilir",
            "plakam yok",
            "plaka guncelle",
            "plaka degistir",
            "arac plakasi",
        ]

        return any(k in text for k in PLATE_KEYWORDS)
  
 

    #control_work_history tarih belirleme
    def extract_history_dates(self, message: str):
        """
        Gelişmiş tarih algılama motoru (Geçmiş + Bugün + Gelecek).
        Amaç:
        - Kullanıcının yazdığı doğal cümleden NET bir tarih aralığı çıkarmak.
        - Hem geçmiş hem gelecek ifadelerini anlamak.
        - Limitler:
                Geçmiş: max 365 gün
                Gelecek: max 30 gün
        RETURN:
            { start: date, end: date, mode: "past|future|mixed", source: "gpt|rule" }
        """

        import json
        import regex as re
        import calendar
        from dateutil.relativedelta import relativedelta


        today = datetime.now().date()
        msg = message.lower().strip()

        msg = self.normalize_month_with_suffixes(msg)

        # === AY HARİTASI (TEK KAYNAK) ===
        month_map = {
            "ocak":1, "şubat":2, "mart":3, "nisan":4, "mayıs":5, "haziran":6,
            "temmuz":7, "ağustos":8, "eylül":9, "ekim":10, "kasım":11, "aralık":12
        }

        # TÜM YIL / YIL BOYUNCA
        if any(k in msg for k in ["tum yil", "tüm yil", "yil boyunca", "butun yil"]):
            return {
                "start": date(today.year, 1, 1),
                "end": date(today.year, 12, 31),
                "mode": "mixed",
                "source": "rule"
            }

        # BU YIL
        if "bu yil" in msg or "bu yıl" in msg:
            return {
                "start": date(today.year, 1, 1),
                "end": today,
                "mode": "past",
                "source": "rule"
            }

        # BU AY
        if "bu ay" in msg:
            start = date(today.year, today.month, 1)
            end = today
            return {
                "start": start,
                "end": end,
                "mode": "past",
                "source": "rule"
            }
        # SON X AY (son 3 ay, son 6 ay)
        m = re.search(r"son\s+(\d+)\s+ay", msg)
        if m:
            months = int(m.group(1))
            end = today
            start = today - relativedelta(months=months)
            return {
                "start": start,
                "end": end,
                "mode": "past",
                "source": "rule"
            }
        # ============================================================
        # 🔥 ÇOKLU AY ALGILAMA (ocak ve şubat / kasım, aralık)
        # ============================================================
        found_months = [
            (name, num)
            for name, num in month_map.items()
            if name in msg
        ]

        if len(found_months) >= 2:
            months = [m[1] for m in found_months]
            min_month = min(months)
            max_month = max(months)

            # yıl hesabı (aralık → ocak gibi durumlar)
            start_year = today.year
            end_year = today.year

            if max_month < min_month:
                # yıl geçişi var (aralık-ocak)
                start_year = today.year - 1
                end_year = today.year

            start = datetime(start_year, min_month, 1).date()
            end = datetime(
                end_year,
                max_month,
                calendar.monthrange(end_year, max_month)[1]
            ).date()

            return {
                "start": start,
                "end": end,
                "mode": "future" if start > today else "past" if end < today else "mixed",
                "source": "rule"
            }


        # === YIL HESAPLAMA YARDIMCI FONKSİYONU ===
        def resolve_year_for_month(month_num, today):
            # Eğer ay bugünün ayından büyükse → geçen yıl
            if month_num > today.month:
                return today.year - 1
            # Aksi halde → bu yıl
            return today.year
        
        for name, num in month_map.items():
            if name in msg:
                year = resolve_year_for_month(num, today)
                start = datetime(year, num, 1).date()
                end = datetime(year, num, calendar.monthrange(year, num)[1]).date()
                mode = "future" if start > today else "past"
                return {
                    "start": start,
                    "end": end,
                    "mode": mode,
                    "source": "rule"
                }
        # RULE FALLBACK — Her ihtimale karşı
        for key, triggers in RELATIVE_DAY_TRIGGERS.items():
            if any(t in msg for t in triggers):

                if key == "today":
                    return {
                        "start": today,
                        "end": today,
                        "mode": "past",
                        "source": "rule"
                    }

                if key == "tomorrow":
                    d = today + timedelta(days=1)
                    return {
                        "start": d,
                        "end": d,
                        "mode": "future",
                        "source": "rule"
                    }

                if key == "yesterday":
                    d = today - timedelta(days=1)
                    return {
                        "start": d,
                        "end": d,
                        "mode": "past",
                        "source": "rule"
                    }

        # === BU HAFTA ===
        if any(k in msg for k in THIS_WEEK_KEYS):
            monday = today - timedelta(days=today.weekday())
            friday = monday + timedelta(days=4)
            return {"start": monday,"end": friday,"mode": "mixed","source": "rule" }

        # === GEÇEN HAFTA ===
        if any(k in msg for k in PAST_WEEK_KEYS):
            monday = today - timedelta(days=today.weekday() + 7)
            return {"start": monday, "end": monday + timedelta(days=4), "mode": "past", "source": "rule"}

        # === HAFTAYA ===
        norm = self.normalize_calendar_text(msg)
        if any(k in norm for k in NEXT_WEEK_KEYS):
            monday = today + timedelta(days=(7 - today.weekday()))
            return {"start": monday, "end": monday + timedelta(days=4), "mode": "future", "source": "rule"}

        # === GEÇEN AY ===
        if "geçen ay" in msg or "gecen ay" in msg:
            last_month = today.month - 1 if today.month > 1 else 12
            year = today.year if today.month > 1 else today.year - 1
            last_day = calendar.monthrange(year, last_month)[1]
            return {
                "start": datetime(year, last_month, 1).date(),
                "end": datetime(year, last_month, last_day).date(),
                "mode": "past",
                "source": "rule"
            }
        # === SAYI + AY (tek gün) ===
        match = re.search(r"(\d{1,2})\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)", msg)
        if match:
            day = int(match.group(1))
            ay_map = {
                "ocak": 1, "şubat":2, "mart":3,"nisan":4,"mayıs":5,"haziran":6,
                "temmuz":7,"ağustos":8,"eylül":9,"ekim":10,"kasım":11,"aralık":12
            }
            month = ay_map[match.group(2)]
            year = resolve_year_for_month(month, today)
            d = datetime(year, month, day).date()
            return {
                "start": d,
                "end": d,
                "mode": "future" if d > today else "past",
                "source": "rule"
            }
    
        # === AY + SAYI (TERS SIRA) → ocak 15 ===
        reverse_match = re.search(
            r"\b(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)\s+(\d{1,2})\b",
            msg
        )
        if reverse_match:
            ay = reverse_match.group(1)
            gun = int(reverse_match.group(2))

            ay_num = month_map[ay]
            year = resolve_year_for_month(ay_num, today)

            d = date(year, ay_num, gun)
            return {
                "start": d,
                "end": d,
                "mode": "future" if d > today else "past",
                "source": "rule"
            }
        # === SAYI SAYI SAYI + AY (ör: 11 12 13 kasım) ===
        multi = re.findall(r"(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+(kasım|ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|aralık)", msg)
        if multi:
            g1,g2,g3,ay = multi[0]
            numbers = sorted([int(g1),int(g2),int(g3)])
            ay_map = {
                "ocak":1,"şubat":2,"mart":3,"nisan":4,"mayıs":5,"haziran":6,
                "temmuz":7,"ağustos":8,"eylül":9,"ekim":10,"kasım":11,"aralık":12
            }
            m = ay_map[ay]

            year = resolve_year_for_month(m, today)
            d1 = datetime(year, m, numbers[0]).date()
            d2 = datetime(year, m, numbers[2]).date()
            return {
                "start": d1,
                "end": d2,
                "mode": "future" if d1 > today else "past",
                "source": "rule"
            }

        #  GPT  cümle oluşturma — Hem geçmiş hem gelecek için
        gpt_prompt = f"""
        Kullanıcı çalışma günlerini incelemek istiyor.
        Mesajdan NET bir tarih aralığı çıkarmalısın.

        Sadece ŞU JSON formatında cevap ver:

        {{
            "start": "YYYY-MM-DD",
            "end":   "YYYY-MM-DD"
        }}

        Kurallar:
        - Eğer kullanıcı tek gün soruyorsa start=end yap.
        - Aşağıdaki ifadeleri anlamalısın:
            * bu ay, geçen ay, gelecek ay
            * bu hafta, geçen hafta, haftaya, önümüzdeki hafta
            * 3 gün sonra, 5 gün önce, dün, yarın, bugün
            * 11 kasım, 11-15 kasım, 11 12 13 kasım
            * ayın başı, ayın ortası, ayın sonu
            * geçen sene kasım değil → sadece son 12 ay
        - Eğer tarih çıkaramazsan:
            {{"start": null, "end": null}}
        """
        try:
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {"role": "system", "content": gpt_prompt},
                    {"role": "user", "content": message}
                ],
                max_tokens=120,
            )

            raw = resp.choices[0].message.content.strip()
            gpt_json = json.loads(raw)

            if gpt_json.get("start") and gpt_json.get("end"):
                start = datetime.fromisoformat(gpt_json["start"]).date()
                end = datetime.fromisoformat(gpt_json["end"]).date()

                # GPT üretti ama start>end olabilir → düzelt
                if start > end:
                    start, end = end, start

                # === GPT TARİH YIL AUTO-DÜZELTME ===
                def fix_year_auto(d):
                    # Eğer ay bugünden büyükse → geçen yıl
                    if d.month > today.month:
                        return d.replace(year=today.year - 1)
                    # Aksi halde → bu yıl
                    return d.replace(year=today.year)

                start = fix_year_auto(start)
                end = fix_year_auto(end)

                return {
                    "start": start,
                    "end": end,
                    "mode": "future" if start > today else "past" if end < today else "mixed",
                    "source": "gpt"
                }


        except Exception as e:
            print("GPT tarih hatası:", e)
        #  HIÇBIR ŞEY ÇÖZÜLEMEZSE
        return {"start": None, "end": None, "mode": "unknown", "source": "none"}

    
    def detect_link_intent(self, user_message: str) -> bool:
        """
        Kullanıcı mesajının bir portal sayfa linki isteme niyetinde olup olmadığını GPT ile tespit eder.
        """
        #  OTOPARK VARSA LINK INTENT ÇALIŞMASIN
        if any(k in self.normalize_people_text(user_message) for k in OTOPARK_KEY):
            return False
        
        prompt = f"""
    Sen Yüce Auto'nun şirket içi portalı olan YücePortal için geliştirilmiş bir dijital asistansın.

    Görevin: Kullanıcının bir portal sayfasını, raporu, modülü veya bağlantıyı arayıp aramadığını tespit etmektir.

    YücePortal’ın içerdiği modüller ve sayfalar (örnekler):
    - Otopark Rezervasyonu 
    - Şarj İstasyonu Rezervasyon
    - Sabah Raporu / Akşam Raporu / Öngörü Raporu
    - Organizasyon Şeması
    - Kafeterya Siparişleri
    - Fiktif Araç Bildirimi / Filo Fiktif Araç Bildirimi
    - Dashboard / Sayaç / İstatistik Paneli
    - Satın Alma
    - Kurgu Excel
    - Jato Verileri
    - PDF Karşılaştırma
    - YS Faaliyet Raporu
    - Ana Sayfa
    - Acil Durum Kat Planları
    - SSH sayfası

    Bu sayfalara yönlendirme isteyen kullanıcılar genellikle şu kelime veya kalıpları kullanır:
    - "sayfa", "sayfası", "panel", "modül", "ekran", "raporu", "rapor"
    - "nerede", "nerde", "nereye", "nasıl açılır", "nasıl giderim"
    - "link", "bağlantı", "linkini", "linki", "bağlantısını"
    - "aç", "göster", "götür", "bul", "ulaş", "getir"
    - "menüsünü", "menü nerede","nerde","nerede"

    AYRICA:
    Kullanıcı yanlış yazabilir. Örneğin:
    - “shh sayfası”
    - “otoprak raporo”
    - “org sçeması”
    - “fiktiv araş”
    - “ak’şam rapooru”
    - “şarj istasonu”

    Bu hatalı yazımları da bir sayfa arama niyeti olarak kabul etmelisin.

    Kullanıcı mesajı:
    "{user_message}"

    SORU:
    Bu mesaj bir portal sayfası / modülü / raporu / bağlantıyı arama, bulma veya açma amacı taşıyor mu?

    Sadece:
    EVET
    veya
    HAYIR
    diye yanıt ver.
    """

        try:
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=5
            )

            raw = resp.choices[0].message.content
            normalized = raw.strip().lower()

            print("🟦 [INTENT DEBUG] RAW:", raw)
            print("🟦 [INTENT DEBUG] NORMALIZED:", normalized)

            return normalized.startswith("e")

        except Exception as e:
            print("Intent hata:", e)
            return False


    def detect_food_intent(self, user_message: str) -> bool:
        msg = user_message.lower()
        food_words = [
            "yemek", "menü", "menu", "yemekte", "bugün ne var", "ne yiyeceğiz",
            "ne var bugün", "yemek listesi","ne yemek","yemekde","yemkde","bugun ne yesek",
            "ne yesek","ne yiyecegiz","yemekte ne var","yemek var mi","menu var mi","bugunku menu",
            "bugun yemek","yarin yemek","cuma menu",
            "pazartesi yemek","yemek kac tl","menu fiyati","yemek fiyati",
            "ucret","kac para","menu degisti mi","bugun ne cikiyor","yemek cikiyor mu"
        ]
        return any(w in msg for w in food_words)


    def resolve_food_date(self, message):
        """
        Kullanıcının istediği günü gerçek tarihe çevirir.
        Sadece bugün + 2 gün (toplam 3 gün) izin verir.
        Daha ileri tarihleri None döndürerek reddettirir.
        """

        msg = message.lower()
        today = datetime.now().date()

        # 1) BUGÜN
        if any(trigger in msg for trigger in RELATIVE_DAY_TRIGGERS["today"]):
            return today

        # 2) YARIN
        if any(trigger in msg for trigger in RELATIVE_DAY_TRIGGERS["tomorrow"]):
            return today + timedelta(days=1)

        # 3) ÖBÜR GÜN / SONRAKİ GÜN
        if "öbür gün" in msg or "sonraki gün" in msg or "ertesi gün" in msg:
            return today + timedelta(days=2)

        # 4) HAFTANIN GÜNLERİ → gün adı geçiyorsa tarihe çevir
        weekday_map = {
            "pazartesi": 0,
            "salı": 1, "sali": 1,
            "çarşamba": 2, "carsamba": 2,
            "perşembe": 3, "persembe": 3,
            "cuma": 4,
            "cumartesi": 5,
            "pazar": 6
        }

        for word, weekday_number in weekday_map.items():
            if word in msg:
                today_weekday = today.weekday()
                diff = weekday_number - today_weekday
                if diff < 0:
                    diff += 7  # gelecek haftanın günü

                target_date = today + timedelta(days=diff)

                # ❗ 3 günlük sınırı (bugün + 2 gün)
                if (target_date - today).days <= 2:
                    return target_date
                else:
                    return None  # ileri tarih

        # Hiç gün belirtilmemişse → default olarak bugün
        return today

    def get_food_menu_for_date(self, target_date):
        try:
            conn = pyodbc.connect(YA_2El_AracSatis)
            cur = conn.cursor()
            cur.execute("""
                SELECT yemek1, yemek2, yemek3, yemek4, yemek5, yemek6,
                       yemek1_fiyat, yemek2_fiyat, yemek3_fiyat,
                       yemek4_fiyat, yemek5_fiyat, yemek6_fiyat,
                       notlar
                FROM [Yuce_Portal].[dbo].[Kafeterya_Menu]
                WHERE CAST(menu_date AS DATE) = CAST(? AS DATE)
            """, (target_date,))
            row = cur.fetchone()
            conn.close()
            return row
        except Exception as e:
            print("DB yemek hata:", e)
            return None
        
    def _format_page_response(self, row):
        return (
            f"🔗 <b>{row.page_name}</b> sayfasına buradan ulaşabilirsin:<br>"
            f"<a href='{row.url}' target='_blank'>{row.url}</a><br><br>"
            f"{row.description or ''}"
        )

    #  DB'DEN SAYFA EŞLEŞTİRME (SYNONYM + FUZZY + EMBEDDING)
    def find_page_link(self, user_message: str) -> str:
        """
        Kullanıcı sayfa / link / nerede dediğinde
        DB'den en uygun portal sayfasını bulur ve
        direkt kullanıcıya cevap döndürür.
        """

        print("🔍 [PORTAL LINK] user_message:", user_message)

        try:
            conn = pyodbc.connect(YA_2El_AracSatis)
            cur = conn.cursor()
            sql = """
                SELECT page_name, url, description, synonyms
                FROM [Yuce_Portal].[dbo].[yugii_page_links]
                WHERE active = 1
            """
            cur.execute(sql)
            self.auto_log_sql(sql, None)
            rows = cur.fetchall()
            conn.close()

            message = self.normalize_people_text(user_message)

            #  SYNONYM (EN GÜÇLÜ)
            best_score = 0
            best_row = None

            for r in rows:
                if not r.synonyms:
                    continue

                synonyms = [
                    s.strip().lower()
                    for s in r.synonyms.split(",")
                    if len(s.strip()) >= 3
                ]

                score = 0
                for s in synonyms:
                    if s in message:
                        score += 2
                    elif fuzz.partial_ratio(message, s) >= 85:
                        score += 1

                if score > best_score:
                    best_score = score
                    best_row = r

            if best_row:
                return self._format_page_response(best_row)

            #  FUZZY (PAGE NAME + SYNONYM)
            FUZZY_THRESHOLD = 70
            best_score = 0
            best_row = None

            for r in rows:
                candidates = []

                if r.page_name:
                    candidates.append(r.page_name.lower())

                if r.synonyms:
                    candidates.extend(
                        s.strip().lower()
                        for s in r.synonyms.split(",")
                        if len(s.strip()) >= 3
                    )

                for c in candidates:
                    score = fuzz.partial_ratio(message, c)
                    if score > best_score:
                        best_score = score
                        best_row = r

            if best_score >= FUZZY_THRESHOLD:
                return self._format_page_response(best_row)

            # EMBEDDING (SON ÇARE)
            if not hasattr(self, "_page_link_embedder"):
                self._page_link_embedder = SentenceTransformer(
                    "paraphrase-multilingual-mpnet-base-v2"
                )

            texts = []
            row_map = []

            for r in rows:
                base = r.page_name or ""
                if r.synonyms:
                    base += " " + r.synonyms.replace(",", " ")
                texts.append(base)
                row_map.append(r)

            msg_emb = self._page_link_embedder.encode(message, convert_to_tensor=True)
            page_embs = self._page_link_embedder.encode(texts, convert_to_tensor=True)

            sims = util.cos_sim(msg_emb, page_embs)[0]
            best_idx = int(sims.argmax())
            best_sim = float(sims[best_idx])

            if best_sim >= 0.75:
                return self._format_page_response(row_map[best_idx])

            return (
                "Aradığın sayfayı net bulamadım 🤔<br>"
                "Biraz daha açık yazar mısın? (örn: haftalık takvim, otopark, ssh)"
            )

        except Exception as e:
            print("❌ PORTAL LINK HATA:", e)
            return "Şu anda sayfa bilgilerine erişemiyorum."


    def get_db_otopark_help_text(self, user_message):


        help_pool = [

            # 1
            (
                "💚 Otopark rezervasyonu yapmak oldukça pratik!<br><br>"
                "🧭 <b>Portal üzerinden:</b><br>"
                "👉 <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>Otopark Rezervasyon Sayfası</a><br><br>"
                "💬 <b>Yugii ile:</b><br>"
                "• “Yarın 10 numaralı parkı ayır”<br>"
                "• “17 kasım için park ayır”<br>"
                "• “Bu hafta park ayır”<br>"
            ),

            # 2
            (
                "🅿️ Park ayırmak çok kolay!<br>"
                "İstersen <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>portal üzerinden</a> yapabilirsin,<br>"
                "ya da bana şöyle yazabilirsin:<br>"
                "• “Yarın park ayır”<br>"
                "• “18 kasım 12 numara park ayır” 💚"
            ),

            # 3
            (
                "🌿 Elbette! Otopark rezervasyonu için:<br>"
                "👉 Portal: <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>Otopark Sayfası</a><br>"
                "👉 Yugii: “Yarın 9 numaralı parkı ayır” yazman yeterli 🙂"
            ),

            # 4
            (
                "💚 Park rezervasyonu oluşturmak için iki yol var:<br>"
                "• <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>Portal ekranından</a> oluşturabilirsin yada <br>"
                "• Ya da bana kısaca 'yarın 8 numaralı parkı ayır' yazabilirsin 🌿"
            ),

            # 5
            (
                "🅿️ Park ayırma çok pratik!<br>"
                "👉 Portal: <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>yüceportal sayfasından yada buradan</a><br>"
                "👉 Yugii: '17 kasım park ayır' demen yeterli 😊"
            ),

            # 6
            (
                "🌿 Yardımcı olayım! Otopark rezervasyonu için:<br>"
                "• Ya da bana 'yarın 11 numarayı ayır' gibi bir mesaj yazabilirsin 💚"
                "• Portal bağlantısı: <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>Buraya tıkla</a><br>"

            ),

            # 7
            (
                "💚 Park rezervasyonu oluşturmak istiyorsan;<br>"
                "👉 Ya da bana 'yarın 10 numara park ayır' yazman yeterli 🙂"
                "👉 <a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve' target='_blank'>Portal sayfasını açarakda kolay bir şekilde</a> yapabilirsin.<br>"

            ),

        ]

        return random.choice(help_pool)
    
    def generate_calendar_help_ai(self, user_message):
        try:
            prompt = f"""
            Kullanıcı haftalık takvimi nasıl işaretleyeceğini soruyor.

            Sadece şu formatta çok kısa bir mesaj üret:
            1️⃣ Kısa bir giriş cümlesi (tek cümle, max 1 emoji).
            2️⃣ Takvimin mantığını TEK cümlede açıkla: “gün + çalışma tipi yazman yeterli”.
            3️⃣ Aynı satırda 2 örnek ver: “Salı home”, “Perşembe ofis”.
            4️⃣ Son satır TEK satır olmalı ve aynen şöyle başlamalı:
            Haftalık Takvim 👇 https://yuceportal.skoda.com.tr/haftalik_takvim

            Kurallar:
            - En fazla 2–3 satır üret.
            - Paragraf yok.
            - Uzun açıklama YOK.
            - Cümleler kısa ve sade olsun.
            - Blok yapısını BOZMA.
            - Kesinlikle 3 satırı geçme.

            Kullanıcı mesajı: "{user_message}"
            """
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=80,
                temperature=0.6
            )
            return response.choices[0].message.content.strip()

        except:
            return (
                "Gün + çalışma tipi yazman yeterli 😊 Örn: Salı home, Perşembe ofis<br>"
                "Haftalık Takvim 👇 https://yuceportal.skoda.com.tr/haftalik_takvim"
            )

    def get_calendar_help_text(self, user_message=None):

        # AI tamamen kapalı — sadece static mesajlar dönecek ilerde geliştirek için ai eklenecek
        # if random.random() < 0.7:
        #     return self.generate_calendar_help_ai(user_message)

        static_pool = [
            (
                "Takvime gün eklemek çok kolay 😊<br>"
                "Örnek: “Cuma ofis” — “Salı home ekle”<br>"
                "Haftalık Takvim 👇 "
                "<a href='https://yuceportal.skoda.com.tr/haftalik_takvim' target='_blank'>Haftalık Takvim</a>"
            ),
            (
                "Gün + çalışma tipi yazman yeterli 💚<br>"
                "Mesela: “Pazartesi ofis” veya “Çarşamba home işaretle”<br>"
                "Haftalık Takvim 👇 "
                "<a href='https://yuceportal.skoda.com.tr/haftalik_takvim' target='_blank'>Haftalık Takvim</a>"
            ),
            (
                "Elbette! Takvimde bir gün işaretlemek için kısa bir komut yazman yeterli 🙂<br>"
                "Örneğin: “Perşembe home” — “Yarın ofis ekle”<br>"
                "Haftalık Takvim 👇 "
                "<a href='https://yuceportal.skoda.com.tr/haftalik_takvim' target='_blank'>Haftalık Takvim</a>"
            )
        ]

        return random.choice(static_pool)

    #help parse modülü için 
    def help_user_with_parse(self, user_message: str) -> dict:
        """
        HELP intent için mesajı parçalar.
        İşlem yapmaz, sadece rehberlik için bağlam çıkarır.
        """

        msg = user_message.lower()

        if any(k in msg for k in ["otopark", "park", "rezervasyon"]):
            topic = "otopark"
        elif any(k in msg for k in [
            "takvim", "haftalık", "ofis", "home", "izin", "işaretle"
        ]):
            topic = "haftalik_takvim"

        # 🔹 Excel / Export HELP (öncelikli!)
        elif any(k in msg for k in [
            "excel", "exel", "liste", "indir", "çıkar",
            "export", "dosya", "tablo"
        ]):
            topic = "company_people_export_help"

        elif any(k in msg for k in [
            "kim", "çalışan", "personel", "mail", "email", "telefon"
        ]):
            topic = "company_people"

        else:
            topic = "unknown"

        if any(k in msg for k in ["nasıl", "nasil", "nasıl yaparım", "nasil yaparim"]):
            question_type = "how"
        elif any(k in msg for k in [
            "yapabiliyor musun", "yapabilir misin",
            "mümkün mü", "mumkun mu"
        ]):
            question_type = "capability"
        else:
            question_type = "guide"

        keywords = []
        for k in [
            "otopark", "park", "rezervasyon",
            "takvim", "haftalık", "ofis", "home",
            "yardım", "nasıl"
        ]:
            if k in msg:
                keywords.append(k)

        return {
            "topic": topic,
            "question_type": question_type,
            "keywords": keywords
        }

    def llm_with_handle_help(self, parsed: dict, user_message: str) -> str:
        topic = parsed.get("topic")

        if topic == "unknown":
            return (
                "Hangi konuda yardım istediğini biraz açar mısın? 😊<br>"
                "Otopark, haftalık takvim veya çalışan bilgileri gibi konularda yardımcı olabilirim."
            )
        
        if topic == "company_people_export_help":
            return (
                "Çalışan listelerini Excel olarak almak için net bir komut yazman yeterli 😊<br><br>"
                "Örnekler:<br>"
                "• 'Tüm çalışanların mail adreslerini Excel al'<br>"
                "• 'IT departmanı çalışanlarını Excel çıkar'<br>"
                "• 'Çalışanların mail ve telefonlarını Excel indir'<br>"
                "• 'Tüm çalışanların tüm bilgilerini Excel ver'<br><br>"
                "Bu şekilde yazdığında Yugii Excel dosyasını otomatik oluşturur."
            )
        if topic == "otopark":
            context = (
                "Yugii, otopark rezervasyonu konusunda rehberlik eder.\n"
                "Kullanıcı isterse portal üzerinden, isterse Yugii'ye yazarak rezervasyon yapabilir.\n"
                "Örnekler:\n"
                "- 'Yarın 10 numaralı parkı ayır'\n"
                "- '17 kasım için park ayır'"
            )

        elif topic == "haftalik_takvim":
            context = (
                "Yugii, haftalık çalışma takvimi işaretleme konusunda rehberlik eder.\n"
                "Gün + çalışma tipi yazman yeterlidir.\n"
                "Örnekler:\n"
                "- 'Salı evden çalışma olarak işaretle'\n"
                "- 'Perşembe ofisten çalışacağım'"
            )

        elif topic == "company_people":
            context = (
                "Yugii, şirket içi çalışan rehberi konusunda bilgi verir.\n"
                "İsim, departman, email ve telefon bilgileri sorulabilir."
            )

        else:  # general
            context = (
                "Yugii, şirket içi işlemler ve bilgi taleplerinde yardımcı olur.\n"
                "Otopark, takvim, yemek menüsü ve çalışan bilgileri gibi konulara destek verir."
            )

        prompt = f"""
    Sen Yüce Auto'nun dijital asistanı Yugii'sin.

    Kurallar:
    - Yalnızca aşağıdaki bilgileri kullan.
    - İşlem yapma.
    - Kısa ve net cevap ver (maks 3–4 cümle).

    Bilgi:
    {context}

    Kullanıcı sorusu:
    "{user_message}"

    SADECE cevabı yaz.
    """

        try:
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=120,
                temperature=0.3
            )

            answer = resp.choices[0].message.content.strip()

            page_row = self.find_page_link(user_message)

            if isinstance(page_row, dict) and page_row.get("url"):
                answer += (
                    "<br><br>"
                    f"<a href='{page_row['url']}' target='_blank'>"
                    f"🔗 {page_row.get('page_name','Sayfa')} sayfasına git"
                    f"</a>"
                )

            return answer   

        except Exception as e:
            print("❌ llm_with_handle_help hata:", e)
            return "Bu konuda sana kısaca yardımcı olabilirim, biraz daha detay sorar mısın?"

    #otopark rezervasyon tarih anlama isim fix yapılacak otopark_parse_tarih() şeklinde
    def otopark_parse_tarih(self, user_message: str):

        today = datetime.now().date()
        text = user_message.lower().replace("'", "").replace("’", "").strip()

        def strip_weekday_suffixes(text: str) -> str:
            suffixes = ["ye","ya","de","da","den","dan","te","ta","nde","nda","inde","ında","ine","ına"]
            for day in ["pazartesi","salı","çarşamba","perşembe","cuma","cumartesi","pazar"]:
                for suf in suffixes:
                    text = re.sub(rf"\b{day}{suf}\b", day, text)
            return text

        def normalize_weekdays_fuzzy(text: str) -> str:
            weekdays = ["pazartesi","salı","çarşamba","perşembe","cuma","cumartesi","pazar"]
            words = text.split()
            for i, w in enumerate(words):
                match = difflib.get_close_matches(w, weekdays, n=1, cutoff=0.75)
                if match:
                    words[i] = match[0]
            return " ".join(words)

        #aylarda yazım yanlışını doğruya çevirmek için
        def normalize_months_fuzzy(text: str) -> str:
            months = [
                "ocak","şubat","mart","nisan","mayıs",
                "haziran","temmuz","ağustos",
                "eylül","ekim","kasım","aralık"
            ]
            words = text.split()
            for i, w in enumerate(words):
                match = difflib.get_close_matches(w, months, n=1, cutoff=0.75)
                if match:
                    words[i] = match[0]
            return " ".join(words)
        
        #ay eklerini temizlemek için
        def strip_month_suffixes(text: str) -> str:
            suffixes = [
                "da","de","ta","te",
                "dan","den","tan","ten",
                "a","e","ya","ye",
                "nda","nde","ında","inde",
                "nın","nin"
            ]
            months = [
                "ocak","şubat","mart","nisan","mayıs",
                "haziran","temmuz","ağustos",
                "eylül","ekim","kasım","aralık"
            ]
            for m in months:
                for suf in suffixes:
                    text = re.sub(rf"\b{m}{suf}\b", m, text)
            return text

        # --- normalize ---
        NORMALIZE_GROUPS = {
            "pazartesi": [
                "pzrtsi","pazrtesi","pazartes","pazartesii","pazartesı",
                "pazartesine","pazartesinde","pazartesinden","pazartesiye",
                "pazartesiyi","paazartesi","pazartesi̇","pzt","pzrtsı"
            ],

            "salı": [
                "sali","salıi","salıı","saliya","saliye","salida",
                "salıda","salidan","salıdan","saali","salii","slı",
            ],

            "çarşamba": [
                "carsamba","çarsamba","çarsmaba","çarşmaba","çarşambaı",
                "carsambaa","carsambada","carsambaya","çarsambaa","carsmba"
            ],

            "perşembe": [
                "persembe","perşmbe","perşm","prsmb","ersembe",
                "perşembeı","perşembede","perşembeye","perşembeden","perşembeee","perş","pers","persenbe","persnbe"
            ],

            "cuma": [
                "cumaa","cumaı","cumaya","cumada","cumadan",
                "cumaaa","cuuma","cumaı","cumeya","cumae"
            ],

            "cumartesi": [
                "cumrtesi","cumartes","cumartesii","cumartesiye","cumartesinde",
                "cumartesinden","cumaratesi","cumartesı","cumartesii","cumartes"
            ],

            "pazar": [
                "pazarr","pazara","pazarda","pazardan","pazarı",
                "pazarr","pazzar","paazar","pazarrr","pazarrda"
            ],
            "haftaya": [
                "hafataya", "haftay", "haftayaa", "haftyaa",
                "haftayaa", "haftayaaa", "haftayae","haftya","hftaya","aftaya","hftaya"
            ],

            # --- GÜNSEL İFADELER ---
            "bugün": [
                "bugun","bugünu","bugüne","bugunu","bugünde",
                "bugunden","bu gun","bugünn","bugünı","bugunn"
            ],

            "yarın": [
                "yarin","yarına","yarını","yarinda","yarından",
                "yarinn","yarınn","yarn","yaarın","yarınnn"
            ],

            "öbür gün": [
                "obur gun","oburgun","öbürgun","obur gunn","öbür günn",
                "obur gunu","obur gune","öbür gune","obur gunden","öbürgün"
            ],

            "ertesi gün": [
                "ertesi gun","ertesigun","ertesi gunn","ertesii gun",
                "ertesi gune","ertesi gunden","ertesigunu","ertesi güni","ertsi gun"
            ],
    
            "ocak": ["ocak"],
            "şubat": ["subat","şubatta","subatta","subata"],
            "mart": ["mart","martta","marta"],
            "nisan": ["nisan","nisanda","nisana"],
            "mayıs": ["mayis","mayista","mayisa"],
            "haziran": ["haziran","haziranda","hazirana"],
            "temmuz": ["temmuz","temmuzda","temmuza"],
            "ağustos": ["agustos","agust","agu","agustosta","agustosa"],
            "eylül": ["eylul","eylule","eylulde"],
            "ekim": ["ekim","ekimde","ekime"],
            "kasım": ["kasim","kasm","kasımda","kasima"],
            "aralık": ["aralik","aralikta","araliga"]
        }

        for correct, variants in NORMALIZE_GROUPS.items():
            for v in variants:
                text = text.replace(v, correct)


        #  GÜN NORMALIZATION (TEK YER)
        text = strip_weekday_suffixes(text)
        text = normalize_weekdays_fuzzy(text)

        #  AY NORMALIZATION (TEK YER)
        text = strip_month_suffixes(text)
        text = normalize_months_fuzzy(text)

        TEMPORAL_ANCHORS = {
            "next_week": [
                "haftaya",
                "gelecek hafta",
                "önümüzdeki hafta",
                "bir sonraki hafta",
                "diger hafta"
            ],
            "this_week": [
                "bu hafta"
            ]
        }
        tokens = text.split()

        # 🔒 GLOBAL ANCHOR (yazım hatasında anchor düşmesin diye)
        detected_global_anchor = None
        for anchor, variants in TEMPORAL_ANCHORS.items():
            for v in variants:
                if v in text:
                    detected_global_anchor = anchor
                    break
        MONTHS = {
            "ocak":1,"şubat":2,"mart":3,"nisan":4,"mayıs":5,
            "haziran":6,"temmuz":7,"ağustos":8,
            "eylül":9,"ekim":10,"kasım":11,"aralık":12
        }
        dates = []
        # BUGÜN / YARIN / ÖBÜR GÜN
        text = text.lower()
        text = re.sub(r"\s+", " ", text)

        # BUGÜN / YARIN / DÜN
        if any(t in text for t in RELATIVE_DAY_TRIGGERS["today"]):
            dates.append(today)

        if any(t in text for t in RELATIVE_DAY_TRIGGERS["tomorrow"]):
            dates.append(today + timedelta(days=1))

        if "öbür gün" in text or "ertesi gün" in text:
            dates.append(today + timedelta(days=2))

        #  HAFTAYA + BUGÜN (SADECE YAN YANA)
        if contains_next_week_relative(text, "today", max_gap=1):
            return {"dates": [today + timedelta(days=7)]}

        #  HAFTAYA + YARIN (SADECE YAN YANA)
        if contains_next_week_relative(text, "tomorrow", max_gap=1):
            return {"dates": [today + timedelta(days=8)]}
        
        #  HAFTAYA ama GÜN YOKSA → Pazartesi–Cuma
        if any(k in text for k in NEXT_WEEK_KEYS) and not any(
            g in text for g in ALL_WEEKDAY_VARIANTS
        ):
            monday = today + timedelta(days=(7 - today.weekday()))
            dates = [monday + timedelta(days=i) for i in range(5)]
            return {"dates": dates}

        # GÜN KAPSAMI (HANGİ GÜNLER?)
        DAY_SCOPES = {
            "all_days": [
                "tum gunler", "tüm günler","hafta boyu her gün ","hafta boynca","tüm hfta",
                "tum hafta", "tüm hafta","tüm hafta içi","hafta içi hergün",
                "hafta boyunca",
                "komple",
                "tamami", "tamamı"
            ],
            "weekdays": [
                "hafta ici", "hafta içi",
                "is gunleri", "iş günleri"
            ],
            "weekend": [
                "hafta sonu","hfta sonu",
                "weekend"
            ]
        }

        #  HER GÜN / HER GÜNE ALGILAMA
        has_every_day = any(k in text for k in EVERY_DAY_KEYS)

        if has_every_day:
            #  Haftaya her gün
            if detected_global_anchor == "next_week":
                base_monday = today + timedelta(days=(7 - today.weekday()))
                dates = [
                    base_monday + timedelta(days=i)
                    for i in range(5)   # hafta içi
                ]
                return {"dates": dates}

            #  Bu hafta her gün
            if detected_global_anchor == "this_week":
                base_monday = today - timedelta(days=today.weekday())
                dates = [
                    base_monday + timedelta(days=i)
                    for i in range(5)
                    if base_monday + timedelta(days=i) >= today
                ]
                return {"dates": dates}

            #  Bu ay her gün
            if any(k in text for k in THIS_MONTH_KEYS):
                start = today
                end = datetime(
                    today.year,
                    today.month,
                    calendar.monthrange(today.year, today.month)[1]
                ).date()

                dates = []
                cur = start
                while cur <= end:
                    if cur.weekday() < 5:   # hafta içi
                        dates.append(cur)
                    cur += timedelta(days=1)

                return {"dates": dates}

            #Sadece "her gün" (anchor yoksa → bugünden itibaren 5 iş günü)
            dates = []
            cur = today
            while len(dates) < 5:
                if cur.weekday() < 5:
                    dates.append(cur)
                cur += timedelta(days=1)

            return {"dates": dates}

        # SCOPE + ANCHOR (tüm hafta / hafta içi / hafta sonu)
        detected_scope = None
        for scope, variants in DAY_SCOPES.items():
            for v in variants:
                if v in text:
                    detected_scope = scope
                    break

        # Anchor zaten normalize sonrası global tespit edildi
        if detected_scope and detected_global_anchor:

            if detected_global_anchor == "next_week":
                base_monday = today + timedelta(days=(7 - today.weekday()))
            else:
                base_monday = today - timedelta(days=today.weekday())

            if detected_scope == "all_days":
                for i in range(7):
                    dates.append(base_monday + timedelta(days=i))

            elif detected_scope == "weekdays":
                for i in range(5):
                    dates.append(base_monday + timedelta(days=i))

            elif detected_scope == "weekend":
                dates.append(base_monday + timedelta(days=5))
                dates.append(base_monday + timedelta(days=6))

        # WEEKDAY ÇÖZÜMÜ (scope YOKSA)
        if not detected_scope:

            WEEKMAP = {
                "pazartesi":0,"salı":1,"çarşamba":2,
                "perşembe":3,"cuma":4,"cumartesi":5,"pazar":6
            }

            def resolve_weekday_with_anchor(gun, anchor):
                idx = WEEKMAP[gun]

                if anchor == "next_week":
                    base = today + timedelta(days=(7 - today.weekday()))
                    return base + timedelta(days=idx)

                if anchor == "this_week":
                    base = today - timedelta(days=today.weekday())
                    return base + timedelta(days=idx)

                diff = idx - today.weekday()
                if diff <= 0:
                    diff += 7
                return today + timedelta(days=diff)

            for token in tokens:
                if token in WEEKMAP:
                    dates.append(
                        resolve_weekday_with_anchor(
                            token,
                            detected_global_anchor
                        )
                    )

        # TÜM AY / TÜM AY BOYUNCA
        if "tüm ay" in text and ("komple" in text or "boyunca" in text or "tamamı" in text):
            start = today
            end = datetime(today.year, today.month,
                        calendar.monthrange(today.year, today.month)[1]).date()
            cur = start
            while cur <= end:
                dates.append(cur)
                cur += timedelta(days=1)

        # BU AY KOMPLE / BU AY BOYUNCA        
        if "bu ay" in text and ("komple" in text or "boyunca" in text or "tamamı" in text):
            start = today
            end = datetime(today.year, today.month,
                        calendar.monthrange(today.year, today.month)[1]).date()
            cur = start
            while cur <= end:
                dates.append(cur)
                cur += timedelta(days=1)

        for ay, ay_num in MONTHS.items():
            if f"{ay} boyunca" in text or f"{ay} komple" in text or f"{ay} tamamı" in text:
                start = datetime(today.year, ay_num, 1).date()
                end = datetime(today.year, ay_num, calendar.monthrange(today.year, ay_num)[1]).date()
                cur = start
                while cur <= end:
                    dates.append(cur)
                    cur += timedelta(days=1)

        # AY + GÜN (15 kasım)
        for gun, ay in re.findall(
            r"(\d{1,2})\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)",
            text
        ):
            try:
                dates.append(datetime(today.year, MONTHS[ay], int(gun)).date())
            except:
                pass

        # TARİH ARALIĞI (15 kasım - 20 kasım)
        m = re.search(
            r"(\d{1,2})\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)"
            r".*?(?:-|–|ile|den|dan).*?"
            r"(\d{1,2})\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)",
            text
        )
        if m:
            g1, a1, g2, a2 = m.groups()
            try:
                d1 = datetime(today.year, MONTHS[a1], int(g1)).date()
                d2 = datetime(today.year, MONTHS[a2], int(g2)).date()
                if d1 > d2:
                    d1, d2 = d2, d1
                cur = d1
                while cur <= d2:
                    dates.append(cur)
                    cur += timedelta(days=1)
            except:
                pass

        # ------------------------------------------------
        # FORMATLI TARİHLER (15.11 / 15-11 / 15/11)
        # ------------------------------------------------
        for s in re.findall(r"\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?", text):
            try:
                dates.append(dateutil.parser.parse(s, dayfirst=True).date())
            except:
                pass

        return {
            "dates": sorted(set(dates))
        }

    def gpt_fix_calendar_text(self, text: str) -> str:
        """
        "Metindeki yazım hatalarını düzelt. "
        "ANLAMI ASLA değiştirme. "
        "⚠️ GÜN VE TARİH ifadelerine asla dokunma. "
        "Perşembe → cuma gibi düzeltmeler YASAK. "
        "Sadece yazım hatalarını düzelt. "
        "Ör: 'perişme' → 'perşembe', 'hafaya' → 'haftaya'. "
        "Ancak anlamı değiştirecek tahminler yapma."

        """
        try:
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Metindeki sadece yazım hatalarını düzelt. "
                            "ANLAMI ASLA değiştirme. "
                            "⚠️ GÜN isimlerini ASLA değiştirme: pazartesi, salı, çarşamba, perşembe, cuma, cumartesi, pazar—"
                            "bu isimlerde düzeltme YAPABİLİRSİN ama BAŞKA bir güne çeviremezsin. "
                            "Örnek: 'perşenbe' → 'perşembe' olur, ama 'perşembe' → 'cuma' OLMAZ. "
                            "Tarihsel veya mantıksal tahmin yapma. "
                            "Kullanıcı ne yazdıysa O kalsın."
                        )
                    },

                    {"role": "user", "content": text}
                ],
                max_tokens=80,
                temperature=0
            )
            return response.choices[0].message.content.strip()
        except:
            return text
        
    def gpt_extract_dates(self, text: str):
        """
        Karmaşık doğal dil tarih ifadelerini anlamak için GPT'yi kullanır.
        Örneğin: 'yarından 3 gün sonra', 'önümüzdeki ay ortası'.
        Çıktı ISO formatlı tarih listesi olur: ['2024-11-15', '2024-11-18'].
        """
        today = datetime.now().date()

        prompt = f"""
        Bugünün tarihi: {today}

        Kullanıcı mesajındaki karmaşık tarih ifadelerini analiz et.
        - 'yarından sonra 3 gün'
        - 'gelecek ayın ortası'
        - 'haftaya salıdan iki gün sonra'
        - '3 gün sonra'
        gibi ifadeleri kesin tarihlere çevir.

        Sadece ISO formatlı Python listesi döndür:
        Örnek: ["2024-11-15", "2024-11-18"]

        Eğer hesaplanacak tarih yoksa: []
        """

        try:
            result = self.client.chat.completions.create(
                model="gpt-4.1",
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": text}
                ],
                max_tokens=120,
                temperature=0
            )
            raw = result.choices[0].message.content.strip()

            if raw.startswith("[") and raw.endswith("]"):
                return eval(raw)
            return []
        except:
            return []

    def resolve_calendar_dates(self, raw_message: str):
        """
        GPT düzeltme (YOL 1) + GPT tarih çıkarımı (YOL 2) +
        klasik tarih parser (parse_takvim_tarih_araligi)
        → En doğru tarih listesi.
        """

        cleaned = raw_message
        work_types_in_text = set()
        tokens = cleaned.split()

        for i in range(len(tokens)):
            wt1 = self.detect_work_type(tokens[i])
            if wt1:
                work_types_in_text.add(wt1)

            if i + 1 < len(tokens):
                wt2 = self.detect_work_type(tokens[i] + " " + tokens[i+1])
                if wt2:
                    work_types_in_text.add(wt2)

        # 🔎 GÖRELİ GÜN VAR MI?
        norm_cleaned = self.normalize_calendar_text(cleaned)

        HAS_RELATIVE_DAY = any(
            k in norm_cleaned for k in ["bugun", "yarin", "obur gun", "ertesi gun"]
        )
        # 🚫 GPT’nin gün isimlerini yanlış düzeltmesini engelle
        gunler = ["pazartesi","salı","sali","çarşamba","carsamba",
                  "perşembe","cuma","cumartesi","pazar"]

        for g in gunler:
            if g in raw_message.lower() and g not in cleaned.lower().replace("ı", "i").replace("İ", "i"):
                print("⚠️ GPT gün ismini yanlış değiştirdi → orijinale dönüldü:", g)
                cleaned = raw_message  # GPT düzeltmesini iptal et
                break


        # 2) GPT tarih çıkarımı dene
        gpt_dates = self.gpt_extract_dates(cleaned)

        norm = self.normalize_calendar_text(cleaned)
        if any(k in norm for k in NEXT_WEEK_KEYS):
            print("❗ [DEBUG] NEXT_WEEK_KEYS tespit edildi → GPT tarihleri yok sayıldı:", gpt_dates)
            gpt_dates = []

        # Eğer GPT bir tarih listesi döndürdüyse → hemen kullan
        if gpt_dates:
            parsed_dates = []
            for iso in gpt_dates:
                try:
                    dt = datetime.fromisoformat(iso).date()
                    parsed_dates.append(dt)
                except:
                    pass

            if parsed_dates:
                # Bu tarihleri klasik parser olmadan direkt döndür
                return {"dates": parsed_dates, "warning": None}

        # 🔥 ÇOKLU AY FULL (ocak ve şubat full / boyunca / komple)
        MONTHS = {
            "ocak":1,"subat":2,"şubat":2,"mart":3,"nisan":4,
            "mayis":5,"mayıs":5,"haziran":6,"temmuz":7,
            "agustos":8,"ağustos":8,"eylul":9,"eylül":9,
            "ekim":10,"kasim":11,"kasım":11,"aralik":12,"aralık":12
        }

        text = self.normalize_calendar_text(cleaned)

        multi_months = re.findall(
            r"\b(ocak|subat|şubat|mart|nisan|mayis|mayıs|haziran|temmuz|agustos|ağustos|eylul|eylül|ekim|kasim|kasım|aralik|aralık)\b",
            text
        )

        if len(multi_months) >= 2 and any(k in text for k in ["full", "komple", "tamami", "boyunca"]):
            dates = []
            today = datetime.now().date()

            for ay in dict.fromkeys(multi_months):  # tekrarları sil
                ay_num = MONTHS.get(ay)
                if not ay_num:
                    continue

                start = datetime(today.year, ay_num, 1).date()
                end = datetime(today.year, ay_num, calendar.monthrange(today.year, ay_num)[1]).date()

                cur = start
                while cur <= end:
                    if cur.weekday() < 5:
                        dates.append(cur)
                    cur += timedelta(days=1)

            return {"dates": sorted(set(dates)), "warning": None}

        # 3) GPT işini yapamadıysa → klasik parser’ı çalıştır
        return self.parse_takvim_tarih_araligi(cleaned)


    def _detect_weekday_count_scope(self, text: str):
        """
        'haftaya 5 gün', 'ilk 3 gün', 'haftanın ilk 2 günü'
        return: int | None
        """
        patterns = [
            r"(haftaya|gelecek hafta|onumuzdeki hafta).*(\d)\s*gun",
            r"haftanin\s*ilk\s*(\d)\s*gun",
            r"ilk\s*(\d)\s*gun"
        ]

        for p in patterns:
            m = re.search(p, text)
            if m:
                return int(m.group(2) if m.lastindex and m.lastindex >= 2 else m.group(1))

        return None
    

    def handle_otopark_create(self, user_message: str, user_info: dict):
        print("🅿️ CREATE modülü çalışıyor...")

        if any(k in user_message.lower() for k in [
            "boş", "dolu", "müsait", "müsaitlik",
            "yer var", "yer varmı", "yer var mı",
            "boşluk", "uygun","boş yer varmı","musaitmi","doluluk nedir","dolumu","dolu mu"
        ]):
            print(" CREATE engellendi → STATUS'a yönlendirildi")

            return self.handle_otopark_status(user_message, user_info)
        
        username = user_info.get("fullname")
        plaka = user_info.get("plaka")
        plaka2 = user_info.get("plaka2") or user_info.get("arac_plaka2")

        is_company_car = bool(plaka2)

        msg_norm = self.normalize_people_text(user_message)

        # 0) Kullanıcının aracı var mı?
        if not plaka and not plaka2:
            return "🚫 Otopark rezervasyonu yapabilmek için sistemde kayıtlı bir aracınız bulunmamaktadır."

        # 1) Kullanıcının tarih talebini çöz
        parsed = self.otopark_parse_tarih(user_message)
        print("📅 Tarih Parse:", parsed)

        #Saat 8 kuralı 
        today = get_effective_today()

        #Son 3 iş günü kuralı  
        #Hafta sonu sayılmıyor, Geriye doğru 3 iş günü bulunuyor
        three_days_ago = today - timedelta(days=3)

        try:
            conn_check = pyodbc.connect(YA_2El_AracSatis)
            cur_check = conn_check.cursor()

            #Gelinmeyen rezervasyon son 3 İŞ GÜNÜ içindeyse yeni rezervasyona izin vermiyor.
            cur_check.execute("""
                SELECT COUNT(*)
                FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                WHERE username = ?
                AND is_active = 1
                AND Geldi = 0
                AND canceled_at IS NULL
                AND rezerv_tarih >= ?
            """, (username, three_days_ago))

            no_show_count = cur_check.fetchone()[0]
            conn_check.close()

        except Exception as e:
            print("❌ No-show kontrol hatası:", e)
            no_show_count = 0  # fail-open (engelleme yapma)

        if no_show_count > 0:
            return (
                "🚫 Son 3 gün içinde oluşturduğunuz bir otopark rezervasyonunda "
                "şirkete giriş yapılmadığı tespit edildi.<br>"
                "Otopark kullanım verimliliğini artırmak adına, "
                "<b>3 gün sonra</b> tekrar rezervasyon oluşturabilirsiniz."
            )

        # === TARİH FİLTRELEME (TEK DOĞRU YER) ===
        dates = parsed.get("dates", [])

        #otopark create rezervasyon oluşturma
        today = get_effective_today()
        limit = today + timedelta(days=7)

        success_rows = []  
        info_rows = []      
        error_rows = []    
        
        valid_dates = []
        past_dates = []
        too_far_dates = []
        weekend_dates = []
        created_any = False


        for d in dates:
            if d.weekday() >= 5:   # 5=Cumartesi, 6=Pazar
                weekend_dates.append(d)
                continue
            if d < today:
                past_dates.append(d)
            elif d > limit:
                too_far_dates.append(d)
            else:
                valid_dates.append(d)

        for d in weekend_dates:
            error_rows.append((
                d,
                f"🚫 {self.format_date_with_day(d)} → Hafta sonu olduğu için işlenmedi."
            ))

        if not dates:
            return (
                "🅿️ Otopark rezervasyonu oluşturabilmem için bir gün belirtmen gerekiyor 😊<br><br>"
                "Örnekler:<br>"
                "• <b>Yarın için park yeri ayır</b><br>"
                "• <b>Pazartesi ve salı ofise geleceğim, otoparkta park 15 rezerve et</b><br>"
                "• <b>17 kasım için otopark ayır</b>"
            )
        
        # 3) Yetki alanı belirle
        if is_company_car:
            allowed_range = list(range(10, 44))
            priority_range = list(range(36, 44))  # 🔥 öncelikli havuz
        else:
            allowed_range = list(range(10, 36))
            priority_range = []

        print("🔐 Yetkili park aralığı:", allowed_range)
        
        if not valid_dates:
            if past_dates and not too_far_dates:
                return "❌ Girdiğiniz tüm tarihler geçmiş olduğu için işlem yapılamadı."
            if too_far_dates and not past_dates:
                return "⏳ Girdiğiniz tarihler 7 gün sınırını aştığı için işlem yapılamadı."
            if past_dates and too_far_dates:
                return "❌ Geçmiş ve 7 günü aşan tarihler olduğu için işlem yapılamadı."

        try:
            conn = pyodbc.connect(YA_2El_AracSatis)
            cur = conn.cursor()
        except Exception as e:
            print("❌ DB hata:", e)
            return "⚠️ Sistem geçici olarak kullanılamıyor."

        requested_park = None

        # Kullanıcı park numarası yazdıysa yakala
        park_no_match = re.search(
            r"\bpark\s*(?:no\s*)?(\d{1,2})\b",
            user_message.lower()
        )
        if park_no_match:
            requested_park = int(park_no_match.group(1))

        # === ANA DÖNGÜ ===
        for t in valid_dates:
            gun_str = self.format_date_with_day(t)

            # 1) Kullanıcının o gün zaten rezervasyonu var mı?
            cur.execute("""
                SELECT COUNT(*)
                FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                WHERE username=? AND rezerv_tarih=? AND is_active=1
            """, (username, t))
            already = cur.fetchone()[0]

            if already > 0:
                info_rows.append((
                    t,
                    f"♻️ {gun_str} → Aktif rezervasyonunuz var."
                ))
                continue

            if requested_park:

                # Yetkisi var mı?
                if requested_park not in allowed_range:
                    error_rows.append((
                        t,
                        f"🚫 {gun_str} → {requested_park} numaralı park için yetkiniz yok."
                    ))
                    continue
                if requested_park < 10:
                    error_rows.append((
                        t,
                        f"🚫 {gun_str} → 1–9 arası parklar rezerve edilemez."
                    ))
                    continue

                # Park dolu mu?
                cur.execute("""
                    SELECT COUNT(*) 
                    FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                    WHERE rezerv_tarih=? AND rezerv_park_no=? AND is_active=1
                """, (t, requested_park))
                dolu = cur.fetchone()[0]

                if dolu > 0:
                    # 🔥 O günün boş parklarını bul
                    cur.execute("""
                        SELECT rezerv_park_no
                        FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                        WHERE rezerv_tarih=? AND is_active=1
                    """, (t,))
                    dolu_liste = {int(r[0]) for r in cur.fetchall()}
                    bos_liste = [p for p in allowed_range if p not in dolu_liste]

                    bos_str = ", ".join(map(str, bos_liste[:12]))

                    error_rows.append((
                        t,
                        f"🚫 {gun_str} → {requested_park} numaralı park dolu. "
                        f"Uygun park sayısı: {len(bos_liste)}"
                    ))
                    continue


                # Park boş → oluştur
                cur.execute("""
                    INSERT INTO [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                        (username, plaka, rezerv_tarih, rezerv_park_no, created_at, is_active, aciklama, Geldi, guvenlik_aciklama)
                    VALUES (?, ?, ?, ?, GETDATE(), 1, '', 1, '')
                """, (username, plaka or plaka2, t, requested_park))
                
                conn.commit()

                created_any = True

                success_rows.append((
                    t,
                    f"🅿️ {gun_str} → {requested_park} numaralı park rezerve edildi."
                ))
                continue

            # === Kullanıcı park numarası belirtmediyse → RANDOM PARK AYIR ===
            
            # O günün dolu parklarını al
            cur.execute("""
                SELECT rezerv_park_no
                FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                WHERE rezerv_tarih=? AND is_active=1
            """, (t,))
            dolu_olanlar = {int(r[0]) for r in cur.fetchall()}

            # Boş parkları belirle
            bos_olanlar = [p for p in allowed_range if p not in dolu_olanlar]

            if not bos_olanlar:
                error_rows.append((
                    t,
                    f"🚫 {gun_str} → Uygun park bulunamadı."
                ))
                continue

            # 🚗 ŞİRKET ARACI ÖNCELİĞİ (36–43)
            if priority_range:
                priority_boslar = [p for p in priority_range if p in bos_olanlar]

                if priority_boslar:
                    secilen = random.choice(priority_boslar)  # önce 36–43
                else:
                    secilen = random.choice(bos_olanlar)      # fallback
            else:
                secilen = random.choice(bos_olanlar)

            # RANDOM → INSERT
            cur.execute("""
                INSERT INTO [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                    (username, plaka, rezerv_tarih, rezerv_park_no, created_at, is_active, aciklama, Geldi, guvenlik_aciklama)
                VALUES (?, ?, ?, ?, GETDATE(), 1, '', 1, '')
            """, (username, plaka or plaka2, t, secilen))
            conn.commit()

            created_any = True

            success_rows.append((
                t,
                f"🅿️ {gun_str} → {secilen} numaralı park rezerve edildi."
            ))

        for d in past_dates:
            error_rows.append((
                d,
                f"❌ {self.format_date_with_day(d)} → Geçmiş tarih."
            ))


        for d in too_far_dates:
            error_rows.append((
                d,
                f"⏳ {self.format_date_with_day(d)} → 7 gün kuralı nedeniyle oluşturulmadı."
            ))

        conn.close()

        def _sort(rows):
            return [text for _, text in sorted(rows, key=lambda x: x[0])]

        success_rows = _sort(success_rows)
        info_rows = _sort(info_rows)
        error_rows = _sort(error_rows)
    
        #mesaj son 
        intro = self.naturalize_intro("otopark_create")
        
        sections = []

        if success_rows:
            sections.append(
                "🅿️ <b>Otopark Rezervasyonu</b><br>" +
                "<br>".join(success_rows)
            )

        if info_rows:
            sections.append(
                "ℹ️ <b>Bilgilendirme</b><br>" +
                "<br>".join(info_rows)
            )

        if error_rows:
            sections.append(
                "⚠️ <b>İşlem Yapılamayan Günler</b><br>" +
                "<br>".join(error_rows)
            )

        body = "<br><br>".join(sections)

        has_calendar_action = any(f in msg_norm for f in self.TAKVİM_KEYS)
        has_work_type = bool(self.detect_work_type(msg_norm))

        has_calendar_intent = has_calendar_action and has_work_type


        if created_any:
            outro_base = "Uygun günler için otopark rezervasyonu oluşturuldu."
        else:
            outro_base = "Otopark rezervasyonu oluşturulamadı."

        # Haftalık takvim niyeti varsa → LLM outro YOK
        if has_calendar_intent:
            outro = outro_base          
        else:
            outro = self.naturalize_text(outro_base, "otopark_create")


        final = ""
        if intro:
            final += intro + "<br>"

        final += body

        if outro:
            final += "<br>" + outro


        if has_calendar_intent:
            final += (
                "<br><br>🗓️ Ayrıca haftalık takvimle ilgili de bir işlem yapmak istediğini fark ettim.<br>"
                "Bu işlem için ayrıca yardımcı olabilirim tekarar söylemen yeterli."
            )
        return final


    def handle_otopark_cancel(self, user_message: str, user_info: dict):
        print("❌ CANCEL modülü çalışıyor...")

        username = user_info.get("fullname")
        plaka = user_info.get("plaka") or user_info.get("plaka2")

        intro = self.naturalize_intro("otopark_cancel")
 
        if not username:
            return "🚫 Kullanıcı bilgisi bulunamadı."

        msg_norm = self.normalize_people_text(user_message)
        skip_bulk = any(k in msg_norm for k in ["var mı", "var mi", "kontrol", "göster"])


        def is_bulk_cancel(text: str) -> bool:
            text = self.normalize_people_text(text)
            tokens = set(text.split())

       
            has_cancel = bool(tokens & CANCEL_KEYS)

            #  2) TOPLULUK (token + substring)
            ALL_TOKEN_KEYS = {
                "tum", "hepsi", "hepsini", "tamami",
                "komple", "mevcut", "aktif", "olan", "varsa"
            }
            has_all = (
                bool(tokens & ALL_TOKEN_KEYS)
                or any(self.normalize_people_text(k) in text for k in FULL_KEYS)
            )

            has_otopark_context = (
                any(self.normalize_people_text(k) in text for k in OTOPARK_KEY)
                or any(self.normalize_people_text(k) in text for k in USER_PARK_STATUS_KEYS)
            )

            return has_cancel and has_all and has_otopark_context

        now = datetime.now()
        current_time = now.time()
        today = get_effective_today()

        # TOPLU İPTAL — EN BAŞTA
        if not skip_bulk and is_bulk_cancel(user_message):

            try:
                conn = pyodbc.connect(YA_2El_AracSatis)
                cur = conn.cursor()

                cur.execute("""
                    SELECT rezerv_tarih, rezerv_park_no
                    FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                    WHERE username=? AND is_active=1
                """, (username,))
                rows = cur.fetchall()

                if not rows:
                    return "📭 Aktif bir otopark rezervasyonunuz bulunmuyor."

                today = get_effective_today()

                iptal_edilebilir = []
                engellenen = []

                for rezerv_tarih, park_no in rows:

                    # geçmiş
                    if rezerv_tarih < today:
                        engellenen.append(
                            f"{self.format_date_with_day(rezerv_tarih)} — iptal edilemez"
                        )
                        continue

                    # bugün ama 08:00 sonrası
                    if (
                        rezerv_tarih == today
                        and current_time > datetime.strptime("08:00", "%H:%M").time()
                    ):
                        engellenen.append(
                            f"{self.format_date_with_day(rezerv_tarih)} — 08:00 sonrası iptal edilemez"
                        )
                        continue

                    # iptal edilebilir
                    iptal_edilebilir.append((rezerv_tarih, park_no))

                if not iptal_edilebilir:
                    return (
                        "🚫 İptal edilebilecek bir otopark rezervasyonunuz bulunmuyor.<br>"
                        "ℹ️ Bugünkü ve geçmiş rezervasyonlar iptal edilemez."
                    )

                for rezerv_tarih, _ in iptal_edilebilir:
                    cur.execute("""
                        UPDATE [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                        SET is_active=0,
                            Geldi=0,
                            canceled_at=GETDATE()
                        WHERE username=? AND rezerv_tarih=? AND is_active=1
                    """, (username, rezerv_tarih))

                conn.commit()

                lines = [
                    f"❌ {self.format_date_with_day(t)} — 🅿️ {p} numaralı park iptal edildi."
                    for t, p in iptal_edilebilir
                ]


                return "🗑️ Gelecek tarihli otopark rezervasyonlarınız iptal edildi:<br>" + "<br>".join(lines)

            except Exception as e:
                print("❌ Toplu iptal hata:", e)
                return "⚠️ Toplu iptal sırasında hata oluştu."
            finally:
                try:
                    conn.close()
                except:
                    pass

        parsed = self.otopark_parse_tarih(user_message)

        print("📅 CANCEL → Tarih Parse:", parsed)

        tarih_list = parsed.get("dates", [])

        if not tarih_list:
            return (
                "📅 İptal etmek istediğiniz tarihi anlayamadım.<br>"
                "Örn: <b>“Pazartesi rezervasyonumu iptal et”</b>"
            )

        # 2) DB bağlan
        try:
            conn = pyodbc.connect(YA_2El_AracSatis)
            cur = conn.cursor()
        except Exception as e:
            print("DB hata:", e)
            return "⚠️ Sistem geçici olarak kullanılamıyor."

        iptal_edilenler = []
        bulunamayanlar = []
        zaman_kural_engeli = []
        canceled_any = False

        #otopark cancel saat ve sınır kontrol
        now = datetime.now()
        current_time = now.time()
        today_date = get_effective_today()

        for t in tarih_list:

            #  GEÇMİŞ TARİH İPTAL EDİLEMEZ
            if t < today_date:
                zaman_kural_engeli.append(
                    f"{self.format_date_with_day(t)} — Geçmiş tarih, iptal edilemez."
                )
                continue

            # 🔒 08:00 sonrası bugünkü rezervasyon iptal edilemez
            if t == today_date and current_time > datetime.strptime("08:00", "%H:%M").time():
                zaman_kural_engeli.append(
                    f"{self.format_date_with_day(t)} — 08:00 sonrası iptal edilemez."
                )
                continue

            # 🔍 O güne ait rezervasyon var mı?
            cur.execute("""
                SELECT rezerv_park_no
                FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                WHERE username=? AND rezerv_tarih=? AND is_active=1
            """, (username, t))
            row = cur.fetchone()

            if not row:
                bulunamayanlar.append(
                    f"{self.format_date_with_day(t)} — Aktif bir rezervasyon bulunamadı."
                )
                continue

            park_no = int(row[0])

            #  İPTAL ET
            try:
                cur.execute("""
                    UPDATE [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                    SET is_active = 0,
                        Geldi = 0,
                        canceled_at = GETDATE()
                    WHERE username=? AND rezerv_tarih=? AND is_active=1
                """, (username, t))
                conn.commit()

                canceled_any = True


                iptal_edilenler.append(
                    f"{self.format_date_with_day(t)} — 🅿️ {park_no} numaralı park iptal edildi."
                )

            except Exception as e:
                print("Cancel UPDATE hata:", e)
                bulunamayanlar.append(
                    f"{self.format_date_with_day(t)} — iptal işlemi başarısız."
                )

        conn.close()

        cevap_satirlari = []

        # İptal edilenler
        for x in iptal_edilenler:
            tarih_raw, _ = x.split(" — ", 1)
            cevap_satirlari.append(f"✅ {tarih_raw} tarihli rezervasyon iptal edildi.")

        # 08:00 kuralı
        for x in zaman_kural_engeli:
            tarih_raw, _ = x.split(" — ", 1)
            cevap_satirlari.append(f"🚫 {tarih_raw} günü için 08:00 sonrası iptal yapılamaz.")

        body = "<br>".join(cevap_satirlari)

        # Outro — doğal kapanış cümlesi
        if canceled_any:
            outro_base = "Otopark rezervasyonu iptal edildi."
        else:
            outro_base = "İptal edilebilecek aktif bir otopark rezervasyonu bulunamadı."

        outro = self.naturalize_text(outro_base, "otopark_cancel")

        # Final mesajı birleştir
        final = ""

        if intro:
            final += intro + "<br>"

        final += body

        if outro:
            final += "<br>" + outro

        return final


    def handle_otopark_status(self, user_message: str, user_info: dict):
        print("📊 STATUS (GENEL MÜSAİTLİK) modülü çalışıyor...")

        # 🔹 1) Intro — doğal giriş cümlesi
        intro = self.naturalize_intro("otopark_status")

        # 🔹 2) Tarih çözümleme
        parsed = self.otopark_parse_tarih(user_message)
        print("📅 STATUS → Tarih Parse:", parsed)

        parsed_type = parsed.get("type")

        if parsed_type is None:
            if parsed.get("dates"):
                parsed_type = "single"
            else:
                parsed_type = "error"

        if parsed_type == "error":
            body = "📅 Hangi gün için park müsaitliğini sorguladığını anlayamadım."
            return f"{intro}<br>{body}"

        # -- Tarih listesi --
        if parsed_type in ["single", "multi"]:
            tarih_list = parsed.get("dates", [])
        elif parsed_type == "range":
            d1, d2 = parsed["start"], parsed["end"]
            tarih_list = [d1 + timedelta(days=i) for i in range((d2 - d1).days + 1)]
        else:
            body = "📅 Geçerli bir tarih bulamadım."
            return f"{intro}<br>{body}"

        #  GÜN SINIRI KONTROLÜ (FİLTRELE)
        today = datetime.now().date()
        max_date = today + timedelta(days=7)

        #  SADECE UYGUN TARİHLERİ TUT
        filtered_dates = [t for t in tarih_list if today <= t <= max_date]

        if not filtered_dates:
            body = (
                "🟡 Park müsaitliği sadece <b>bugün + 7 gün</b> arasında sorgulanabilir.<br>"
                "Bu aralıkta uygun bir gün bulunamadı."
            )
            return f"{intro}<br>{body}"

        tarih_list = filtered_dates

        if not tarih_list:
            body = "🚫 Hafta sonları için park müsaitlik sorgulanamaz."
            return f"{intro}<br>{body}"

        #   DB bağlantısı
        try:
            conn = pyodbc.connect(YA_2El_AracSatis)
            cur = conn.cursor()
        except:
            body = "⚠️ Şu anda park bilgilerine ulaşılamıyor."
            return f"{intro}<br>{body}"

        cevaplar = []

        # --------------------------------------
        # 🔹 4) Park durumunu hesapla
        # --------------------------------------
        for t in tarih_list:
            gun_str = self.format_date_with_day(t)

            cur.execute("""
                SELECT rezerv_park_no
                FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                WHERE rezerv_tarih=? AND is_active=1
            """, (t,))
            dolu_olanlar = {int(r[0]) for r in cur.fetchall()}

            all_spots = list(range(10, 44))
            boslar = [p for p in all_spots if p not in dolu_olanlar]

            if not boslar:
                cevaplar.append(f"🚫 {gun_str} → Uygun park bulunamadı.")
                continue

            ilk_on = ", ".join(map(str, boslar[:10])) + ("..." if len(boslar) > 10 else "")

            cevaplar.append(
                f"🅿️ {gun_str} → <b>{len(boslar)}</b> uygun yer var.<br>"
                f"Müsait yerler: {ilk_on}"
            )

        conn.close()

        body = "<br><br>".join(cevaplar)

        # ❗ ARTIK LLM OUTRO YOK — SADECE GERÇEK SONUÇ
        return f"{intro}<br>{body}"


    def handle_otopark_status_user(self, user_message: str, user_info: dict):
        print("📘 STATUS_USER (Kullanıcı Rezervasyonları) çalışıyor...")

        username = user_info.get("fullname")
        if not username:
            return "🚫 Kullanıcı bilgisi bulunamadı."

        # --------------------------------------
        # 🔹 Intro — kullanıcıyı karşılayan cümle
        # --------------------------------------
        intro = self.naturalize_intro("otopark_status_user")

        # --------------------------------------
        # 🔹 Tarih çözümleme
        # --------------------------------------
        parsed = self.otopark_parse_tarih(user_message)

        print("📅 STATUS_USER PARSE:", parsed)

        tarih_list = parsed.get("dates", [])

        today = datetime.now().date()

        if not tarih_list:
            today = datetime.now().date()
            tarih_list = [today + timedelta(days=i) for i in range(7)]

        # --------------------------------------
        # 🔹 DB bağlantısı
        # --------------------------------------
        try:
            conn = pyodbc.connect(YA_2El_AracSatis)
            cur = conn.cursor()
        except:
            body = "⚠️ Şu anda rezervasyon bilgilerine ulaşılamıyor."
            return f"{intro}<br>{body}"

        cevaplar = []

        # --------------------------------------
        # 🔹 Kullanıcının her gün için kaydını kontrol et
        # --------------------------------------
        for t in tarih_list:
            gun_str = self.format_date_with_day(t)

            cur.execute("""
                SELECT rezerv_park_no
                FROM [Yuce_PortalTest].[dbo].[YA_otopark_TEST]
                WHERE username=? AND rezerv_tarih=? AND is_active=1
            """, (username, t))

            row = cur.fetchone()

            if row:
                park_no = int(row[0])
                cevaplar.append(
                    f"♻️ {gun_str} → {park_no} numaralı parkınız aktif görünüyor."
                )
        conn.close()
        #  HİÇ AKTİF REZERVASYON YOKSA
        if not cevaplar:
            if "yarın" in user_message.lower():
                body = "📭 Yarın için aktif bir otopark rezervasyonunuz bulunmuyor."
            elif "bugün" in user_message.lower():
                body = "📭 Bugün için aktif bir otopark rezervasyonunuz bulunmuyor."
            else:
                body = "📭 Belirtilen tarihlerde aktif bir otopark rezervasyonunuz bulunmuyor."
        else:
            body = "<br>".join(cevaplar)
        # 🔹 Final mesaj
        final = f"{intro}<br>{body}"

        return final
 
    def parse_workday_assignments(self, text: str):
        """
        Kullanıcı cümlesini analiz eder ve
        [(date, work_type), ...] döndürür.

        TÜM ZEKA BURADADIR.
        handle_haftalik_takvim sadece uygular.
        """
        print("\n🧠 [STATE PARSER INPUT]:", text)
        today = datetime.now().date()

        # -------------------------------------------------
        #  NORMALIZE
        norm = self.normalize_calendar_text(text)
        tokens = norm.split()

        # -------------------------------------------------
        # 2️⃣ SABİTLER
        # -------------------------------------------------
        WEEKDAYS = {
            "pazartesi": 0,
            "salı": 1,
            "çarşamba": 2,
            "perşembe": 3,
            "cuma": 4,
        }
        RELATIVE_DAYS = {
            "bugun": 0,
            "yarin": 1,
            "obur gun": 2,
            "ertesi gun": 2,
        }
        OTHER_DAY_KEYS = {
            "diger", "diğer", "diger gunler", "diğer günler",
            "kalan", "kalan gunler", "kalan günler",
            "geri kalan", "geri kalan gunler", "geri kalan günler","dger","kln günler"
        }

        # -------------------------------------------------
        # 3️⃣ HANGİ HAFTA?
        # -------------------------------------------------
        is_next_week = any(k in norm for k in NEXT_WEEK_KEYS)
        is_this_week = not is_next_week  # default

        if is_next_week:
            base_monday = today + timedelta(days=(7 - today.weekday()))
        else:
            base_monday = today - timedelta(days=today.weekday())

        # bu haftada bugünden önceki günleri alma
        def valid_week_days():
            return [
                base_monday + timedelta(days=i)
                for i in range(5)
            ]

        week_days = valid_week_days()

        # -------------------------------------------------
        # 4️⃣ STATE DEĞİŞKENLERİ
        # -------------------------------------------------
        current_type = None            # aktif çalışma tipi
        pending_days = []              # tipi henüz gelmemiş günler (index)
        assigned = {}                  # date -> work_type
        other_days_requested = False   # "diğer günler" flag
        seen_work_types = set()

        # -------------------------------------------------
        # 5️⃣ TOKEN TOKEN OKUMA (ASIL MANTIK)
        # -------------------------------------------------
        i = 0
        while i < len(tokens):

            tok = tokens[i]
            tok = re.sub(r"[^\wçğıöşü]", "", tok).lower()  
            print(f"🔹 tok='{tok}' | current_type={current_type} | pending={pending_days}")

            # ---- göreli günler: bugün / yarın / öbür gün ----
            window2 = " ".join(tokens[i:i+2])
            
            if tok in RELATIVE_DAYS or window2 in RELATIVE_DAYS:
                offset = RELATIVE_DAYS.get(tok) if tok in RELATIVE_DAYS else RELATIVE_DAYS.get(window2)
                d = today + timedelta(days=offset)

                if current_type:
                    assigned[d] = current_type
                    print(f"✅ ASSIGN {d} → {current_type}")

                else:
                    pending_days.append(("absolute", d))

                i += 1
                continue

            # ---- "diğer günler" ----
            window2 = " ".join(tokens[i:i+2])
            if tok in OTHER_DAY_KEYS or window2 in OTHER_DAY_KEYS:
                other_days_requested = True
                i += 1
                continue

            # ---- AMA / FAKAT / ANCAK = KESİN BLOK AYIRICI ----
            if tok in {"ama", "fakat", "ancak"}:
                current_type = None
                pending_days.clear()
                i += 1
                continue

            # ---- çalışma tipi ----
            window2 = " ".join(tokens[i:i+2])
            wt = self.detect_work_type(window2) or self.detect_work_type(tok)


            if wt:
                seen_work_types.add(wt)

                #  pending varsa HEPSİNİ bu tipe bağla
                if pending_days:
                    for item in pending_days:
                        if isinstance(item, tuple) and item[0] == "absolute":
                            d = item[1]
                        else:
                            d = base_monday + timedelta(days=item)

                        if d.weekday() < 5:
                            assigned[d] = wt
                            print(f"✅ ASSIGN {d} → {wt}")

                    pending_days.clear()

                current_type = wt
                i += 1
                continue

            # ---- gün ----
            if tok in WEEKDAYS:
                day_idx = WEEKDAYS[tok]

                candidate = base_monday + timedelta(days=day_idx)

                # Eğer kullanıcı haftaya demediyse ve gün geçmişse
                # haftayı komple kaydır
                if not is_next_week and candidate < today:
                    base_monday += timedelta(days=7)
                    week_days = [
                        base_monday + timedelta(days=i)
                        for i in range(5)
                    ]
                    candidate = base_monday + timedelta(days=day_idx)

                d = candidate
                    
                if d.weekday() >= 5:
                    i += 1
                    continue

                #  ASLA hemen atama yapma
                pending_days.append(day_idx)

                i += 1
                continue

            i += 1

        # PENDING KALAN GÜNLER (sona kalanlar)
        if pending_days and current_type:
            for item in pending_days:
                # bugün / yarın gibi absolute günler
                if isinstance(item, tuple) and item[0] == "absolute":
                    d = item[1]
                    if d not in assigned:
                        assigned[d] = current_type
                else:
                    # pazartesi–cuma gibi indexli günler
                    day_idx = item
                    d = base_monday + timedelta(days=day_idx)
                    if d in week_days and d not in assigned:
                        assigned[d] = current_type

        # -------------------------------------------------
        # 8️⃣ DİĞER GÜNLER MANTIĞI
        # -------------------------------------------------
        if other_days_requested and current_type:
            for d in week_days:
                if d not in assigned:
                    assigned[d] = current_type


        print("🧠 [STATE PARSER RESULT]:")
        for d, t in assigned.items():
            print(f"   {d} → {t}")

        #  ÇIKTI
        #  Hiçbir gün atanmadıysa → bu parser bu cümleye ait değil
        if not assigned:
            return []

        return [(d, assigned[d]) for d in sorted(assigned.keys())]


    def parse_takvim_tarih_araligi(self, message: str):
        """
        Haftalık takvim için doğal dilde geçen 'bu hafta', 'haftaya', 'hafta içi', 'hafta sonu',
        'bu haftayı full işaretle', 'haftaya hafta sonu' vb. ifadeleri tarih listesine çevirir.
        Geçmiş günleri atlar, gerekiyorsa uyarı mesajı döndürür.
        """
        print("🔍 [DEBUG] parse_takvim_tarih_araligi() ÇALIŞTI → Mesaj:", message)

        msg_lower = (message or "").lower().strip()
        today = datetime.now().date()
        year = today.year
        #  HER ZAMAN VAR olmalı
        found_days = []

        MONTHS = {
            "ocak":1,"şubat":2,"mart":3,"nisan":4,"mayıs":5,
            "haziran":6,"temmuz":7,"ağustos":8,
            "eylül":9,"ekim":10,"kasım":11,"aralık":12
        }
        # 🔹 Bugün / Yarın ifadeleri
        if re.search(r"\bbug[uü]n(e|ü|u|un|ün)?\b", msg_lower):
            return {"dates": [today], "warning": None}

        if re.search(r"\byar[iı]n(a|e|ı|i|un|ün)?\b", msg_lower):
            yarin = today + timedelta(days=1)
            if yarin.weekday() < 5:
                return {"dates": [yarin], "warning": None}
            else:
                return {"dates": [], "warning": "🚫 Yarın hafta sonu, işaretlenmedi."}
    
        #  EN ÖNCELİKLİ: HAFTAYA BUGÜN (SADECE TEK GÜN İSE)
        if any(k in msg_lower for k in [
            "haftaya bugün",
            "haftaya bugun",
            "haftaya bugünü",
            "haftaya bugunu","haftya bugn","haftaya bgn","önümüzdeki hafta bugün","önüzümüzde ki hafta bugün","haftaya bu gün","ileriki hafta bugün","gelecek hafta bu gün","gelecek hafta bugün"
        ]):
            # başka gün var mı?
            other_days = found_days.copy()

            # "bugün" canonical bir gün değil, onu hariç tut
            # found_days sadece pazartesi-salı vb. içerir
            if not other_days:
                t = today + timedelta(days=7)
                if t.weekday() >= 5:
                    return {
                        "dates": [],
                        "warning": "🚫 Haftaya bugün hafta sonuna denk geliyor."
                    }
                return {"dates": [t], "warning": None}
            
        VARIANT_TO_CANONICAL_DAY = {}

        for d in PAZARTESI_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "pazartesi"
        for d in SALI_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "salı"
        for d in CARSAMBA_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "çarşamba"
        for d in PERSEMBE_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "perşembe"
        for d in CUMA_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "cuma"
        for d in CUMARTESI_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "cumartesi"
        for d in PAZAR_VARIANTS:
            VARIANT_TO_CANONICAL_DAY[d] = "pazar"

        CANONICAL_DAY_TO_INDEX = {
            "pazartesi": 0,
            "salı": 1,
            "çarşamba": 2,
            "perşembe": 3,
            "cuma": 4,
            "cumartesi": 5,
            "pazar": 6,
        }


        for day in ALL_WEEKDAY_VARIANTS:
            if day in msg_lower:
                canonical = VARIANT_TO_CANONICAL_DAY.get(day)
                if canonical:
                    found_days.append(canonical)

        # tekrarları temizle
        found_days = list(dict.fromkeys(found_days))

            
        has_context = any(
            k in msg_lower
            for k in (BU_WEEK_DAY_PREFIXES + NEXT_WEEK_KEYS + THIS_MONTH_KEYS)
        )
        if found_days and not has_context:

            monday = today - timedelta(days=today.weekday())
            tarih_listesi = []

            for g in found_days:
                idx = CANONICAL_DAY_TO_INDEX.get(g)
                if idx is None:
                    continue

                t = monday + timedelta(days=idx)

                # bu hafta geçtiyse → haftaya kaydır
                if t < today:
                    t += timedelta(days=7)

                if t.weekday() < 5:
                    tarih_listesi.append(t)

            return {"dates": tarih_listesi, "warning": None}
        
        
        #  AY VAR AMA GÜN YOKSA → AYIN TÜM İŞ GÜNLERİ
        found_months = [
            ay for ay in MONTHS
            if re.search(rf"\b{ay}\b", msg_lower)
        ]

        has_day_info = bool(found_days) or bool(re.search(r"\d{1,2}", msg_lower))


        if found_months and not has_day_info:
            dates = []
            year_match = re.search(r"\b(20\d{2})\b", msg_lower)
            yil = int(year_match.group(1)) if year_match else today.year

            for ay in found_months:
                ay_num = MONTHS[ay]

                # aynı ay ise bugünden başla
                if yil == today.year and ay_num == today.month:
                    start_date = today
                else:
                    start_date = date(yil, ay_num, 1)

                end_date = date(
                    yil,
                    ay_num,
                    calendar.monthrange(yil, ay_num)[1]
                )

                cur = start_date
                while cur <= end_date:
                    if cur.weekday() < 5:  # sadece hafta içi
                        dates.append(cur)
                    cur += timedelta(days=1)

            return {"dates": sorted(set(dates)), "warning": None}

        for ay_adi, ay_num in MONTHS.items():

            if ay_adi in msg_lower and any(k in msg_lower for k in FULL_MONTH_KEYS):

                # yıl
                year_match = re.search(r"\b(20\d{2})\b", msg_lower)
                yil = int(year_match.group(1)) if year_match else year

                # başlangıç
                if yil == today.year and ay_num == today.month:
                    start_date = today
                else:
                    start_date = datetime(yil, ay_num, 1).date()

                # bitiş
                end_date = datetime(
                    yil,
                    ay_num,
                    calendar.monthrange(yil, ay_num)[1]
                ).date()

                dates = []
                cur = start_date
                while cur <= end_date:
                    if cur.weekday() < 5:   # hafta sonu yok
                        dates.append(cur)
                    cur += timedelta(days=1)

                print(f"📅 AY FULL ALGILANDI → {ay_adi} {yil}")
                return {"dates": dates, "warning": None}
            
        # BİRDEN FAZLA AY + TAMAMI / FULL / KOMPLE 
        ay_full_triggers = ["full", "komple", "tamami", "tamamı", "tümü", "hepsi", "ful", "boyunca"]

        all_dates = []
        matched_any_month = False

        # YIL TESPİTİ (tek kez)
        year_match = re.search(r"\b(20\d{2})\b", msg_lower)
        yil = int(year_match.group(1)) if year_match else year

        for ay_adi, ay_num in MONTHS.items():
            if ay_adi in msg_lower and any(k in msg_lower for k in ay_full_triggers):
                matched_any_month = True

                #  AY BAŞLANGICI
                if yil == today.year and ay_num == today.month:
                    start_date = today
                else:
                    start_date = datetime(yil, ay_num, 1).date()

                #  AY SONU
                end_date = datetime(
                    yil,
                    ay_num,
                    calendar.monthrange(yil, ay_num)[1]
                ).date()

                cur = start_date
                while cur <= end_date:
                    if cur.weekday() < 5:  # 🚫 hafta sonu yok
                        all_dates.append(cur)
                    cur += timedelta(days=1)

                print(f"📅 AY FULL ALGILANDI → {ay_adi} {yil}")

        #  EN SONDA TEK RETURN
        if matched_any_month:
            all_dates = sorted(set(all_dates))  # duplicate temizle
            return {"dates": all_dates, "warning": None}

        # 🔹 "5 ekim ve 10 ekim" gibi, sadece iki farklı günü ifade eden durumları yakala
        if (" ve " in msg_lower or "," in msg_lower) and not any(x in msg_lower for x in ["arası", "arasında", "kadar", "ile", "–", "-"]):
            match_coklu = re.findall(
                r"(\d{1,2})\s+(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)?",
                msg_lower
            )
            if match_coklu and len(match_coklu) >= 2:
                tarih_listesi = []
                gecmis_tarihler = []
                for gun, ay in match_coklu:
                    try:
                        gun = int(gun)
                        ay_num = MONTHS.get(ay, today.month)
                        yil = year
                        tarih = datetime(yil, ay_num, gun).date()
                        if tarih.weekday() < 5:  # sadece hafta içi
                            if tarih < today:
                                gecmis_tarihler.append(tarih)
                            else:
                                tarih_listesi.append(tarih)
                    except Exception as e:
                        print(f"⚠️ Çoklu tarih çözüm hatası: {e}")
                        continue

                # 🎯 Kullanıcıya hem geçmiş hem güncel tarihleri bildir
                if gecmis_tarihler or tarih_listesi:
                    lines = []
                    for t in tarih_listesi:
                        lines.append(f"✅ {t.strftime('%d.%m.%Y (%a)')} → Güncellendi.")
                    warning_text = "📅 Takvim işlemi tamamlandı:<br>" + "<br>".join(lines)
                    print(warning_text)
                    return {"dates": tarih_listesi, "warning": warning_text}


        # 🔹 "5 kasım 11 kasım 15 kasım" gibi birden fazla ay adı içeren tarihleri yakala
        if not any(x in msg_lower for x in ["arası", "arasında", "kadar", "ile", "–", "-"]):
            # Tüm sayı + ay çiftlerini sırayla yakala
            month_pattern = "|".join(MONTHS.keys())
            match_multi_ay = re.findall(
                rf"(\d{{1,2}})\s+({month_pattern})",
                msg_lower
            )

            if len(match_multi_ay) >=2:
                tarih_listesi = []
                gecmis_tarihler = []
                for gun, ay in match_multi_ay:
                    try:
                        gun = int(gun)
                        ay_num = MONTHS[ay]
                        yil = year
                        tarih = datetime(yil, ay_num, gun).date()
                        if tarih.weekday() < 5:
                            if tarih < today:
                                gecmis_tarihler.append(tarih)
                            else:
                                tarih_listesi.append(tarih)
                    except Exception as e:
                        print(f"⚠️ Çoklu ay çözüm hatası: {e}")
                        continue

                if tarih_listesi:
                    lines = [f"✅ {t.strftime('%d.%m.%Y (%a)')} → Güncellendi." for t in tarih_listesi]
                    warning_text = "📅 Takvim işlemi tamamlandı:<br>" + "<br>".join(lines)
                    return {"dates": tarih_listesi, "warning": warning_text}

                if gecmis_tarihler and not tarih_listesi:
                    return {
                        "dates": [],
                        "warning": "🚫 Belirttiğiniz tarihler geçmişte, işaretlenemez."
                    }

        # ===============================================================
        # 🧩 TÜRKÇE TAM DESTEKLİ TARİH ARALIĞI ALGILAMA (örnek: 5 kasımdan 15 kasıma kadar)
        # ===============================================================
        match_aralik_yeni = re.search(
            r"(?P<g1>\d{1,2})"                                # ilk gün (ör: 5)
            r"\s*[,\.']*\s*"                                   # olası ayırıcılar (, . veya ')
            r"(?P<ay1>ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)?"  # ilk ay (opsiyonel)
            r"(?:['’]?(?:dan|den|tan|ten|a|e|ya|ye|ın|in|un|ün))?"  # Türkçe ekler (ör: kasımdan, kasıma)
            r"\s*(?:-|–|,|ile|ve|’|’dan|’den|’a|’e|’ya|’ye|’dan|’den|’tan|’ten|\s)+\s*"  # bağlaçlar veya tire
            r"(?P<g2>\d{1,2})"                                # ikinci gün (ör: 15)
            r"\s*[,\.']*\s*"                                   # ayırıcılar
            r"(?P<ay2>ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)?"  # ikinci ay (opsiyonel)
            r"(?:['’]?(?:a|e|ya|ye|dan|den|tan|ten|kadar|arası|arasında)?)?",            # ikinci kelime ekleri
            msg_lower,
            flags=re.IGNORECASE  # büyük/küçük harf duyarsız
        )


        if match_aralik_yeni:
            try:
                g1 = int(match_aralik_yeni.group("g1"))
                g2 = int(match_aralik_yeni.group("g2"))
                ay1 = match_aralik_yeni.group("ay1") or match_aralik_yeni.group("ay2") or None
                month = MONTHS.get(ay1, today.month)
                start_date = datetime(year, month, g1).date()
                end_date = datetime(year, month, g2).date()

                if end_date < start_date:
                    end_date = start_date  # yanlış sıra yazılmışsa düzelt

                tarih_listesi = []
                for d in range((end_date - start_date).days + 1):
                    t = start_date + timedelta(days=d)
                    if t.weekday() < 5:
                        tarih_listesi.append(t)

                # 🔹 Geçmiş / Gelecek / Limit Kontrolü (iyileştirilmiş versiyon)
                valid_dates = []
                message_lines = []
                max_end = today + timedelta(days=30)

                for t in tarih_listesi:
                    if t < today:
                        message_lines.append(f"❌ {t.strftime('%d.%m.%Y (%A)')} → Geçmiş tarih, işaretlenemedi.")
                        continue
                    if t > max_end:
                        message_lines.append(f"⚠️ {t.strftime('%d.%m.%Y (%A)')} → 30 gün sınırını aşıyor, eklenmedi.")
                        continue
                    if t.weekday() >= 5:
                        message_lines.append(f"🚫 {t.strftime('%d.%m.%Y (%A)')} → Hafta sonu, eklenmedi.")
                        continue

                    valid_dates.append(t)
                    message_lines.append(f"✅ {t.strftime('%d.%m.%Y (%A)')} → Güncellendi.")

                # 🔸 Hiç geçerli tarih yoksa tamamen iptal
                if not valid_dates:
                    warning_text = "📅 Takvim işlemi tamamlanamadı:<br>" + "<br>".join(message_lines)
                    print(warning_text)
                    return {"dates": [], "warning": warning_text}

                # 🔸 Hem geçmiş hem geçerli tarih varsa (kısmi başarı)
                if len(valid_dates) < len(tarih_listesi):
                    warning_text = "📅 Takvim işlemi kısmen tamamlandı:<br>" + "<br>".join(message_lines)
                else:
                    warning_text = "📅 Takvim işlemi tamamlandı:<br>" + "<br>".join(message_lines)

                print(warning_text)
                return {"dates": valid_dates, "warning": None}


            except Exception as e:
                print(f"⚠️ Yeni aralık çözüm hatası: {e}")


        #  HAFTAYA + SAYI (haftaya 5 gün, ilk 3 gün vb.)
        count = self._detect_weekday_count_scope(msg_lower)

        if count and any(k in msg_lower for k in NEXT_WEEK_KEYS):
            next_monday = today + timedelta(days=(7 - today.weekday()))

            # maksimum 5 iş günü
            count = min(count, 5)

            dates = [
                next_monday + timedelta(days=i)
                for i in range(count)
                if (next_monday + timedelta(days=i)).weekday() < 5
            ]

            return {"dates": dates, "warning": None}
        
        norm = BrainAIYugii().normalize_calendar_text(msg_lower)

        if any(k in norm for k in NEXT_WEEK_KEYS):

            next_monday = today + timedelta(days=(7 - today.weekday()))

            # 1️⃣ haftaya + full / hafta içi
            if any(k in norm for k in FULL_KEYS):
                dates = [next_monday + timedelta(days=i) for i in range(5)]
                return {"dates": dates, "warning": None}

            # 2️⃣ haftaya + BELİRLİ GÜNLER (canonical)
            if found_days:
                dates = []
                for g in found_days:              # g = "salı", "çarşamba"
                    idx = CANONICAL_DAY_TO_INDEX.get(g)
                    if idx is None:
                        continue
                    dates.append(next_monday + timedelta(days=idx))

                return {"dates": dates, "warning": None}

            # 3️⃣ sadece "haftaya"
            return {
                "dates": [next_monday + timedelta(days=i) for i in range(5)],
                "warning": None
            }

        # ===============================================================
        # 🔹 Çoklu tekil tarih desteği (örnek: 11 kasım 12 kasım 13 kasım)
        # ===============================================================
        match_coklu_tekil = re.findall(
            r"(\d{1,2})\s*(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)?",
            msg_lower
        )
        if match_coklu_tekil and len(match_coklu_tekil) >= 2 and any(ay for _, ay in match_coklu_tekil):
            try:
                tarih_listesi = []
                for gun, ay in match_coklu_tekil:
                    gun = int(gun)
                    ay_num = MONTHS.get(ay, today.month)
                    t = datetime(year, ay_num, gun).date()
                    if t.weekday() < 5:
                        tarih_listesi.append(t)

                valid_dates = [t for t in tarih_listesi if today <= t <= today + timedelta(days=30)]
                if not valid_dates:
                    warning = "🚫 Belirttiğiniz tarihler geçmişte veya 30 günü aşıyor."
                    print(warning)
                    return {"dates": [], "warning": warning}

                lines = [f"✅ {t.strftime('%d.%m.%Y (%A)')} → Güncellendi." for t in valid_dates]
                warning_text = "📅 Takvim işlemi tamamlandı:<br>" + "<br>".join(lines)
                print(warning_text)
                return {"dates": valid_dates, "warning": warning_text}
            except Exception as e:
                print(f"⚠️ Çoklu tekil tarih çözüm hatası: {e}")

        # 🔄 Mesaj içindeki hatalı yazılmış MONTHSı düzelt
        for k, v in AY_NORMALIZE.items():
            msg_lower = re.sub(rf"\b{k}\b", v, msg_lower)


        # 📅 Tarih aralığı tespiti (örnek: 30 ekim - 5 kasım arası)
        match_aralik = re.search(
            r"(?P<g1>\d{1,2})\s*(?P<ay1>ocak|şubat|mart|nisan|mayıs|haziran|temmuz|"
            r"ağustos|eylül|ekim|kasım|aralık)?"
            r".*?(?:-|–|ile|ve|,|'?(?:den|dan|ten|tan)|'?(?:a|e|ye|ya)|kadar|dek)\s*"
            r"(?P<g2>\d{1,2})\s*(?P<ay2>ocak|şubat|mart|nisan|mayıs|haziran|temmuz|"
            r"ağustos|eylül|ekim|kasım|aralık)?"
            r".*?(arası|arasında|kadar|dek)?",
            msg_lower
        )
        if match_aralik:
            try:
                g1 = int(match_aralik.group("g1"))
                g2 = int(match_aralik.group("g2"))
                ay1 = match_aralik.group("ay1")
                ay2 = match_aralik.group("ay2")

                # Ay bilgilerini çöz (tek ay veya iki farklı ay)
                if ay1 and ay2:
                    start_month, end_month = MONTHS[ay1], MONTHS[ay2]
                elif ay1 and not ay2:
                    start_month = end_month = MONTHS[ay1]
                elif not ay1 and ay2:
                    start_month = end_month = MONTHS[ay2]
                else:
                    start_month = end_month = today.month  # ay belirtilmemişse mevcut ay

                start_date = datetime(year, start_month, g1).date()
                end_date = datetime(year, end_month, g2).date()

                # Eğer bitiş tarihi başlangıçtan küçükse (ör. 30 kasım - 5 aralık)
                if end_date < start_date:
                    end_date = datetime(year + 1, end_month, g2).date()

                warning = None
                max_end = today + timedelta(days=30)

                #  Tarihleri oluştur (sadece hafta içi)
                tarih_listesi = []
                current = start_date
                while current <= end_date:
                    if current.weekday() < 5:
                        tarih_listesi.append(current)
                    current += timedelta(days=1)

                # Geçmiş / Gelecek / Limit Kontrolü
                # Geçmiş / Gelecek / Limit Kontrolü (iyileştirilmiş versiyon)
                valid_dates = []
                message_lines = []
                max_end = today + timedelta(days=30)

                for t in tarih_listesi:
                    if t < today:
                        message_lines.append(f"❌ {t.strftime('%d.%m.%Y (%A)')} → Geçmiş tarih, işaretlenemedi.")
                        continue
                    if t > max_end:
                        message_lines.append(f"⚠️ {t.strftime('%d.%m.%Y (%A)')} → 30 gün sınırını aşıyor, eklenmedi.")
                        continue
                    if t.weekday() >= 5:
                        message_lines.append(f"🚫 {t.strftime('%d.%m.%Y (%A)')} → Hafta sonu, eklenmedi.")
                        continue

                    valid_dates.append(t)
                    message_lines.append(f"✅ {t.strftime('%d.%m.%Y (%A)')} → Güncellendi.")

                #  Hiç geçerli tarih yoksa tamamen iptal
                if not valid_dates:
                    warning_text = "📅 Takvim işlemi tamamlanamadı:<br>" + "<br>".join(message_lines)
                    print(warning_text)
                    return {"dates": [], "warning": warning_text}

                #  Hem geçmiş hem geçerli tarih varsa (kısmi başarı)
                if len(valid_dates) < len(tarih_listesi):
                    warning_text = "📅 Takvim işlemi kısmen tamamlandı:<br>" + "<br>".join(message_lines)
                else:
                    warning_text = "📅 Takvim işlemi tamamlandı:<br>" + "<br>".join(message_lines)

                print(warning_text)
                return {"dates": valid_dates, "warning": warning_text}

            except Exception as e:
                print(f"⚠️ Aralık çözüm hatası: {e}")

        month_pattern = "|".join(MONTHS.keys())

        match_tek_tarih = re.search(
            rf"(?P<gun>\d{{1,2}})\s+(?P<ay>{month_pattern})(?:\s+(?P<yil>20\d{{2}}))?",
            msg_lower
        )

        if match_tek_tarih:
            gun = int(match_tek_tarih.group("gun"))
            ay = match_tek_tarih.group("ay")
            ay_num = MONTHS[ay]

            yil = int(match_tek_tarih.group("yil")) if match_tek_tarih.group("yil") else today.year

            tarih = datetime(yil, ay_num, gun).date()

            if tarih < today:
                return {
                    "dates": [],
                    "warning": f"❌ {tarih.strftime('%d.%m.%Y (%A)')} geçmiş tarih, işaretlenemez."
                }

            return {"dates": [tarih], "warning": None}


        # 🧩 HAFTAYA / GELECEK / ÖNÜMÜZDEKİ HAFTA DESTEK BLOĞU (BURAYA EKLENİYOR!)
        weekdays = {
            "pazartesi": 0,
            "salı": 1,
            "çarşamba": 2,
            "perşembe": 3,
            "cuma": 4,
            "cumartesi": 5,
            "pazar": 6
        }


        # 🔹 Bu haftanın ve gelecek haftanın başlangıç tarihleri
        monday = today - timedelta(days=today.weekday())
        next_monday = monday + timedelta(days=7)

        if found_days:
            tarih_listesi = []

            norm = self.normalize_calendar_text(msg_lower)


            is_next_week = any(k in norm for k in NEXT_WEEK_KEYS)
            is_this_week = "bu hafta" in norm or "bu pazartesi" in norm

            # anchor pazartesi
            if is_next_week:
                base_monday = today - timedelta(days=today.weekday()) + timedelta(days=7)
            else:
                base_monday = today - timedelta(days=today.weekday())

            # 🔹 "bu pazartesi", "bu hafta pazartesi" ifadelerinde geçmiş tarih kontrolü
            if any(
                prefix in norm and any(day in norm for day in PAZARTESI_VARIANTS)
                for prefix in BU_WEEK_DAY_PREFIXES
            ):
                tarih = base_monday
                if tarih < today:
                    return {
                        "dates": [],
                        "warning": f"❌ {tarih.strftime('%d.%m.%Y (%A)')} → Geçmiş tarih (bu pazartesi) işaretlenemedi."
                    }
                
            # 🔹 Her bir günü sırayla işle
            for g in found_days:
                idx = CANONICAL_DAY_TO_INDEX[g]


                # seçilen haftadaki gün
                tarih = base_monday + timedelta(days=idx)

                has_explicit_date = bool(re.search(r"\d", msg_lower))
                has_bu = "bu " in msg_lower
                has_gecen = "geçen" in msg_lower or "onceki" in msg_lower

                # 🔹 Eğer açık tarih veya "bu cuma" gibi net ifade varsa ASLA kaydırma
                if tarih < today:
                    if has_explicit_date or has_bu or has_gecen:
                        return {
                            "dates": [],
                            "warning": f"❌ {tarih.strftime('%d.%m.%Y (%A)')} geçmiş tarih, işaretlenemez."
                        }
                    else:
                        # sadece çıplak gün adıysa ileri kaydır
                        tarih += timedelta(days=7)

                if tarih.weekday() < 5:
                    tarih_listesi.append(tarih)

            if tarih_listesi:
                print(f"📅 Gün(ler) algılandı → {tarih_listesi}")
                return {"dates": tarih_listesi, "warning": None}


        if any(x in msg_lower for x in ["boyunca", "tüm ay", "bu ay", "gelecek ay", "önümüzdeki ay"]):
            if "gelecek" in msg_lower or "önümüzdeki" in msg_lower:
                ay = today.month + 1 if today.month < 12 else 1
                year = year if today.month < 12 else year + 1
            else:
                ay = today.month

            last_day = calendar.monthrange(year, ay)[1]
            tarih_listesi = [
                datetime(year, ay, d).date()
                for d in range(1, last_day + 1)
                if datetime(year, ay, d).weekday() < 5
            ][:30]

            return {"dates": tarih_listesi, "warning": None}

        # 🔥 AY İSMİ + BOYUNCA (ör: "aralık boyunca", "mart ayı boyunca", "temmuz komple")
        for ay_adi, ay_num in MONTHS.items():
            if f"{ay_adi} boyunca" in msg_lower or f"{ay_adi} ayı boyunca" in msg_lower or f"{ay_adi} komple" in msg_lower:
                
                last_day = calendar.monthrange(year, ay_num)[1]
                max_date = today + timedelta(days=30)

                tarih_listesi = []

                for d in range(1, last_day + 1):
                    t = datetime(year, ay_num, d).date()

                    # Geçmiş gün → at
                    if t < today:
                        continue

                    # Hafta sonu → at
                    if t.weekday() >= 5:
                        continue

                    # 30 gün sonrası → at
                    if t > max_date:
                        continue

                    tarih_listesi.append(t)

                return {"dates": tarih_listesi, "warning": None}

        if "ayın başı" in msg_lower or "ay başı" in msg_lower:
            tarih_listesi = [
                today + timedelta(days=i)
                for i in range(0, 10)
                if (today + timedelta(days=i)).weekday() < 5
            ]
            return {"dates": tarih_listesi, "warning": None}

        if "ayın ortası" in msg_lower or "ay ortası" in msg_lower:
            tarih_listesi = [
                today + timedelta(days=i)
                for i in range(10, 20)
                if (today + timedelta(days=i)).weekday() < 5
            ]
            return {"dates": tarih_listesi, "warning": None}

        if "ayın sonu" in msg_lower or "ay sonu" in msg_lower:
            tarih_listesi = [
                today + timedelta(days=i)
                for i in range(20, 31)
                if (today + timedelta(days=i)).weekday() < 5
            ]
            return {"dates": tarih_listesi, "warning": None}
        
    


        match_weeks = re.search(r"(?:(önümüzdeki|gelecek)\s*)?(\d+)\s*hafta", msg_lower)
        if match_weeks:
            try:
                week_count = int(match_weeks.group(2))
                start = next_monday if any(x in msg_lower for x in ["önümüzdeki", "gelecek"]) else monday
                end = start + timedelta(days=7 * week_count)
                max_end = today + timedelta(days=30)
                if end > max_end:
                    end = max_end
                tarih_listesi = []
                current = start
                while current < end:
                    if current.weekday() < 5:
                        tarih_listesi.append(current)
                    current += timedelta(days=1)
                return {"dates": tarih_listesi, "warning": None}
            except Exception as e:
                print(f"⚠️ 'Önümüzdeki X hafta' çözüm hatası: {e}")

        izin_durumu = any(x in msg_lower for x in ["izin", "tatil", "izinliyim", "izinli"])

        norm = self.normalize_calendar_text(msg_lower)


        is_this_week = any(k in norm for k in BU_WEEK_DAY_PREFIXES)
        is_next_week = any(k in norm for k in NEXT_WEEK_KEYS)

        if is_this_week and not is_next_week:
            monday = today - timedelta(days=today.weekday())
            week_days = [monday + timedelta(days=i) for i in range(5)]
            week_days = [d for d in week_days if d >= today]
            return {"dates": week_days, "warning": None}
        
        #(TEKİL / TOPLU AMA SADECE HAFTA SONU KELİMESİ VARSA)
        HAS_WEEKEND_KEYWORD = any(
            k in msg_lower
            for k in ["cumartesi", "pazar", "hafta sonu", "haftasonu","cmrtesi","pzr","pazr","hfta sonu","cmrt","cumartesi"]
        )

        # Daha önce hiç tarih bulunamadıysa ve hafta sonu deniyorsa
        if HAS_WEEKEND_KEYWORD:
            base_monday = (
                today + timedelta(days=(7 - today.weekday()))
                if any(k in msg_lower for k in NEXT_WEEK_KEYS)
                else today - timedelta(days=today.weekday())
            )

            weekend_dates = [
                base_monday + timedelta(days=5),  # Cumartesi
                base_monday + timedelta(days=6),  # Pazar
            ]

            return {"dates": weekend_dates, "warning": None}
        
        if (
            "tüm hafta" in msg_lower or "tum hafta" in msg_lower
            or "komple" in msg_lower
            or "full" in msg_lower
            or "haftanın tamamı" in msg_lower or "haftanin tamami" in msg_lower
            or "tamamı" in msg_lower or "tamami" in msg_lower
            or "hepsi" in msg_lower
            or "hafta içi" in msg_lower or "hafta içini" in msg_lower or "hafta ici" in msg_lower
        ):
            # Bu haftanın pazartesisi
            monday = today - timedelta(days=today.weekday())

            # Eğer bugün hafta sonu (Cumartesi–Pazar) ise → gelecek haftaya başla
            if today.weekday() >= 5:
                monday = monday + timedelta(days=7)

            # Eğer kullanıcı "haftaya" dediyse → gelecek hafta
            if any(x in msg_lower for x in ["haftaya", "gelecek", "önümüzdeki"]):
                monday = monday + timedelta(days=7)

            # Pazartesi–Cuma arasındaki günler
            week_days = [monday + timedelta(days=i) for i in range(5)]

            #  Geçmiş günleri AT
            week_days = [d for d in week_days if d >= today]

            #  Hafta sonu zaten yok (pazartesi-cuma) ama yine de filtre:
            week_days = [d for d in week_days if d.weekday() < 5]

            return {"dates": week_days, "warning": None}

    
    def handle_haftalik_takvim(self, user_message: str, user_info: dict):
        print("🗓️ HAFTALIK TAKVİM MODÜLÜ ÇALIŞIYOR...")

        username = user_info.get("fullname")
        user_id = user_info.get("id")
        departman = user_info.get("departman")
        
        IGNORE_SIGNALS = True
        #  users1_id güvenlik kontrolü
        if not user_id:
            # username üzerinden users1 tablosundan id çek
            conn_tmp = _mysql_conn()
            if conn_tmp:
                cur_tmp = conn_tmp.cursor(dictionary=True)
                cur_tmp.execute(
                    "SELECT id FROM users1 WHERE username = %s LIMIT 1",
                    (username,)
                )
                row = cur_tmp.fetchone()
                cur_tmp.close()
                conn_tmp.close()

                if row:
                    user_id = row["id"]

        if not user_id:

            parsed = {
                "intent": "haftalik_takvim",
                "action": "status_mark",
                "raw_text": user_message,
                "error": "NO_USER_ID"
            }
            return self.llm_weekly_calendar_response(parsed, [])
        
        # 🔹 1) INTRO — işlem başlamadan doğal giriş cümlesi
        intro = self.naturalize_intro("haftalik_takvim")
        outro = self.naturalize_text("", "haftalik_takvim")  # 

    
        # 🔹 2) TAKVİM NORMALIZE 
        calendar_text = self.normalize_calendar_text(user_message)

        # 🔹 3) GPT ile SADECE yazım düzelt -azaltıldı 
        cleaned = self.gpt_fix_calendar_text(calendar_text)
        
        # === AYLIK TEKRAR (AY + GÜN + HER/HEP) KONTROLÜ ===
        norm = self.normalize_calendar_text(cleaned)

        # 🔧 GÜN EKLERİNİ TEMİZLE (cuma'yı → cuma, pazartesi'ne → pazartesi)
        norm = re.sub(
            r"\b(pazartesi|salı|çarşamba|perşembe|cuma)"
            r"(?:'?(yi|yı|yu|yü|ya|ye|da|de|den|ten|nin|nın|nun|nün))\b",
            r"\1",
            norm
        )
        for raw, canonical in AY_NORMALIZE.items():
            norm = re.sub(rf"\b{raw}\b", canonical, norm)


        has_month = (
            any(k in norm for k in THIS_MONTH_KEYS)
            or any(
                re.search(rf"\b{m}\b", norm)
                for m in AY_NORMALIZE.values()
            )
        )

        has_weekday = any(d in norm for d in ALL_WEEKDAY_VARIANTS)
        has_repeat = any(k in norm for k in ["her", "hep"])
        has_full = any(k in norm for k in FULL_MONTH_KEYS )
        default_this_month = has_weekday and has_repeat and not has_month

        assignments = []
        dates = []

        #  ÇALIŞMA TİPİ SAYIMI (SAĞLAM KAYNAK)
        work_types_in_text = set()
        tokens = cleaned.split()

        for i in range(len(tokens)):
            # tek kelime
            wt1 = self.detect_work_type(tokens[i])
            if wt1:
                work_types_in_text.add(wt1)

            # iki kelimelik pencere (home office, evden calisma)
            if i + 1 < len(tokens):
                window2 = tokens[i] + " " + tokens[i + 1]
                wt2 = self.detect_work_type(window2)
                if wt2:
                    work_types_in_text.add(wt2)


        # 🔎 GÖRELİ GÜN VAR MI?
        HAS_RELATIVE_DAY = any(
            k in cleaned for k in ["bugun", "yarin", "obur gun", "ertesi gun"]
        )
        has_weekend_word = any(
            k in cleaned
            for k in ["cumartesi", "pazar", "hafta sonu", "haftasonu","pzr","cmrtes","cumrtesine","pzara","pzara","pazara"]
        )
        # Eğer GPT metni anlam değiştiriyorsa (ör: perşembe → cuma), dolayı güvenmeyelim:
        if "perşembe" in user_message.lower() and "cuma" in cleaned.lower():
            cleaned = user_message  # GPT düzeltmesini iptal et

        assignments = []

        # 1️⃣ Önce numeric / tarih aralığı
        date_result = self.parse_takvim_tarih_araligi(norm)

        if date_result:

            # 1️⃣ Eğer tarih yok ama warning varsa direkt döndür
            if not date_result.get("dates") and date_result.get("warning"):
                return date_result["warning"]

            # 2️⃣ Geçerli tarih varsa normal devam
            if date_result.get("dates"):
                work_type = self.detect_work_type(norm)
                if work_type:
                    assignments = [(d, work_type) for d in date_result["dates"]]

        # 2️⃣ AY KOMPLE
        elif has_month and has_full and not has_weekday:
            dates = self._build_full_month_dates(norm)

        # 3️⃣ AYLIK TEKRAR
        elif (has_month and has_weekday and has_repeat) or default_this_month:
            dates = self._build_monthly_weekday_repeat(norm)

        # 4️⃣ Haftalık / tek gün
        else:
            assignments = self.parse_workday_assignments(norm)

        # === ORTAK ÇIKIŞ (TEK DOĞRU YER) ===
        work_type = self.detect_work_type(norm)

        if not assignments and dates and work_type:
            assignments = [(d, work_type) for d in dates]

        print("🧠 [PARSE RESULT] assignments:")
        # 🔴 GEÇMİŞ TARİH VARSA → TAMAMEN İPTAL
        today = datetime.now().date()

        past_dates = [d for d, _ in assignments if d < today]

        if past_dates:
            return (
                "🚫 Geçmiş tarihler için takvim işaretlemesi yapamıyorum.<br>"
                "Lütfen bugünden sonraki bir tarih seçin."
            )
        for d, t in assignments:
            print(f"   {d} -> {t}")

        if assignments:
            weekday_items = [(d, t) for d, t in assignments if d.weekday() < 5]
            weekend_items = [(d, t) for d, t in assignments if d.weekday() >= 5]

            # sadece hafta sonu varsa → NET HATA
            if not weekday_items and weekend_items:
                return (
                    "🚫 Belirttiğiniz günler hafta sonuna denk geliyor.<br>"
                    "Hafta sonları için haftalık takvim işaretlemesi yapamıyorum , haftalık takvim sayfasından yapabilirsin."
                )
            
            assignments = weekday_items

        date_type_map = {}

        for d, work_type in assignments:
            if d not in date_type_map:
                date_type_map[d] = work_type

        # 🔹 4) Çalışma tipi belirleme (YENİ)
        # -----------------------------------------
        parsed = {
            "intent": "haftalik_takvim",
            "action": "status_mark",
            "raw_text": user_message,
            "items": [
                {"date": d, "status": t}
                for d, t in date_type_map.items()
            ]
        }
        print("🧠 [DB INPUT ITEMS kayıt son parse]:")
        for item in parsed["items"]:
            print(f"   {item['date']} -> {item['status']}")

        if not parsed["items"]:
            return "Hangi günleri işaretlemek istiyorsunuz, gün veya çalışma tipini netleştirir misiniz?"
        
        #  TEK GÜN + HAFTA SONU → NET MESAJ (DOĞRU YER)
        if len(parsed["items"]) == 1:
            d = parsed["items"][0]["date"]
            if d.weekday() >= 5:  # Cumartesi / Pazar
                return (
                    f"🚫 {self.format_date_with_day(d)} hafta sonuna denk geliyor.<br>"
                    "Hafta sonları çalışma günü olarak işaretleyemiyorum."
                )
        #  TÜM GÜNLER HAFTA SONUYSA → NET RED
        if parsed["items"]:
            only_weekend = all(
                item["date"].weekday() >= 5
                for item in parsed["items"]
            )

            if only_weekend:
                return (
                    "🚫 Seçtiğiniz tüm günler hafta sonuna denk geliyor.<br>"
                    "Hafta sonları çalışma günü olarak işaretlenemez."
                )
        # 🔹 5) MySQL Bağlantısı
        conn = _mysql_conn()
        if conn is None:
            body = "⚠️ Takvim DB bağlantısı kurulamadı."
            outro = self.naturalize_text("", "haftalik_takvim")
            final = ""

            if intro:
                final += intro + "<br>"
            final += body
            if outro:
                final += "<br>" + outro

            return final

        cursor = conn.cursor(dictionary=True)

        results = []
        today = datetime.now().date()
        max_date = today + timedelta(days=60)


        # 💾 DB DEBUG LISTELERİ
        saved_dates = []
        updated_dates = []
        skipped_dates = []

        # 🔹 6) Günleri işaretle (UPDATE / INSERT)
        for item in parsed["items"]:


            t = item["date"]
            detected_type = item["status"]

            if t < today:
                skipped_dates.append((t, "GEÇMİŞ TARİH"))
                continue


            if t.weekday() >= 5:
                skipped_dates.append((t, "HAFTA SONU"))
                results.append(f"🚫 {self.format_date_with_day(t)} → Hafta sonu, işaretlenemedi.")
                continue

            if t > max_date:
                skipped_dates.append((t, "60 GÜN SINIRI"))
                results.append(f"⚠️ {self.format_date_with_day(t)} → 60 gün sonrası işaretlenemez.")
                continue

            db_value = detected_type
            if detected_type == "Izin":
                db_value = None

            cursor.execute("""
                SELECT COUNT(*) AS cnt FROM ofis_gunleri
                WHERE users1_id = %s AND selected_date = %s
            """, (user_id, t))
            exists = cursor.fetchone()["cnt"]

            label = (
                "Ofisten Çalışma" if detected_type == "Ofis"
                else "Uzaktan Çalışma" if detected_type == "Home"
                else "Bayi / Müşteri Ziyareti" if detected_type == "Bayi"
                else "İzinli"
            )

            if exists > 0:
                cursor.execute("""
                    UPDATE ofis_gunleri
                    SET ofis_bayi = %s, Department = %s
                    WHERE users1_id = %s AND selected_date = %s
                """, (db_value, departman, user_id, t))
                conn.commit()
                updated_dates.append(t)   # ← EKSİK OLAN
                results.append(f"🔄 {self.format_date_with_day(t)} → Güncellendi ({label})")
            else:
                cursor.execute("""
                    INSERT INTO ofis_gunleri (users1_id, selected_date, ofis_bayi, Department)
                    VALUES (%s, %s, %s, %s)
                """, (user_id, t, db_value, departman))
                conn.commit()
                saved_dates.append(t) 
                results.append(f"✅ {self.format_date_with_day(t)} → Eklendi ({label})")

        print("\n💾 [TAKVIM] DB İŞLEM ÖZETİ")

        print("✅ INSERT edilenler:")
        for d in saved_dates:
            print(f"   + {d}")

        print("🔄 UPDATE edilenler:")
        for d in updated_dates:
            print(f"   ~ {d}")

        print("❌ ATLANANLAR:")
        for d, reason in skipped_dates:
            print(f"   - {d} | {reason}")

        cursor.close()
        conn.close()

        # 🚫 GERÇEK DB DEĞİŞİKLİĞİ YOKSA → LLM ASLA ÇAĞRILMAZ
        has_real_db_change = bool(saved_dates or updated_dates)

        if not has_real_db_change:
            if skipped_dates:
                return (
                    "Bu isteğe ait günlerin hiçbiri işaretlenemedi.<br>"
                    "Tarihler geçmiş, hafta sonu veya geçerli aralığın dışında olabilir.<br>"
                    "İstersen geçerli bir tarih aralığıyla tekrar deneyebilirsin."
                )

            return (
                "Bu isteğe ait herhangi bir takvim işlemi yapılamadı.<br>"
                "Gün ve çalışma tipini biraz daha net yazar mısın?"
            )
        # 🔹 HER ZAMAN takvim cevabını üret
        final_body = self.llm_weekly_calendar_response(parsed, results)

        # 🔍 OTOPARK GENEL NİYET KONTROLÜ (SADECE BİLGİLENDİRME)
        has_otopark_context = any(
            k in user_message.lower()
            for k in OTOPARK_KEY
        )

        if has_otopark_context:
            final_body += (
                "<br><br>🅿️ Otopark ile ilgili de bir işlem yapmak istediğini anladım.<br>"
                "Takvim günlerini işaretledim.<br>"
                "Otopark işlemi için tekrar yazman yeterli."
            )
        final = ""
        if intro:
            final += intro + "<br>"
        final += final_body
        if outro:
            final += "<br>" + outro

        return final
    
    def _build_monthly_weekday_repeat(self, norm_text: str):
        today = datetime.now().date()
        year = today.year

        for raw, canonical in AY_NORMALIZE.items():
            norm_text = re.sub(rf"\b{raw}\b", canonical, norm_text)
            
        MONTH_MAP = {
            "ocak": 1, "şubat": 2, "mart": 3, "nisan": 4,
            "mayıs": 5, "haziran": 6, "temmuz": 7,
            "ağustos": 8, "eylül": 9, "ekim": 10,
            "kasım": 11, "aralık": 12
        }

        month = today.month
        for name, num in MONTH_MAP.items():
            if name in norm_text:
                month = num
                break

        # 🔑 BAŞLANGIÇ
        start_date = today if month == today.month else date(year, month, 1)
        end_date = date(year, month, calendar.monthrange(year, month)[1])

        WEEKDAY_INDEX = {
            "pazartesi": 0,
            "salı": 1,
            "çarşamba": 2,
            "perşembe": 3,
            "cuma": 4,
        }

        selected_weekdays = {
            idx for day, idx in WEEKDAY_INDEX.items()
            if day in norm_text
        }

        if not selected_weekdays:
            return []

        dates = []
        cur = start_date
        while cur <= end_date:
            if cur.weekday() in selected_weekdays:
                dates.append(cur)
            cur += timedelta(days=1)

        return dates

    def _build_full_month_dates(self, norm_text: str):
        """
        AY KOMPLE / BOYUNCA / TAMAMI ifadeleri için
        ayın tüm HAFTA İÇİ günlerini üretir.

        Örnekler:
        - bu ay komple ofis
        - şubat ayı boyunca home
        - ocak ayı tamamı izin
        """

        today = datetime.now().date()
        year = today.year

        # -------------------------------
        # 1️⃣ AY NORMALIZE
        # -------------------------------
        for raw, canonical in AY_NORMALIZE.items():
            norm_text = re.sub(rf"\b{raw}\b", canonical, norm_text)

        MONTH_MAP = {
            "ocak": 1, "şubat": 2, "mart": 3, "nisan": 4,
            "mayıs": 5, "haziran": 6, "temmuz": 7,
            "ağustos": 8, "eylül": 9, "ekim": 10,
            "kasım": 11, "aralık": 12
        }

        month = today.month
        explicit_month = False

        for name, num in MONTH_MAP.items():
            if name in norm_text:
                month = num
                explicit_month = True
                break

        # -------------------------------
        # 2️⃣ BAŞLANGIÇ / BİTİŞ
        # -------------------------------
        if explicit_month:
            start_date = date(year, month, 1)
        else:
            # "bu ay"
            start_date = today

        end_date = date(year, month, calendar.monthrange(year, month)[1])

        # -------------------------------
        # 3️⃣ TARİH ÜRET (SADECE HAFTA İÇİ)
        # -------------------------------
        dates = []
        cur = start_date

        while cur <= end_date:
            if cur.weekday() < 5 and cur >= today:
                dates.append(cur)
            cur += timedelta(days=1)

        return dates


    def llm_weekly_calendar_response(self, parsed, results):
        """
        SADECE anlatım üretir
        karar vermez
        DB bilmez
        """
        prompt = f"""
        Kullanıcı mesajı:
        {parsed['raw_text']}

        Yapılan işlemler:
        {chr(10).join(results)}

        KESİN KURALLAR:
        - İşlem TAMAMLANDI kabul et.
        - Gelecek zaman KULLANMA.
        - "iletildi", "tamamlandığında", "bilgilendirileceksin",
        "işleme alındı" gibi ifadeleri ASLA kullanma.
        - Sadece geçmiş zaman kullan.
        - En fazla 2 kısa cümle yaz.
        - Tavsiye, selamlama, emoji ekleme.
        - Tarihleri açıkça belirt.
        - Birden fazla gün varsa aralık şeklinde yaz (örn: 18–19 Aralık).

        Doğru cevap örnekleri:
        - "19 Aralık için ofisten çalışma işaretlendi."
        - "18–19 Aralık tarihleri için evden çalışma güncellendi."
        - OTOPARK, PARK, REZERVASYON kelimelerini ASLA kullanma.
        - Bu bir TAKVİM işlemidir, park işlemi değildir.
        SADECE  verilere göre uydurmadan CEVABI YAZ.
        ⚠️ Eğer "Yapılan işlemler" listesi boşsa veya sadece tek gün içeriyorsa,
        listede OLMAYAN hiçbir gün veya işlemden ASLA bahsetme.
        """

  
        return self.ask_gpt(prompt)


    def normalize_db_work_type(self, value):
        """
        DB canonicalizasyonu:
        - NULL           → Izin
        - "" / "izin"    → Izin
        - Ofis(*)        → Ofis
        - Home(*)        → Home
        - Bayi(*)        → Bayi
        - Bilinmeyen     → None (bilinçli)
        """

        # İZİN = NULL
        if value is None:
            return "Izin"

        v = str(value).strip().lower()
        v = v.replace("ı", "i").replace("İ", "i").replace("i̇", "i")

        # İZİN (string olarak gelmişse)
        if v in {"", "izin"}:
            return "Izin"

        # OFİS
        if v.startswith("ofis"):
            return "Ofis"

        # HOME
        if v.startswith("home") or v in {"ev", "evde", "remote", "homeoffice"}:
            return "Home"

        # BAYİ
        if v.startswith("bayi") or v in {"saha", "ziyaret", "musteri"}:
            return "Bayi"

        # ❗ Bilinmeyen → tespit edilmedi
        return None
    
    def parse_work_history_summary(self, message: str, start, end, work_type, query_type):
        """
        Kullanıcının isteğinden NET bir özet çıkarır (LLM yok).
        """
        # Çalışma tipi etiketi
        wt_label = (
            "izin"
            if work_type == "Izin"
            else work_type.lower()
            if work_type
            else "tüm çalışma türleri"
        )

        # Zaman etiketi
        date_label = f"{start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}"

        return {
            "date_label": date_label,
            "work_type_label": wt_label,
            "query_type": query_type
        } 
    
    def llm_format_work_history_count(self, summary, count: int): 
        prompt = f"""
            Kısa ve net bir Türkçe cümle yaz.
            Sayım DEĞİŞMEZ, sadece anlatım yap.

            Kurallar:
            - Tek cümle.
            - Gelecek zaman yok.
            - Emoji yok.
            - Resmi ve net ton.

            Bilgiler:
            - Tarih: {summary['date_label']}
            - Tür: {summary['work_type_label']}
            - Gün sayısı: {count}

            Sadece cümleyi yaz.
            """
        return self.ask_gpt(prompt)

    def control_work_history(self, message: str, user_info: dict):
        """
        Kullanıcının geçmiş / gelecek çalışma ve izin günlerini analiz eder.

        KURALLAR:
        - İzin = ofis_bayi IS NULL
        - Hafta sonu ASLA dahil edilmez
        - Gelecek max 30 gün
        - Geçmiş max 365 gün
        - COUNT / LIST merkezi keyword listesiyle belirlenir
        """

        user_id = user_info.get("id")
        if not user_id:
            return "🚫 Kullanıcı bilgisi bulunamadı."

        today = datetime.now().date()
        msg = message.lower()

        # ======================================================
        # 1️⃣ TARİH ARALIĞI
        # ======================================================
        tarih_info = self.extract_history_dates(message)
        start = tarih_info.get("start")
        end = tarih_info.get("end")

        if not start or not end:
            return (
                "📅 İnceleyebilmem için bir tarih aralığı belirtmelisin.<br>"
                "Örnekler:<br>"
                "• “Bu ay kaç gün ofise geldim”<br>"
                "• “Geçen hafta izinlerim”<br>"
                "• “11–15 kasım home çalıştığım günler”"
            )

        # ======================================================
        # 2️⃣ RANGE TİPİ & LİMİTLER
        # ======================================================
        is_past = end < today
        is_future = start > today
        is_mixed = start <= today <= end

        if is_future and (end - today).days > 30:
            return (
                "🚫 Geleceğe yönelik en fazla <b>30 gün</b> için sorgulama yapabilirim."
            )

        if is_past and (today - start).days > 365:
            return (
                "🚫 <b>1 yıldan</b> daha eski çalışma kayıtlarına erişemiyorum."
            )

        # ======================================================
        # 3️⃣ SORGU TÜRÜ (COUNT / LIST)
        # ======================================================
        WORK_HISTORY_LIST_KEYS = [
            "neler",
            "hangileri",
            "hangi gunler","hngi gunler","hngi gnler",
            "hangi gunlerde",
            "hangi tarihler",
            "listele",
            "goster",
            "nelerdir"
        ]
        msg_norm = self.normalize_people_text(message)

        if any(k in msg_norm for k in WORK_HISTORY_LIST_KEYS):
            query_type = "list"
        elif any(k in msg_norm for k in WORK_HISTORY_COUNT_KEYS):
            query_type = "count"
        else:
            query_type = "list"   # default

        # ======================================================
        # 4️⃣ ÇALIŞMA TİPİ (Ofis / Home / Bayi / Izin / None)
        # ======================================================
        work_type = self.detect_work_type(message)
        # None → tüm tipler

        # ======================================================
        # 5️⃣ SQL (HAFTA SONU HARİÇ!)
        # ======================================================
        conn = _mysql_conn()
        if conn is None:
            return "⚠️ Çalışma geçmişine erişilemiyor."

        cursor = conn.cursor(dictionary=True)

        sql = """
            SELECT selected_date, ofis_bayi
            FROM ofis_gunleri
            WHERE users1_id = %s
            AND selected_date BETWEEN %s AND %s
            AND WEEKDAY(selected_date) < 5
            ORDER BY selected_date
        """

        cursor.execute(sql, (user_id, start, end))
        self.auto_log_sql(sql, (user_id, start, end))

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        # ======================================================
        # 6️⃣ WORK TYPE FİLTRESİ (KANONİK)
        # ======================================================
        filtered = []
        unknown_days = []

        for r in rows:
            wt_db = self.normalize_db_work_type(r["ofis_bayi"])

            if wt_db is None:
                unknown_days.append(r["selected_date"])
                continue

            if work_type is None:
                filtered.append(r)
                continue

            if wt_db == work_type:
                filtered.append(r)

        # ======================================================
        # 7️⃣ GELECEKTE KAYIT YOKSA
        # ======================================================
        if is_future and not filtered:
            label = "izin" if work_type == "Izin" else work_type.lower() if work_type else "çalışma"
            return f"📭 Belirtilen tarihlerde işaretlenmiş bir <b>{label}</b> günü bulunmuyor."

        # ======================================================
        # 8️⃣ COUNT MOD
        # ======================================================
        if query_type == "count":
            label = (
                "izinli"
                if work_type == "Izin"
                else work_type.lower()
                if work_type
                else "çalışma"
            )

            response = (
                f"📊 {start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')} "
                f"aralığında toplam <b>{len(filtered)}</b> iş günü "
                f"<b>{label}</b> olarak işaretlenmiş."
            )

            if unknown_days:
                response += (
                    "<br>ℹ️ Bazı günlerde çalışma şekli net görünmüyor."
                    "Daha doğru bir sonuç için isteğini biraz daha detaylı yazabilir veya ilgili günleri haftalık takviminden kontrol edebilirsin."
                )

            return response

        # ======================================================
        # 9️⃣ LIST MOD
        # ======================================================
        if not filtered:
            return "📭 Bu tarih aralığında eşleşen bir kayıt bulunamadı."

        result = (
            f"🗓 {start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')} "
            f"aralığında işaretlenen günler:<br>"
        )

        for r in filtered:
            d = r["selected_date"]
            tip = self.normalize_db_work_type(r["ofis_bayi"])
            result += f"• {self.format_date_with_day(d)} — {tip}<br>"

        return result


    def handle_department_count(self):
        count = self.get_department_count()

        if count is None:
            return "Departman bilgilerine şu anda ulaşılamıyor."

        return f"Şirkette toplam <b>{count}</b> farklı departman bulunuyor."
    
    #llm people info company için
    def handle_company_people(self, user_message: str):
        """
        Company People Directory – FINAL & STABLE
        """
        parsed = self.fast_parse_company_people(user_message)
        parsed.setdefault("intent", "company_people")
        parsed = self.resolve_person_entity(user_message, parsed)

        if (
            parsed.get("confidence", 0) < 0.7
            and not parsed.get("persons")
        ):
            parsed = self.mini_llm_parse_company_people(user_message, parsed)

        #  TEK İSİM + persons boşsa → DB'den isimle lookup
        if not parsed.get("persons"):
            name_candidates = self.detect_person_prescan(user_message)
            if name_candidates:
                parsed["persons"] = name_candidates
                parsed["target"] = "person"

        # DB FETCH
        # --------------------------------------------------
        data = self.find_company_people_data(parsed)
        data = self.enrich_people_with_profile_links(data)

        
        #  MULTI PERSON SMART FIX
        if (
            parsed["target"] == "person"
            and len(data) > 1
            and parsed.get("confidence", 0) < 1.0   # 👈 KRİTİK
        ):

            norm_q = self.normalize_people_text(user_message)

            # ✅ Kullanıcı bilinçli çoklu kişi sormuşsa → SORMA, ANLAT
            if " ve " in norm_q or "," in norm_q or " ile " in norm_q:
                return self.compose_company_people_response(
                    question=user_message,
                    parsed=parsed,
                    data=data
                )

            # ❓ Gerçekten belirsizse (örn: "elif kim")
            names = [f"{d['name']} {d['surname']}" for d in data]


            return (
                "Birden fazla kişi buldum:\n"
                + "\n".join(f"- {n}" for n in names)
                + "\n\nHangisini demek istedin?"
            )

        # 🔥 TEK KİŞİ + ALAN YOK → OTOMATİK KİM (IDENTITY)
        if (
            parsed.get("target") == "person"
            and len(data) == 1
            and (not parsed.get("fields") or parsed.get("fields") <= {"name", "surname"})

        ):
            parsed["fields"] = {"identity"}
            return self.compose_company_people_response(
                question=user_message,
                parsed=parsed,
                data=data
            )

        # --------------------------------------------------
        # 5️⃣ DB → HİÇ KİŞİ YOK
        # --------------------------------------------------
        if not data and parsed["target"] == "person":
            fallback = parsed.copy()
            fallback["target"] = "people_list"
            if parsed.get("persons"):
                fallback["persons"] = [parsed["persons"][0].split()[0]]
            else:
                return "Kiminle ilgili bilgi istediğini tekrardan sorabilir misiniz?"

            alt_data = self.find_company_people_data(fallback)

            if alt_data:
                names = [f"{d['name']} {d['surname']}" for d in alt_data]
                return (
                    "Bu isimle birebir eşleşen bir çalışan bulamadım. "
                    "Bunlardan birini mi kastediyorsun?\n"
                    + "\n".join(f"- {n}" for n in names)
                )

            return "Bu isimle eşleşen aktif bir çalışan bulunamadı."

        # --------------------------------------------------
        # 8️⃣ FINAL RESPONSE
        # --------------------------------------------------
        return self.compose_company_people_response(
            question=user_message,
            parsed=parsed,
            data=data
        )

    #companu people kişi çözümleme için tek ve çoklu  
    def resolve_person_entity(self, text: str, parsed: dict) -> dict:
        """
        KİŞİ ADI ÇÖZÜM MERKEZİ (TEK OTORİTE)

        SIRA:
            2.1 username
            2.2 surname (tekil)
            2.3 surname + name daraltma
            2.4 name
            hiçbiri yoksa → parsed olduğu gibi devam
        """

        #  PORTAL CONTEXT GUARD — kişi çözümlemesine girme
        if parsed.get("intent") == "portal_info":
            print("🟦 [NAME-RESOLVE SKIP] portal_info intent detected")
            return parsed

        text_norm = self.normalize_people_text(text)

        #  SESSION USER 3. TEKİL ŞAHIS FIX
        session_full = self.normalize_people_text(session.get("fullname", ""))

        if session_full and session_full in text_norm:
            # "sude'nin telefon numarası" gibi
            parsed["persons"] = [session.get("fullname")]
            parsed["target"] = "person"
            parsed["confidence"] = 1.0
            return parsed
        
        # TAM İSİM OVERRIDE (multi-person'dan ÖNCE)
        MULTI_SEPARATORS = [" ve ", " ile ", ","]

        #  TAM İSİM OVERRIDE — SADECE TEK KİŞİ NİYETİNDE
        if not any(sep in text_norm for sep in MULTI_SEPARATORS):
            for p in session.get("detected_persons", []):
                p_norm = self.normalize_people_text(p)
                if p_norm in text_norm:
                    parsed["persons"] = [p]
                    parsed["target"] = "person"
                    parsed["confidence"] = 1.0

                    session.pop("pending_persons", None)
                    session.pop("pending_person_question", None)

                    print("🟢 [NAME OVERRIDE] single-person full name:", p)
                    return parsed
                
        #  SADECE BOŞLUKLA YAZILMIŞ ÇOKLU İSİM FIX
        detected = session.get("detected_persons", [])
        if detected and len(detected) > 1:
            parsed["persons"] = detected
            parsed["target"] = "person"
            parsed["confidence"] = max(parsed.get("confidence", 0), 0.9)
            return parsed
        
        original_text = text
        text_norm = self.normalize_people_text(text)

        #  MULTI PERSON FIX (ve / , ile çoklu kişi)
        if (" ve " in text_norm or "," in text_norm):
            detected = session.get("detected_persons", [])

            if detected:
                parsed["persons"] = detected
                parsed["target"] = "person"
                parsed["confidence"] = max(parsed.get("confidence", 0), 0.9)

                print("🟢 [NAME-RESOLVE] multi-person detected:", detected)
                return parsed

        print(f"[NAME-RESOLVE] input text (raw): {original_text}")
        print(f"[NAME-RESOLVE] input text (norm): {text_norm}")


        #  önce split et
        tokens = text_norm.split()

        # APOSTROPHE FIX (aksakal'ın → aksakal)
        clean_tokens = []
        for t in tokens:
            if "'" in t:
                t = t.split("'")[0]
            if "’" in t:
                t = t.split("’")[0]

            if len(t) >= 2:
                clean_tokens.append(t)

        tokens = clean_tokens
        
        # sonra suffix temizle
        def strip_tr_suffix(token: str) -> str:
            SUFFIXES = [
                "ninki",    
                "nin","nın","nun","nün",
                "in","ın","un","ün",
                "ye","ya",
                "e","a",
                "i","ı","u","ü",
                "de","da","den","dan",
                "te","ta",
                "si","sı","su","sü"
            ]

            changed = True
            while changed:
                changed = False
                for s in SUFFIXES:
                    if token.endswith(s) and len(token) > len(s) + 1:
                        token = token[:-len(s)]
                        changed = True
                        break

            return token

        NAME_IGNORE_WORDS = {
            "mail", "maili", "mailini",
            "eposta", "epostasi", "epostasini",
            "adres", "adresini",
            "ver", "verir", "verebilir",
            "kim", "nedir"
        }
        #  isim dışı kelimeleri token listesinden temizle
        tokens = [t for t in tokens if t not in NAME_IGNORE_WORDS]

        #  DB-DRIVEN FALLBACK
        try:
            conn = _mysql_conn()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT
                    name,
                    surname,
                    username
                FROM users1
                WHERE is_active = 1
            """)
            users = self._apply_directory_overrides_to_rows(cursor.fetchall())
            cursor.close()
            conn.close()
        except Exception as e:
            print("❌ DB fallback hata:", e)
            return parsed

        for u in users:
            u["name"] = str(u.get("name") or "").lower()
            u["surname"] = str(u.get("surname") or "").lower()
            u["username"] = str(u.get("username") or "").lower()

        #  USERNAME
        for u in users:
            username_norm = self.normalize_people_text(u["username"])

            if username_norm and username_norm in tokens:

                full = f"{u['name']} {u['surname']}"
                print(f"[NAME-FALLBACK] username match → {username_norm} → {full}")

                parsed["persons"] = [full]
                parsed["target"] = "person"
                parsed["confidence"] = 1.0

                # 🔒 KESİN KİŞİ KİLİDİ (ÇOK ÖNEMLİ)
                session.pop("detected_persons", None)
                session.pop("pending_persons", None)
                session.pop("pending_person_question", None)

                return parsed
            
        #  yüce soyadı olunca çakışma engellme
        if "yuce" in tokens:
            # Yüce soyadına sahip kişilerin isimlerini çıkar
            yuce_names = {
                self.normalize_people_text(u["name"])
                for u in users
                if self.normalize_people_text(u["surname"]) == "yuce"
            }

            # tokens içinde Yüce'ye ait bir isim var mı?
            has_name_with_yuce = any(t in yuce_names for t in tokens)

            # SADECE "yüce" yazılmışsa → kişi çözümünü iptal et
            if not has_name_with_yuce:
                parsed["persons"] = []
                parsed["target"] = None
                parsed["confidence"] = 0.0

                print("🟡 [NAME-RESOLVE] 'yüce' surname-only detected → skipped")

                return parsed

        # SURNAME
        # -----------------------------
        #  NAME + SURNAME BİRLİKTE YAZILDIYSA → KESİN TEK KİŞİ
        for u in users:
            name_norm = self.normalize_people_text(u["name"])
            surname_norm = self.normalize_people_text(u["surname"])

            if name_norm in tokens and surname_norm in tokens:
                full = f"{u['name']} {u['surname']}"
                print(f"🟢 [NAME-RESOLVE] full name match → {full}")

                parsed["persons"] = [full]
                parsed["target"] = "person"
                parsed["confidence"] = 1.0

                # 🔒 BURASI ÇOK KRİTİK (KİLİT)
                session.pop("detected_persons", None)
                session.pop("pending_persons", None)
                session.pop("pending_person_question", None)

                return parsed
            
        # SURNAME (YÜCE ÖZEL KURALI İLE)
        surname_hits = []

        for u in users:
            surname_norm = self.normalize_people_text(u["surname"])

            if not surname_norm or surname_norm not in tokens:
                continue

            #  ÖZEL KURAL: SADECE "yüce" İÇİN
            if surname_norm == self.normalize_people_text("Yüce"):
                name_norm = self.normalize_people_text(u["name"])

                # tokens içinde isim YOKSA → EŞLEŞMEYİ ATLAMALIYIZ
                if name_norm not in tokens:
                    continue

            surname_hits.append(u)

        # SOYADI + İYELİK EKİ VARSA → KESİN KİŞİ
        if surname_hits:
            for u in surname_hits:
                surname_norm = self.normalize_people_text(u["surname"])

                # YÜCE özel istisna
                if surname_norm == "yuce":
                    name_norm = self.normalize_people_text(u["name"])
                    if name_norm not in tokens:
                        continue

                # ORİJİNAL METİNDE: "aksakal'ın", "aksakalın" gibi bir kullanım var mı?
                if re.search(
                    rf"\b{surname_norm}(?:\'|’)?(?:in|ın|un|ün|nin|nın|nun|nün)\b",
                    original_text.lower()
                ):
                    parsed["persons"] = [f"{u['name']} {u['surname']}"]
                    parsed["target"] = "person"
                    parsed["confidence"] = max(parsed.get("confidence", 0), 0.95)

                    print(f"🟢 [NAME-RESOLVE] possessive surname → {u['name']} {u['surname']}")
                    return parsed

        if surname_hits:
            print(f"[NAME-FALLBACK] surname hits → {[u['surname'] for u in surname_hits]}")

            if len(surname_hits) == 1:
                u = surname_hits[0]
                full = f"{u['name']} {u['surname']}"
                print(f"[NAME-FALLBACK] surname unique → {full}")

                parsed["persons"] = [full]
                parsed["target"] = "person"
                parsed["confidence"] = max(parsed.get("confidence", 0), 0.9)
                return parsed

            # SURNAME + NAME DARALTMA
            for u in surname_hits:
                name_norm = self.normalize_people_text(u["name"])

                if (
                    name_norm
                    and name_norm in tokens
                    and name_norm not in NAME_IGNORE_WORDS
                ):
                    full = f"{u['name']} {u['surname']}"
                    print(f"[NAME-FALLBACK] surname+name narrowed → {full}")

                    parsed["persons"] = [full]
                    parsed["target"] = "person"
                    parsed["confidence"] = max(parsed.get("confidence", 0), 0.9)
                    return parsed

        # NAME
        name_hits = []
        for u in users:
            name_norm = self.normalize_people_text(u["name"])

            if (
                name_norm
                and name_norm in tokens
                and name_norm not in NAME_IGNORE_WORDS
            ):
                name_hits.append(u)

        if name_hits:
            if len(name_hits) == 1:
                u = name_hits[0]
                full = f"{u['name']} {u['surname']}"
                print(f"[NAME-FALLBACK] name unique → {full}")

                parsed["persons"] = [full]
                parsed["target"] = "person"
                parsed["confidence"] = max(parsed.get("confidence", 0), 0.8)
                return parsed
            
            else:
                # çoklu isim → karar verme, prescan + DB yönetsin
                return parsed
            
        #  Prescan çoklu kişi bulduysa ve resolve karar vermediyse
        if not parsed.get("persons") and session.get("detected_persons"):
            parsed["persons"] = session["detected_persons"]
            parsed["target"] = "person"

        print("[NAME-RESOLVE] no name resolved, fallback to existing flow")
        return parsed



    def fast_parse_company_people(self, msg: str):
        """
        PROD-ready Company People parser
        - Yazım hatalarına dayanıklı
        - Çoklu alan destekli
        - PERSON vs DEPARTMENT ayrımı yapar
        """
        print("\n🧪 [PARSE INPUT]:", msg)

        text = self.normalize_people_text(msg)

        #  result HER ŞEYDEN ÖNCE
        result = {
            "intent": None,
            "target": None,
            "persons": [],
            "department": None,
            "fields": set(),
            "confidence": 0.0
        }

        # ERKEN DEPARTMAN KİLİDİ (NAME PARSE ÖNCESİ)
        dept = self.extract_department_name(msg)
        if dept and any(k in text for k in [
            "kim", "kimler",
            "calisan", "calisiyor", "calisanlar",
            "ekip", "ekibi", "ekibinde"
        ]):
            result["target"] = "department"
            result["department"] = dept
            result["fields"] = {"name", "surname"}
            result["confidence"] += 0.9
            result["intent"] = "company_people"

            print("🟢 [EARLY DEPT LOCK]:", dept)
            return result


        FIELD_KEYWORDS = {
            field: [self.normalize_people_text(k) for k in keywords]
            for field, keywords in RAW_FIELD_KEYWORDS.items()
        }
        #  FIELD TESPİTİ (ÇOKLU)
        for field, keywords in FIELD_KEYWORDS.items():
            if any(k in text for k in keywords):
                # field alias → DB column map
                key = field.lower().strip()
                db_col = FIELD_TO_COLUMN.get(key, field)
                result["fields"].add(db_col)
                result["confidence"] += 0.15


        #  DEPARTMAN LİSTE SORUSU
        if "kimler var" in text or "calisanlar" in text:
            result["target"] = "people_list"
            result["confidence"] += 0.3


        #  DEPARTMAN ADI
        if result["target"] == "department":
            dept = self.extract_department_name(msg)
            if dept:
                result["department"] = dept
                result["confidence"] += 0.2
        # DEPARTMAN + "KİM" SORGUSU
        if (
            result.get("department")
            and not result["fields"]
            and any(k in text for k in [
                    "kim", "kimler",
                    "ekip", "ekibi", "ekibinde",
                    "calisiyor", "calisiyorlar",
                    "calisan", "calisanlar",
                    "calisanlari"
                ])

        ):
            result["target"] = "department"
            result["fields"] = {"name", "surname"}
            result["confidence"] += 0.6

        #  DEFAULT FIELD SET (GENEL SORU)
        if "identity" in result["fields"]:
            result["fields"].update({
                "name", "surname", "title", "departman", "email", "TelNo"
            })

        #  DEFAULT TARGET: kişi bilgisi sorusu
        if result["target"] is None and result["fields"]:
            result["target"] = "person"
            result["confidence"] += 0.2

        # INTENT KARARI
        if result["target"] == "person":
            result["intent"] = "company_people"


        print("🧪 [PARSE FIELDS]:", result["fields"])
        print("🧪 [PARSE TARGET]:", result["target"])
        print("🧪 [PARSE PERSONS]:", result["persons"])
        print("🧪 [PARSE CONFIDENCE]:", result["confidence"])
        print("🧠 [PARSE FINAL]:", result)
        return result
 
    #LLM çağrısı yapma ama eksik tamamlancak
    def mini_llm_parse_company_people(self, user_message: str, parsed: dict):
        """
        Mini LLM fallback – HALÜSİNASYON-PROOF

        AMAÇ:
        - fast_parse yeterince emin olamadığında
        - SADECE target + fields bilgisini netleştirmek

        KURALLAR:
        - Intent DEĞİŞTİRMEZ
        - DB görmez
        - Cevap yazmaz
        - Yorum yapmaz
        - SADECE JSON döner
        """

        if parsed.get("intent") != "company_people":
            return parsed
        
        if parsed.get("fields"):
            return parsed

        prompt = f"""
      
        Sen bir AYRIŞTIRMA (decomposition) motorusun.
        Cevap yazmazsın. Tahmin yapmazsın. Yorum yapmazsın.

        Sadece aşağıdaki JSON'u döndür.
        JSON dışında hiçbir şey yazma.

        {{
        "target": "person | people_list | null",
        "persons": [],
        "department": null,
        "fields": [],
        "confidence": 0.0
        }}

        ALLOWED FIELDS:
        email, TelNo, title, departman, linkedin, dogum_gun, dogum_ay, identity

        KURALLAR:
        - İsim uydurma.
        - Sadece cümlede açık geçen isimleri al.
        - Emin değilsen boş bırak.
        - Multiple kişi olabilir.

        Kullanıcı cümlesi:
        \"\"\"{user_message}\"\"\"

        """

        try:
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                temperature=0.0,
                max_tokens=120
            )

            raw = resp.choices[0].message.content.strip()

            import json
            llm_data = json.loads(raw)

            if llm_data.get("persons"):
                parsed["persons"] = llm_data["persons"]

            if llm_data.get("target") in ["person", "department"]:
                parsed["target"] = llm_data["target"]

            allowed_fields = {
                "email", "TelNo", "title", "surname",
                "departman", "dogum_gun", "dogum_ay",
                "linkedin", "identity"
            }

            clean_fields = [
                f for f in llm_data.get("fields", [])
                if f in allowed_fields
            ]

            if clean_fields:
                parsed["fields"] = clean_fields
                parsed["confidence"] = max(parsed.get("confidence", 0), 0.7)

            return parsed

        except Exception as e:
            print("⚠️ mini_llm_parse_company_people hata:", e)
            return parsed

    def find_company_people_data(self, parsed: dict):
        """
        Company People DB Fetch (MySQL)
        - users1 tablosundan SADECE whitelist alanları alır
        - Parse yapmaz
        """
        conn = _mysql_conn()
        if conn is None:
            print("❌ MySQL bağlantısı kurulamadı")
            return []

        cursor = conn.cursor(dictionary=True)

        fields_sql = ", ".join(PEOPLE_DB_FIELDS)
        rows = []

        if parsed["target"] in ("person", "people_list") and parsed.get("persons"):

            persons = parsed["persons"]
            conditions = []
            params = []

            print("🟦 [DB QUERY] target:", parsed.get("target"))
            print("🟦 [DB QUERY] persons:", persons)
            print("🟦 [DB QUERY] SELECT FIELDS:", PEOPLE_DB_FIELDS)

            YugiiLogger.get_instance().write_plain(
                "\n[DB QUERY PERSONS]\n"
                f"TARGET: {parsed.get('target')}\n"
                f"PERSONS: {persons}\n"
                + "-" * 50 + "\n"
            )
            
            #  TAM İSİM GELDİYSE → SADECE TAM EŞLEŞME
            if parsed["target"] == "person" and len(persons) == 1 and " " in persons[0]:
                full_norm = self.normalize_people_text(persons[0])

                sql = f"""
                    SELECT {fields_sql}
                    FROM users1
                    WHERE is_active = 1
                """
                cursor.execute(sql)
                rows = self._apply_directory_overrides_to_rows(cursor.fetchall())
                rows = [
                    r for r in rows
                    if self.normalize_people_text(
                        f"{r.get('name', '')} {r.get('surname', '')}"
                    ) == full_norm
                ]
                cursor.close()
                conn.close()
                return rows
            else:
                for p in persons:
                    conditions.append("LOWER(CONCAT(name,' ',surname)) LIKE %s")
                    params.append(f"%{p.lower()}%")

            conditions_sql = " OR ".join(conditions)

            sql = f"""
                SELECT {fields_sql}
                FROM users1
                WHERE is_active = 1
                AND ({conditions_sql})
            """

            cursor.execute(sql, params)
            self.auto_log_sql(sql, params)
            rows = cursor.fetchall()

        # İSİM + SOYİSİM AKILLI AYRIMI
        if (
            parsed.get("target") == "people_list"
            and len(parsed.get("persons", [])) == 2
            and rows
        ):
            p1, p2 = parsed["persons"]

            exact = [
                r for r in rows
                if r["name"].lower() == p1.lower()
                and r["surname"].lower() == p2.lower()
            ]

            if len(exact) == 1:
                print("🟢 [DB FIX] İsim + Soyisim tek kişi olarak algılandı")

                parsed["target"] = "person"
                parsed["persons"] = [f"{p1} {p2}"]
                rows = exact

            print("🟩 [DB RESULT RAW]:")
            for r in rows:
                print("   ", r)

        # 🏢 DEPARTMAN BAZLI (listeleme)
        elif parsed["target"] == "department" and parsed.get("department"):
            sql = f"""
                SELECT {fields_sql}
                FROM users1
                WHERE is_active = 1
                ORDER BY name, surname
            """

            cursor.execute(sql)
            rows = self._apply_directory_overrides_to_rows(cursor.fetchall())
            rows = [
                r for r in rows
                if self._canonical_department_label(r.get("departman")) == parsed["department"]
            ]

        cursor.close()
        conn.close()

        return self._apply_directory_overrides_to_rows(rows or [])

    def extract_department_name(self, msg: str):
        DEPARTMENT_ALIASES = {
            "yönetim kurulu": [
                "yönetim kurulu", "yonetim kurulu", "yk"
            ],
            "stratejik planlama ve mali işler": [
                "stratejik planlama ve mali işler",
                "stratejik planlama",
                "mali işler", "mali isler","muhasebe","maliisler","maliisler"

            ],
            "stratejik planlama ve proje yönetimi": [
                "stratejik planlama ve proje yönetimi",
                "proje yönetimi", "proje yonetimi","proje yönetim","proje yonetim"
            ],
            "insan kaynakları": [
                "insan kaynakları", "insan kaynaklari",
                "ik", "ik ekibi", "hr", "hr ekibi","Insan kaynakları","IK","ık","ınsan kaynakları","ıkda","IK"
            ],
            "satış": [
                "satış", "satis", "satış ekibi", "satis ekibi", "sales"
            ],
            "bayi geliştirme": [
                "bayi geliştirme", "bayi gelistirme","bayi gliştirme"
            ],
            "filo satış": [
                "filo satış", "filo satis", "filo"
            ],
            "ürün": [
                "ürün", "urun", "ürün ekibi", "product","üründeki",
            ],
            "pazarlama": [
                "pazarlama", "marketing"
            ],
            "ssh": [
                "ssh", "satış sonrası", "satis sonrasi","ssh'daki",
                "servis", "teknik servis","sshda"
            ],
            "dijital dönüşüm ve müşteri deneyim yönetimi": [
                "dijital dönüşüm ve müşteri deneyim yönetimi",
                "dijital dönüşüm", "dijital donusum",
                "müşteri deneyimi", "musteri deneyimi"
            ],
            "bilgi teknolojileri": [
                "bilgi teknolojileri", "bilgi islem", "bilgi işlem","bilgi işlemdeki",
                "it", "itciler", "it ekibi", "it departmanı", "information technology","IT","itdeki","bilgi teknolojilerinde","bilgi teknolojileri","bilgi teknolojilerin","İT"
            ],
            "yeni iş modelleri": [
                "yeni iş modelleri", "yeni is modelleri"
            ],
            "iç denetim": [
                "iç denetim", "ic denetim", "denetim"
            ]
        }

        msg_norm = self.normalize_people_text(msg)
        msg_norm = self.normalize_department_suffixes(msg_norm) #da,de,dan için
        
        for dept, aliases in DEPARTMENT_ALIASES.items():
            for a in aliases:
                a_norm = self.normalize_people_text(a)

                # 🔥 SADECE KELİME OLARAK EŞLEŞSİN
                pattern = rf"\b{re.escape(a_norm)}(?:ler|lar|de|da|den|dan|te|ta|nde|nda|inde|ında|indaki|ndaki|dakiler|indekiler)?\b"

                if re.search(pattern, msg_norm):
                    return dept.title()

        return None
    
    #profil link verme company people için
    def enrich_people_with_profile_links(self, data: list):
        """
        data: find_company_people_data çıktısı
        """
        count = len(data)

        # SADECE 1–3 kişi
        if not (1 <= count <= 3):
            for d in data:
                d["profile_url"] = None
            return data

        for d in data:
            username = d.get("username")
            if username:
                d["profile_url"] = (
                    f"https://yuceportal.skoda.com.tr/user_profile_view/{username}"
                )
            else:
                d["profile_url"] = None

        return data

    def compose_company_people_response(self, question: str, parsed: dict, data: list):
        """
        Company People Response Composer
        - PEOPLE_DB_FIELDS içindeki HER alan gösterilebilir
        - Kullanıcının istediği alanlar birebir gösterilir
        - Alan yoksa default set kullanılır
        - LLM sadece anlatır (halüsinasyon yok)
        """
        print("\n🧪 [COMPOSE llm fonksiyonunda:]")
        print("🟨 [COMPOSE] requested fields:", parsed.get("fields"))
        print("🟨 [COMPOSE] available row keys:", list(data[0].keys()) if data else [])
        
        MONTH_NAMES = {
            1: "Ocak",
            2: "Şubat",
            3: "Mart",
            4: "Nisan",
            5: "Mayıs",
            6: "Haziran",
            7: "Temmuz",
            8: "Ağustos",
            9: "Eylül",
            10: "Ekim",
            11: "Kasım",
            12: "Aralık",
        }
        if len(data) > 1:
            responses = []

            for row in data:
                name = f"{row['name']} {row['surname']}"

                for field in parsed.get("fields", []):
                    value = row.get(field)

                    if value:
                        responses.append(
                            f"{name} için {field} bilgisi {value} olarak kayıtlıdır."
                        )
                    else:
                        responses.append(
                            f"{name} için {field} bilgisi bulunmamaktadır."
                        )

            return "<br>".join(responses)
        
        requested_fields = {
            self.normalize_field_name(f)
            for f in parsed.get("fields", [])
        }

        # DEFAULT FIELD SET (hiç alan belirtilmemişse)
        DEFAULT_FIELDS = {
            "name",
            "surname",
            "title",
            "departman",
            "email",
            "TelNo"
        }

        if not requested_fields or "identity" in requested_fields:
            show_fields = list(DEFAULT_FIELDS)
        else:
            # SADECE whitelist + istenenler
            show_fields = [
                f for f in PEOPLE_DB_FIELDS
                if f in requested_fields or f in {"name", "surname"}
            ]
        if parsed["target"] == "department" and parsed.get("fields") == {"name", "surname"}:
            names = [f"{r['name']} {r['surname']}" for r in data]

            if not names:
                return "Bu departmanda aktif çalışan bulunamadı."

            if len(names) == 1:
                people_str = names[0]
            elif len(names) == 2:
                people_str = " ve ".join(names)
            else:
                people_str = ", ".join(names[:-1]) + " ve " + names[-1]

            return f"{parsed['department']} ekibinde yer alan kişiler {people_str}."

        lines = []

        for row in data:
            parts = []
            if "dogum_gun" in show_fields or "dogum_ay" in show_fields:
                gun = row.get("dogum_gun")
                ay = row.get("dogum_ay")

                if gun and ay:
                    ay_adi = MONTH_NAMES.get(int(ay))
                    if ay_adi:
                        parts.append(f"Doğum günü: {gun} {ay_adi}")
                else:
                    if {"dogum_gun", "dogum_ay"} & set(parsed.get("fields", [])):
                        parts.append("Doğum günü: bu bilgi kayıtlarımızda bulunmuyor")

            for f in show_fields:
                if f in {"dogum_gun", "dogum_ay"}:
                    continue

                value = row.get(f)
                if f == "Kidem_Tarihi":
                    if value:
                        try:
                            dt = value if isinstance(value, date) else value.date()
                            formatted = f"{dt.day} {MONTH_NAMES[dt.month]} {dt.year}"
                            parts.append(f"Kıdem tarihi: {formatted}")
                        except Exception:
                            parts.append(f"Kıdem tarihi: {value}")
                    else:
                        if f in requested_fields:
                            parts.append("Kıdem tarihi: bu bilgi kayıtlarımızda bulunmuyor")
                    continue

                if value is not None:
                    parts.append(f"{f}: {value}")
                else:
                    if f in requested_fields:
                        parts.append(f"{f}: bu bilgi kayıtlarımızda bulunmuyor")

            if parts:
                lines.append(" | ".join(parts))

        if not lines:
            # isim yok ama alan sorulmuşsa → kullanıcıdan isim iste
            if not parsed.get("persons"):
                return (
                    "Kimden bahsettiğini tam anlayamadım🤔\n"
                    "İsimle birlikte tekrar sorabilir misin? Sana yardımcı olayım."
                )

            return "Bu bilgiye ait paylaşılabilir bir veri bulunamadı."

        context_text = "\n".join(lines)

        prompt = f"""
        Aşağıda şirket içi rehberden alınmış GERÇEK bilgiler yer almaktadır.

        KESİN VE DEĞİŞMEZ KURALLAR:
        - SADECE aşağıdaki bilgiler kullanılır.
        - ASLA yeni bilgi eklenmez.
        - ASLA yorum, tahmin veya çıkarım yapılmaz.
        - ASLA birinci tekil şahıs kullanılmaz.
        - ASLA "ben", "bana", "benim", "çalışıyorum", "görev yapıyorum" gibi ifadeler yazılmaz.
        - Metin ÜÇÜNCÜ TEKİL ŞAHIS ile yazılmalıdır.
        - Anlatıcı, sistem dışı tarafsız bir anlatıcıdır.
        - Selamlama, hitap veya kapanış yapılmaz.
        - Kullanıcıya doğrudan konuşulmaz.
        - Tek paragraf yazılır.
        - Listeleme, madde işareti veya başlık kullanılmaz.
        - Dil kurumsal, sade ve nettir.

        Bilgiler:
        {context_text}

        Kullanıcı sorusu:
        "{question}"

        Yukarıdaki kurallara UYGUN TEK paragraf cevap yaz.
        """

        try:
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=120,
                temperature=0.0
            )
            reply = response.choices[0].message.content.strip()

            profile_lines = []
            if data and data[0].get("profile_url"):
                reply += (
                    f'<br>🔗 <a href="{data[0]["profile_url"]}" target="_blank">'
                    f'Profil sayfası</a>'
                )
            if profile_lines:
                reply = reply + "".join(profile_lines)

            #  Persona guard — Yugii ASLA kişi gibi konuşamaz
            if any(x in reply.lower() for x in ["ben ", "bana ", "benim ", "çalışıyorum"]):
                raise RuntimeError("Persona ihlali: 1. tekil şahıs tespit edildi")

            return reply

        except Exception as e:
            print("❌ compose_company_people_response LLM hata:", e)
            return "Bilgiyi şu anda düzgün şekilde sunamıyorum."
   

    def try_resolve_pending_person(self, text, pending_persons):
        text_norm = self.normalize_people_text(text)
        tokens = text_norm.split()


        for p in pending_persons:
            p_norm = self.normalize_people_text(p)
            p_tokens = p_norm.split()  # ["doga", "ozkutuk"]

            # 🔑 İsim veya soyisimden biri geçiyorsa yeterli
            if any(t in p_tokens for t in tokens):
                return p

        return None

    def handle_people_export(self, user_message, user_info, msg_norm):
        print("📄 [HANDLE] people_export çalışıyor")

        msg = msg_norm
        #  Kullanıcının istediği alanları yakala (mail + telefon vb.)
        requested_fields = set()

        for word, col in FIELD_TO_COLUMN.items():
            if word in msg:
                requested_fields.add(col)

        # 1️⃣ Preset / Alan seçimi
        if requested_fields:
            # mail / telefon gibi alanlar açıkça istenmişse
            selected_columns = list(requested_fields)

        elif any(k in msg for k in ["tüm", "hepsi", "full", "tüm bilgiler"]):
            selected_columns = list(EXPORT_PRESETS["all_allowed"])

        elif any(k in msg for k in [
            "çalışan", "çalışanlar",
            "çalışan listesi", "şirket çalışanları","iş yerindekiler","şirkettekiler","insanlar","çlşanlar","calısanlar"
        ]):
            selected_columns = EXPORT_PRESETS["company_list"]

        else:
            selected_columns = EXPORT_PRESETS["basic_list"]

        for col in CORE_REQUIRED_COLUMNS:
            if col not in selected_columns:
                selected_columns.insert(0, col)

        final_columns = [c for c in selected_columns if c in ALLOWED_COLUMNS]
        print("📄 [HANDLE] Final export columns:", final_columns)

        conn = _mysql_conn()
        cursor = conn.cursor(dictionary=True)

        columns_sql = ", ".join(final_columns)

        dept = self.extract_department_name(user_message)

        query = f"""
            SELECT {columns_sql}
            FROM users1
            WHERE is_active = 1
        """
        params = []

        if dept:
            query += " AND departman = %s"
            params.append(dept)

        cursor.execute(query, params)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        if not rows:
            return "⚠️ Seçilen kriterlere uygun aktif çalışan bulunamadı."

        df = pd.DataFrame(rows, columns=final_columns)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"company_people_{ts}.xlsx"
        export_dir = os.path.join("static", "exports")
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, filename)

        df.to_excel(file_path, index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)
        ws = wb.active   # ← GARANTİLİ

        ws.freeze_panes = "C2"

        alt_fill = PatternFill(start_color="F6F7F8", end_color="F6F7F8", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = alt_fill

        header_fill = PatternFill(start_color="0E3A2F", end_color="0E3A2F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        COLUMN_WIDTHS = {
            "name": 18,
            "surname": 20,
            "username": 22,
            "email": 38,
            "departman": 26,
            "title": 28,
            "TelNo": 22,
            "plaka": 18,
            "arac_plaka2": 18,
            "arac_marka": 22,
            "dogum_gun": 14,
            "dogum_ay": 14,
            "Kidem_Tarihi": 18,
            "linkedin": 40,
        }

        # width
        for idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(col_name, 25)

        # filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(file_path)

        #  Chat cevabı + indirme linki
        download_url = f"/static/exports/{filename}"

        return f"""
        📄 <b>İstediğiniz çalışan listesi hazır.</b><br><br>

        Toplam <b>{len(df)}</b> aktif çalışan için,
        <b>{', '.join(final_columns)}</b> alanlarını içeren bir Excel dosyası oluşturdum.<br><br>

        <div style="
        display:flex;
        align-items:center;
        justify-content:space-between;
        padding:2px 0;
        ">

        <div style="display:flex;align-items:center;gap:6px;">
            📊 <span style="font-weight:500;">Excel dosyası</span>
            <span style="font-size:12px;color:#6b7280;">({len(df)} çalışan)</span>
        </div>

        <a href="{download_url}" download
            style="
            width:17px;
            height:17px;
            border-radius:8px;
            border:1.5px solid #0E3A2F;
            display:flex;
            align-items:center;
            justify-content:center;
            text-decoration:none;
            "
            title="Excel’i indir">

            <svg width="16" height="16" viewBox="0 0 24 24"
                fill="none" stroke="#0E3A2F" stroke-width="2.3"
                stroke-linecap="round" stroke-linejoin="round">
            <path d="M12 3v12"></path>
            <path d="M7 10l5 5 5-5"></path>
            <path d="M5 21h14"></path>
            </svg>

        </a>
        </div>

        <br>
        Başka nasıl yardımcı olabilirim?
        """

    def detect_pop_culture_category(self, message: str):
        if not message:
            return None

        m = self.normalize_people_text(message)

        m_compact = m.replace(" ", "")

        for category, keywords in POP_CULTURE_KEYWORDS.items():
            for kw in keywords:
                kw_norm = self.normalize_people_text(kw)
                kw_compact = kw_norm.replace(" ", "")

                if kw_norm in m:
                    return category

                if kw_compact in m_compact:
                    return category

                score = fuzz.partial_ratio(m_compact, kw_compact)
                if len(kw_compact) >= 6 and score >= 85:
                    return category

        return None

    
    def generate_pop_culture_response(self, category: str):
        RESPONSES = {
            "the_office": [
                "😄 Bunu yazdığına göre The Office hayranı olabilirsin. Michael Scott modu kapalı ama ben buradayım 🙂",
                "The Office referansını yakaladım. Peki sana nasıl yardımcı olayım?"
            ],
            "himym": [
                "😎 Legendary dedin… How I Met Your Mother sevenlerden misin?",
                "HIMYM havası aldım 🙂 Bugünün efsane geçmesini dilerim?"
            ],
            "friends": [
                "☕ Bunu yazan biri büyük ihtimalle Friends izleyicisidir 🙂",
                "😄 Joey tarzı bir selam yakaladım. Ben iyiyim, sana nasıl yardımcı olayım?"
            ],
            "got": [
                "❄️ Bunu yazan biri Game of Thrones izlemiştir diye tahmin ediyorum 🙂",
                "Kış geliyor havası aldım ama burada işler net ilerliyor 😄"
            ],
            "batman": [
                "🦇 Gotham havası sezdim 🙂 Ben Batman değilim ama yardımcı olabilirim."
            ],
            "harry_potter": [
                "🪄 Harry Potter evrenini sevdiğini fark ettim 🙂 Ama ben daha çok portal işlerinde güçlüyüm."
            ],
            "se7en": [
            "😅 O sahneyi hatırladım… Seven izleyenler genelde bunu kaçırmaz.",
            "🎬 Bu repliği yazan biri muhtemelen Seven izlemiştir. Merak uyandırıcı bir sahneydi. Peki sana nasıl yardımcı olabilirim🙂"
        ]
        }
        return random.choice(RESPONSES.get(category, []))
    
    
    def handle_fun_intent(self, user_message: str):

        pop_category = self.detect_pop_culture_category(user_message)
        if pop_category:
            return self.generate_pop_culture_response(pop_category)

        JOKES = [
            
            "😄 Kısa bir ara ver ki, bilgisayar senden önce pes etmesin.",
            "😄 Bir kahve molası ver, en azından kahve seni yarı yolda bırakmaz.",
            "😄 Ayağa kalkıp iki adım yürü, sandalyeyle aranı biraz soğut.",
            "😄 Bir mola ver, belki çözüm dönünce seni bekliyordur.",
            "😄 Kısa bir ara ver; işler kaçmıyor ama sen kaçabilirsin.",
            "😄 Bir su iç, bilgisayar değilsen yeniden başlatmaya gerek yok."
        ]

        return random.choice(JOKES)
     

    def handle_portal_info(self):

        RESPONSES = [
            (
                "YücePortal, Yüce Auto çalışanlarının şirket içi işleyişini kolaylaştırmak için "
                "geliştirilmiş kapsamlı bir dijital platformdur. Otopark rezervasyonu, haftalık "
                "çalışma takvimi işaretleme, raporlar ve  araç bildirimleri gibi birçok "
                "şirket içi modül bu portal üzerinden yönetilir."
            ),
            (
                "YücePortal; otopark rezervasyonu, haftalık takvim, raporlama ve araç bildirimleri "
                "gibi pek çok şirket içi sürecin tek bir çatı altında toplandığı Yüce Auto’ya "
                "özel bir iç portaldır. Amaç, günlük iş akışını sadeleştirmektir."
            ),
            (
                "YücePortal, çalışanların otopark rezervasyonu yapabildiği, haftalık çalışma "
                "takvimini işaretleyebildiği ve raporlara erişebildiği; bunun gibi birçok "
                "kurumsal modülü barındıran merkezi bir platformdur."
            ),
            (
                "YücePortal sayesinde çalışanlar; otopark, haftalık takvim, raporlar ve benzeri "
                "birçok şirket içi modülü tek bir ekran üzerinden hızlı ve pratik şekilde "
                "kullanabilir."
            ),
        ]

        return random.choice(RESPONSES)
    
    def handle_yuce_auto_reputation(self):
        responses = [
            (
                "Yüce Auto, uzun yıllardır otomotiv sektöründe faaliyet gösteren ve "
                "Škoda’nın Türkiye distribütörlüğünü yürüten köklü ve güvenilir bir firmadır."
            ),
            (
                "Yüce Auto; müşteri deneyimi, satış ve satış sonrası hizmetler alanlarında "
                "elde ettiği uluslararası başarılarla sektörde güçlü bir konuma sahiptir."
            ),
            (
                "1989’dan bu yana Škoda markasını Türkiye’de temsil eden Yüce Auto, "
                "istikrarlı büyümesi ve yaygın bayi ağıyla öne çıkan bir şirkettir."
            ),
            (
                "Yüce Auto, hem operasyonel gücü hem de müşteri memnuniyetine verdiği önemle "
                "otomotiv sektöründe saygın bir yere sahiptir."
            ),
            (
                "Yüce Auto; köklü geçmişi, güçlü organizasyonu ve uluslararası ödülleriyle "
                "güven veren bir şirkettir."
            )
        ]

        return random.choice(responses)
    
    def handle_skoda_reputation(self):
        responses = [
            (
                "Škoda hakkında kısaca bilgi vereyim: Václav Laurin ve Václav Klement "
                "1895 yılında Mladá Boleslav’da kendi işlerini kurdu. "
                "Bugün Škoda, dünyanın en köklü otomotiv markalarından biri olarak faaliyet göstermektedir."
            ),
            (
                "Škoda hakkında genel bir bilgi paylaşmak gerekirse, markanın kökeni "
                "1895 yılına dayanır ve uzun yıllara yayılan üretim tecrübesiyle "
                "otomotiv sektöründe güçlü bir konuma sahiptir."
            ),
            (
                "Škoda hakkında kısaca bilgi vereyim: 1895 yılında kurulan marka, "
                "mühendislik mirası ve köklü geçmişiyle otomotiv dünyasında "
                "önemli bir yere sahiptir."
            )
        ]

        return random.choice(responses)


    def get_department_list(self):
        conn = _mysql_conn()
        if conn is None:
            return []

        cursor = conn.cursor()

        cursor.execute("""
            SELECT DISTINCT TRIM(departman)
            FROM users1
            WHERE 
                is_active = 1
                AND departman IS NOT NULL
                AND TRIM(departman) <> ''
            ORDER BY departman
        """)

        rows = cursor.fetchall()

        cursor.close()
        conn.close()

        departments = [
            self._canonical_department_label(r[0])
            for r in rows
            if r and r[0]
        ]
        unique_departments = list(dict.fromkeys(d for d in departments if d))
        return sorted(unique_departments, key=self.normalize_people_text)
    def handle_department_list(self):
        departments = self.get_department_list()

        if not departments:
            return "Şu anda departman bilgilerine ulaşılamıyor."

        if len(departments) == 1:
            return f"Şirkette yalnızca <b>{departments[0]}</b> departmanı bulunuyor."

        # düzgün Türkçe liste
        if len(departments) == 2:
            dept_text = " ve ".join(departments)
        else:
            dept_text = ", ".join(departments[:-1]) + " ve " + departments[-1]

        return (
            f"Şirkette toplam <b>{len(departments)}</b> departman bulunuyor:<br>"
            f"{dept_text}."
        )
    
    def handle_charge_reservation_redirect(self):
        return (
            "🔌 Şu an için şarj rezervasyonu konusunda yardımcı olamıyorum.<br><br>"
            "Bu özellik ilerleyen dönemlerde geliştirilecektir.<br><br>"
            "Şimdilik Yüce Portal’daki "
            "<b><a href='https://yuceportal.skoda.com.tr/YA_otopark_rezerve'>"
            "şarj rezervasyonu sayfası</a></b> 🔗"
        )

    def detect_weather_day(self, msg: str) -> str:
        msg = self.normalize_people_text(msg)

        if any(w in msg for w in ALL_WEEKDAY_VARIANTS):
            return "weekday_unsupported"

        if "yarin" in msg or "yarinki" in msg:
            return "tomorrow"

        if "bugun" in msg:
            return "today"

        return "today"  # default
    
    def detect_weather_question_type(self, msg: str) -> str:
        msg = self.normalize_people_text(msg)

        if any(k in msg for k in ["kac derece", "kaç derece", "derece"]):
            return "temperature"

        if any(k in msg for k in ["soguk mu", "soğuk mu", "sicak mi", "sıcak mı","yağmurlumu","güneşlimi","soğukmu","sıcakmı"]):
            return "comfort"

        return "general"   # hava nasıl
    
    @staticmethod
    def weather_advice_tag(weathercode: int, temp: float) -> str:
        # 🔥 SICAKLIK BANTLARI (ÖNCELİKLİ)
        if temp < 0:
            return "freezing"
        if 0 <= temp < 10:
            return "very_cold"
        if 10 <= temp <= 25:
            # Ilıman ama yağmur/fırtına varsa onu öne al
            if weathercode in [61, 63, 80]:
                return "rain"
            if weathercode in [95, 99]:
                return "storm"
            return "mild"
        if temp > 25:
            return "hot"

        return "neutral"
    
    def handle_weather_redirect(self, user_message: str):
        try:
            msg_norm = self.normalize_people_text(user_message)

            city_key = None
            for c in self.cities.keys():
                if c in msg_norm:
                    city_key = c
                    break

            if not city_key:
                city_key = "istanbul"

            lat, lon = self.cities[city_key]

            day = self.detect_weather_day(user_message)
            if day == "weekday_unsupported":
                return (
                    "📅 Belirttiğin gün için hava durumu bilgisini şu an gösteremiyorum.<br><br>"
                    "Şimdilik sadece <b>bugün</b> ve <b>yarın</b> için hava durumunu paylaşabiliyorum."
                )

            url = (
                f"https://api.open-meteo.com/v1/forecast"
                f"?latitude={lat}&longitude={lon}&current_weather=true"
            )

            data = requests.get(url, timeout=5).json()
            weather = data.get("current_weather")

            if not weather:
                return "🌦️ Hava durumu bilgisi şu anda alınamıyor."

            temp = weather["temperature"]
            code = weather["weathercode"]

            desc_map = {
                0: "☀️ Açık",
                1: "🌤 Az bulutlu",
                2: "⛅ Parçalı bulutlu",
                3: "☁️ Bulutlu",
                45: "🌫 Sisli",
                48: "🌫 Sisli",
                51: "🌦 Hafif yağmurlu",
                61: "🌧 Yağmurlu",
                63: "🌧️ Kuvvetli yağmur",
                80: "🌧 Sağanak",
                95: "⛈ Gök gürültülü",
                99: "⛈ Fırtınalı"
            }
            desc = desc_map.get(code, "🌦 Belirsiz hava durumu")

            tag = self.weather_advice_tag(code, temp)
            question_type = self.detect_weather_question_type(user_message)

            return self.llm_weather_sentence(
                user_message=user_message,
                city=city_key.title(),
                temp=round(temp, 1),
                desc=desc,
                advice_tag=tag,
                question_type=question_type
            )

        except Exception as e:
            print("❌ Weather hata:", e)
            return (
                "🌦️ Hava durumu bilgisine şu anda ulaşılamıyor.<br>"
                "İnternet bağlantısı veya servis geçici olarak yanıt vermiyor olabilir."
            )
        
    def llm_weather_sentence(
        self,
        user_message: str,
        city: str,
        temp: float,
        desc: str,
        advice_tag: str,
        question_type: str
    ) -> str:
        """
        Gerçek hava durumu verilerinden,
        kontrollü ve halüsinasyonsuz tek cümle üretir.
        """
        print("🧠 LLM WEATHER ÇALIŞTI:", question_type, advice_tag)

        prompt = f"""
    Aşağıda GERÇEK hava durumu verileri ve
    sistem tarafından belirlenmiş bir öneri etiketi vardır.

    KURALLAR (ÇOK ÖNEMLİ):
    - SADECE verilen verileri kullan.
    - Yeni bilgi ekleme.
    - Tahmin yapma.
    - Gelecek zaman kullanma.
    - Tek cümle yaz.
    - Emoji en fazla 3 tane.
    - Öneri SADECE advice_tag'e göre yapılır.
    - advice_tag = neutral ise öneri ekleyebilirsin.
    - temperature soru tipinde emoji KULLANABİLİRSİN ama öneri  kısaca yapabilirsin genel.

    SORU TİPİ KURALLARI:
    - temperature → dereceyi vurgula
    - comfort → soğuk/sıcak yorumunu yap
    - general → genel hava durumunu anlat

    Advice tag kuralları:
    - freezing  → dondurucu soğuk vurgusu (buzlanma)
    - very_cold → çok soğuk, kalın giyin
    - mild      → ılıman, rahat hava
    - hot       → sıcak, hafif giyin / su için
    - rain      → şemsiye öner
    - storm     → dikkatli olun uyarısı
    - neutral   → genel, kısa yorum

    VERİLER:
    - Şehir: {city}
    - Sıcaklık: {temp}°C
    - Durum: {desc}
    - Advice tag: {advice_tag}
    - Soru tipi: {question_type}

    Kullanıcı mesajı:
    "{user_message}"
    """

        try:
            resp = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                temperature=0.4,
                max_tokens=60
            )
            return resp.choices[0].message.content.strip()

        except Exception as e:
            print("❌ LLM weather hata:", e)
            return f"{city} için hava durumu {desc}, {temp:.1f}°C."
        

    
    def handle_today_question(self, msg_norm: str) -> str:
        gunler = ["pazartesi","salı","çarşamba","perşembe","cuma","cumartesi","pazar"]

        # 🔑 bugün mü yarın mı?
        is_tomorrow = "yarin" in msg_norm
        offset = 1 if is_tomorrow else 0

        target_date = datetime.now() + timedelta(days=offset)
        day_name = gunler[target_date.weekday()]

        for canonical, variants in {
            "pazartesi": PAZARTESI_VARIANTS,
            "salı": SALI_VARIANTS,
            "çarşamba": CARSAMBA_VARIANTS,
            "perşembe": PERSEMBE_VARIANTS,
            "cuma": CUMA_VARIANTS,
            "cumartesi": CUMARTESI_VARIANTS,
            "pazar": PAZAR_VARIANTS,
        }.items():
            if any(v in msg_norm for v in variants):
                if canonical == day_name:
                    return f"📅 Evet, {'yarın' if is_tomorrow else 'bugün'} {day_name.title()} 😊"
                else:
                    return f"📅 Hayır, {'yarın' if is_tomorrow else 'bugün'} {day_name.title()} 😊"

        responses_today = [
            f"Bugün {day_name.title()}, güzel bir gün — çalışmaya devam edelim 😊",
            f"{day_name.title()}, ne güzel bir gün değil mi, tempoyu bozmayalım.",
            f"Bugün {day_name.title()}, gün güzel görünüyor, keyifle devam edelim.",
            f"{day_name.title()}, hafta akıyor; istersen hava durumuna da bakabilirim 🌦️",
            f"Bugün {day_name.title()}, enerjimiz yerindeyse devam edelim 💪",
            f"{day_name.title()}, günün havasını merak edersen sorman yeterli 😊",
        ]

        responses_tomorrow = [
            f"Yarın {day_name.title()}, şimdiden plan yapmak iyi fikir 🙂",
            f"Yarın {day_name.title()} görünüyor, hazırlıklı olabilirsin.",
            f"Yarın {day_name.title()}, güzel bir gün olacak gibi; istersen hava durumuna da bakabilirim 🌦️",
            f"Yarın {day_name.title()}, tempoya hazır olalım 💪",
        ]

        return "📅 " + random.choice(
            responses_tomorrow if is_tomorrow else responses_today
        )
    
    
    #handle_workforce_status tarih tespiti
    def parse_workforce_dates(self, message: str):
        """
        Workforce için TEK GÜN tarih parser.
        Öncelik sırası:
        1) bugün / yarın / öbür gün
        2) bu + gün / bu hafta + gün
        3) haftaya + gün
        4) sadece gün (en yakın ileri)
        5) net tarih (16 ocak / 16.01 / 16.01.2026)
        6) extract_history_dates fallback
        """
        print("\n🧪 [DATE PARSER] INPUT:", message)

        msg = self.normalize_people_text(message)
        msg = self.normalize_month_with_suffixes(msg)


        # ŞU AN = BUGÜN
        msg = re.sub(
        r"(ofis(?:ten|den|te|de)?)(mi|mı|mu|mü)\b",
        r"\1 \2",
        msg
        )
        msg = re.sub(r"(mi|mı|mu|mü)\b", r" \1", msg)

        #zaman referansı
        today = datetime.now().date()
        current_year = today.year

        #suffix strip (SORU TESPİTİ İÇİN)
        msg_stripped = self._strip_tr_suffixes(msg)

        has_question_particle = (
            bool(re.search(r"(mi|mı|mu|mü)\b", msg))
            or bool(re.search(r"(mi|mı|mu|mü)\b", msg_stripped))
        )
        #  ŞU AN = BUGÜN (SADECE SORU VARSA)
        if has_question_particle and any(k in msg for k in ["suan", "su an"]):
            return {
                "start": today,
                "end": today,
                "mode": "today"
            }
        
        #  BUGÜN / YARIN / ÖBÜR GÜN
        if "bugun" in msg:
            return {"start": today, "end": today, "mode": "today"}
        
        if "yarin" in msg:
            d = today + timedelta(days=1)
            return {"start": d, "end": d, "mode": "future"}

        if "obur gun" in msg or "öbür gun" in msg:
            d = today + timedelta(days=2)
            return {"start": d, "end": d, "mode": "future"}
        
        #  GÜN HARİTASI (VARIANT → WEEKDAY INDEX)
        DAY_VARIANT_MAP = {}

        for v in PAZARTESI_VARIANTS:
            DAY_VARIANT_MAP[v] = 0
        for v in SALI_VARIANTS:
            DAY_VARIANT_MAP[v] = 1
        for v in CARSAMBA_VARIANTS:
            DAY_VARIANT_MAP[v] = 2
        for v in PERSEMBE_VARIANTS:
            DAY_VARIANT_MAP[v] = 3
        for v in CUMA_VARIANTS:
            DAY_VARIANT_MAP[v] = 4
        for v in CUMARTESI_VARIANTS:
            DAY_VARIANT_MAP[v] = 5
        for v in PAZAR_VARIANTS:
            DAY_VARIANT_MAP[v] = 6

        has_day_name = any(v in msg for v in DAY_VARIANT_MAP)
        has_month_name = (
            any(k in msg for k in AY_NORMALIZE.keys())
            or any(m in msg for m in AY_NORMALIZE.values())
        )
        has_day_number = bool(re.search(r"\b\d{1,2}\b", msg))

        #PERIOD ANCHOR VAR + GÜN YOK → TEK GÜN PARSER DEVRE DIŞI
        has_period_anchor = any(k in msg for k in (
            THIS_WEEK_KEYS + PAST_WEEK_KEYS + THIS_MONTH_KEYS
        ))

        has_explicit_day = (
            any(v in msg for v in DAY_VARIANT_MAP)      
            or bool(re.search(r"\b\d{1,2}\b", msg))     
        )

        if has_period_anchor and not has_explicit_day:
            return {"start": None, "end": None, "mode": None}


        for variant, weekday_idx in DAY_VARIANT_MAP.items():
            if variant in msg:

                base_monday = today - timedelta(days=today.weekday())

                # ⏪ GEÇEN HAFTA
                if any(k in msg for k in PAST_WEEK_KEYS):
                    base_monday -= timedelta(days=7)

                # ⏩ HAFTAYA / GELECEK HAFTA
                elif any(k in msg for k in NEXT_WEEK_KEYS):
                    base_monday += timedelta(days=7)

                # 🟡 AKSİ HALDE → bu hafta
                target = base_monday + timedelta(days=weekday_idx)

                return {
                    "start": target,
                    "end": target,
                    "mode": "past" if target < today else "future" if target > today else "today"
                }


        #  NET TARİH — NUMERİK (16.01 / 16.01.2026)
        match_numeric = re.search(r"\b(\d{1,2})\.(\d{1,2})(?:\.(\d{4}))?\b", msg)
        if match_numeric:
            day = int(match_numeric.group(1))
            month = int(match_numeric.group(2))
            year = int(match_numeric.group(3)) if match_numeric.group(3) else current_year
            try:
                d = date(year, month, day)
                if d < today:
                    mode = "past"
                elif d > today:
                    mode = "future"
                else:
                    mode = "today"
                return {"start": d, "end": d, "mode": mode}
            except ValueError:
                pass
        #  NET TARİH — AY YAZILI (16 ocak)
        for raw, normalized in AY_NORMALIZE.items():
            msg = re.sub(rf"\b{raw}\b", normalized, msg)

        match_textual = re.search(
            r"\b(\d{1,2})\s+"
            r"(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)"
            r"(?:\s+(\d{4}))?\b",
            msg
        )

        if match_textual:
            day = int(match_textual.group(1))
            month_name = match_textual.group(2)
            year = int(match_textual.group(3)) if match_textual.group(3) else current_year

            MONTH_NUM = {
                "ocak": 1, "şubat": 2, "mart": 3, "nisan": 4,
                "mayıs": 5, "haziran": 6, "temmuz": 7, "ağustos": 8,
                "eylül": 9, "ekim": 10, "kasım": 11, "aralık": 12
            }

            try:
                d = date(year, MONTH_NUM[month_name], day)
                if d < today:
                    mode = "past"
                elif d > today:
                    mode = "future"
                else:
                    mode = "today"
                return {"start": d, "end": d, "mode": mode}
            except ValueError:
                pass

        #  AY + GÜN (TERS SIRA) — ocak 15 / ocakta 15
        reverse_match = re.search(
            r"\b(ocak|şubat|mart|nisan|mayıs|haziran|temmuz|ağustos|eylül|ekim|kasım|aralık)\s+(\d{1,2})\b",
            msg
        )
        if reverse_match:
            month_name = reverse_match.group(1)
            day = int(reverse_match.group(2))

            MONTH_NUM = {
                "ocak": 1, "şubat": 2, "mart": 3, "nisan": 4,
                "mayıs": 5, "haziran": 6, "temmuz": 7, "ağustos": 8,
                "eylül": 9, "ekim": 10, "kasım": 11, "aralık": 12
            }

            try:
                d = date(current_year, MONTH_NUM[month_name], day)
                if d < today:
                    mode = "past"
                elif d > today:
                    mode = "future"
                else:
                    mode = "today"
                return {"start": d, "end": d, "mode": mode}
            except ValueError:
                pass
        
        # 2️⃣ BU / BU HAFTA + GÜN
        is_this_week = any(p in msg for p in BU_WEEK_DAY_PREFIXES)

        for variant, weekday_idx in DAY_VARIANT_MAP.items():
            if variant in msg and is_this_week:
                base_monday = today - timedelta(days=today.weekday())
                target = base_monday + timedelta(days=weekday_idx)

                # bu hafta denmiş ama gün geçmişse → GEÇMİŞ
                mode = (
                    "today" if target == today
                    else "future" if target > today
                    else "past"
                )

                return {"start": target, "end": target, "mode": mode}
        
        is_past_query = any(k in msg for k in [
            "calisti", "calismis", "ofisteydi", "geldi", "oldu", "yapildi","evdeydi","uzaktan çalıştı","remote","calıstı"
        ])

        is_future_query = any(k in msg for k in [
            "calisacak", "gelecek", "olacak", "olucak",
            "bulunacak", "bulunucak", "gelicek", "çalisacak","isaretlemis"
        ])

        is_present_query = any(k in msg for k in [
            "calisiyor", "çalışıyor",
            "nerede", "nerden","nerde","nrde","isaretlemis"
        ])
        has_person = bool(
            self.detect_person_prescan(message)
            or self.has_person_in_db(message)
        )
        if (
            has_person 
            and (is_present_query or has_question_particle)
            and not is_past_query
            and not is_future_query
            and not has_day_name
            and not has_month_name
            and not has_day_number
        ):
            return {
                "start": today,
                "end": today,
                "mode": "today"
            }

        # HAFTAYA + GÜN
        if "haftaya" in msg or "gelecek hafta" in msg:
            for variant, weekday_idx in DAY_VARIANT_MAP.items():
                if variant in msg:
                    base_monday = today - timedelta(days=today.weekday())
                    target = base_monday + timedelta(days=7 + weekday_idx)

                    #  güvenlik: 14 günden ileri gitmesin
                    if (target - today).days > 14:
                        target = base_monday + timedelta(days=7)

                    return {"start": target, "end": target, "mode": "future"}

        
        print("🧪 [DATE PARSER] NO MATCH → None")


        return {"start": None, "end": None, "mode": None}

    def format_workforce_day_label(self, target_date: date, today: date):
        days = ["Pazartesi","Salı","Çarşamba","Perşembe","Cuma","Cumartesi","Pazar"]
        day_name = days[target_date.weekday()]

        if target_date == today:
            prefix = "Bugün"
        elif target_date == today + timedelta(days=1):
            prefix = "Yarın"
        else:
            prefix = day_name

        return f"{prefix} ({target_date.strftime('%d.%m.%Y')} {day_name})"

    def is_workforce_period_query(self, msg_norm: str, date_info: dict):
        """
        Workforce için PERIOD sorgusu mu?
        Eğer PERIOD ise → (True, start, end)
        Değilse        → (False, None, None)
        """
        today = date.today()
        #  GÜN VARSA → ASLA PERIOD DEĞİL
        if any(d in msg_norm for d in ALL_WEEKDAY_VARIANTS):
            return False, None, None

        if re.search(r"\b\d{1,2}\b", msg_norm):  # 6, 7, 15 ocak
            return False, None, None
        
        #  BU HAFTA
        if any(k in msg_norm for k in THIS_WEEK_KEYS):
            monday = today - timedelta(days=today.weekday())
            friday = monday + timedelta(days=4)
            return True, monday, friday

        # GEÇEN HAFTA
        if any(k in msg_norm for k in PAST_WEEK_KEYS):
            last_monday = today - timedelta(days=today.weekday() + 7)
            return True, last_monday, last_monday + timedelta(days=4)

        #  BU AY
        if any(k in msg_norm for k in THIS_MONTH_KEYS):
            start = today.replace(day=1)
            end_day = calendar.monthrange(today.year, today.month)[1]
            return True, start, today.replace(day=end_day)

        #  GEÇEN AY
        if any(k in msg_norm for k in PAST_MONTH_KEYS):
            first_this_month = today.replace(day=1)
            last_month_end = first_this_month - timedelta(days=1)
            start = last_month_end.replace(day=1)
            return True, start, last_month_end

        #  AY ADI (ocak, şubat, …) – GÜN YOKSA
        for raw, normalized in AY_NORMALIZE.items():
            if raw in msg_norm or normalized in msg_norm:
                month_map = {
                    "ocak": 1, "şubat": 2, "mart": 3, "nisan": 4,
                    "mayıs": 5, "haziran": 6, "temmuz": 7,
                    "ağustos": 8, "eylül": 9, "ekim": 10,
                    "kasım": 11, "aralık": 12
                }
                m = month_map.get(normalized)
                if m:
                    start = date(today.year, m, 1)
                    end = date(today.year, m, calendar.monthrange(today.year, m)[1])
                    return True, start, end

        #  PERIOD DEĞİL none dönmeli
        return False, None, None

    def handle_workforce_period_summary(self, user_message, start, end, msg_norm):

        if not start or not end:
            return "Bu tarih aralığını net anlayamadım."

        period_label = f"{start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}"

        work_type = self.detect_work_type(user_message)
        # KULLANICIYA GÖSTERİLECEK LABEL
        if work_type == "Ofis":
            label = "ofiste"
        elif work_type == "Home":
            label = "evden"
        elif work_type == "Bayi":
            label = "sahada"
        elif work_type == "Izin":
            label = "izinli"
        else:
            label = "şirkette"   

        conn = _mysql_conn()
        cursor = conn.cursor()

        if work_type is None:
            #  Kullanıcı çalışma tipi söylememiş → TÜM KAYITLAR
            cursor.execute("""
                SELECT COUNT(DISTINCT users1_id)
                FROM ofis_gunleri
                WHERE selected_date BETWEEN %s AND %s
            """, (start, end))

        elif work_type == "Izin":
            cursor.execute("""
                SELECT COUNT(DISTINCT users1_id)
                FROM ofis_gunleri
                WHERE selected_date BETWEEN %s AND %s
                AND (ofis_bayi IS NULL OR ofis_bayi = '')
            """, (start, end))

        else:
            #  Ofis / Home / Bayi
            cursor.execute("""
                SELECT COUNT(DISTINCT users1_id)
                FROM ofis_gunleri
                WHERE selected_date BETWEEN %s AND %s
                AND ofis_bayi = %s
            """, (start, end, work_type))

        count = cursor.fetchone()[0]
        cursor.close()
        conn.close()

    
        return (
            f"📅 <b>{period_label}</b> aralığında {label} olan çalışanlar "
            f"günlere göre değişiklik gösteriyor.<br>"
            f"<b>Özet:</b><br>"
            f"• Bu dönemde toplam <b>{count}</b> farklı çalışan en az bir gün {label} bulunmuş.<br>"
            f"<br>"
            f"📆 Gün bazında görmek ister misin?<br>"
            f"Örnekler:<br>"
            f"• “ocak ayı pazartesi kimler ofiste”<br>"
            f"• “cuma kimler ofiste”"
        )


    #şirkete bugün kaç kişi gelicek yada isim bugün ofistemi olucak?
    def handle_workforce_status(self, user_message: str, user_info: dict):
        """
        WORKFORCE STATUS HANDLER
        - Kişi / Departman / Şirket bazlı
        - Ofis / Home / Bayi / İzin
        - Count / List / Status
        """

        print("\n🧪 [WORKFORCE] RAW INPUT:", user_message)

        msg = self.normalize_people_text(user_message)
        today = datetime.now().date()

        # TARİH ANALİZİ
        date_info = self.parse_workforce_dates(user_message)
        is_period, p_start, p_end = self.is_workforce_period_query(msg, date_info)

        if is_period:
            return self.handle_workforce_period_summary(
                user_message=user_message,
                start=p_start,
                end=p_end,
                msg_norm=msg
            )

        # MODE KULLANMIYORUZ
        start = date_info.get("start") or today
        end   = date_info.get("end") or today

        print("📅 [WORKFORCE] start:", start, "end:", end)

        # 3 AY SINIRI (mode bağımsız)
        if abs((today - start).days) > 90:
            return "En fazla 3 ay aralığında sorgulama yapabilirim."

        # ÇALIŞMA TİPİ
        work_type = self.detect_work_type(user_message)
        print("🏢 [WORKFORCE] work_type:", work_type)

        persons = self.detect_person_prescan(user_message)
        department = self.extract_department_name(user_message)

        if (
            not persons
            and self.normalize_people_text(user_info.get("fullname","").split()[0]) 
                in self.normalize_people_text(user_message).split()
        ):
            persons = [user_info.get("fullname")]

        if self.has_person_in_db(user_message) and not persons:
            return (
                "Bu isimle eşleşen birden fazla veya belirsiz çalışan buldum. "
                "Lütfen isim ve soyisimle birlikte tekrar sorabilir misin?"
            )

        if persons:
            scope = "person"
        elif department:
            scope = "department"
        else:
            scope = "company"



        print("👥 [WORKFORCE] scope:", scope)

        if scope == "person" and len(persons) > 1:
            names = ", ".join(persons)
            return (
                f"Birden fazla kişi buldum: {names}.<br>"
                "Lütfen isim ve soyisimle birlikte tekrar sorabilir misiniz?"
            )

        # SORU TİPİ
        STATUS_KEYWORDS = [
            "ne isaretlemis",
            "ne isaretledi",
            "ne yazmis",
            "ne girmis",
            "nerede",
            "nerden",
            "nereden",
            "nereden calisacak",
            "nerede olacak",
            "ofiste mi",
            "evde mi",
            "bayide mi",
            "izinli mi"          
        ]

        if any(k in msg for k in ["kim", "kimler","kim","kimler","kişiler","çalışanlar","calisanlar","çalışan"]):
            query_type = "list"
        elif any(k in msg for k in STATUS_KEYWORDS):
            query_type = "status"
        elif self.has_work_history_count_signal(msg):
            query_type = "count"
        else:
            query_type = "status"

        if persons and query_type != "list":
            query_type = "status"

        print("❓ [WORKFORCE] query_type:", query_type)

        # 🔥 SELF STATUS FIX (GENEL)
        if (
            scope == "company"
            and query_type == "status"
            and not persons
            and not department
        ):
            persons = [user_info.get("fullname")]
            scope = "person"

        # 🔥 PERSON + STATUS sorgusunda filtre uygulama
        if scope == "person" and query_type == "status":
            work_type = None

        if scope == "person" and not persons:
            return (
                "Bu isimle eşleşen bir çalışan bulamadım. "
                "İsim ve soyisimle birlikte tekrar sorabilir misin?"
            )

        conn = _mysql_conn()
        if not conn:
            return "⚠️ Şu anda çalışma bilgilerine erişemiyorum."

        cursor = conn.cursor(dictionary=True)

        base_sql = """
            FROM ofis_gunleri og
            INNER JOIN users1 u ON u.id = og.users1_id
            WHERE og.selected_date BETWEEN %s AND %s
        """
        params = [start, end]

        if work_type:
            if work_type == "Izin":
                base_sql += " AND (og.ofis_bayi IS NULL OR og.ofis_bayi = '')"
            else:
                base_sql += " AND og.ofis_bayi = %s"
                params.append(work_type)

        if scope == "person":
            name, surname = persons[0].split(" ", 1)
            base_sql += " AND LOWER(u.name) = %s AND LOWER(u.surname) = %s"
            params.extend([name.lower(), surname.lower()])

        print("DEBUG persons:", persons)
        print("DEBUG fullname:", user_info.get("fullname"))

        # COUNT
        if query_type == "count":
            if scope == "department":
                sql = "SELECT u.username, u.name, u.surname, u.departman " + base_sql
                cursor.execute(sql, params)
                self.auto_log_sql(sql, params)
                rows = self._apply_directory_overrides_to_rows(cursor.fetchall())
                count = sum(
                    1
                    for r in rows
                    if self._canonical_department_label(r.get("departman")) == department
                )
            else:
                sql = "SELECT COUNT(*) AS cnt " + base_sql
                cursor.execute(sql, params)
                self.auto_log_sql(sql, params)
                count = cursor.fetchone()["cnt"]
            cursor.close()
            conn.close()

            label = (
                "izinli"
                if work_type == "Izin"
                else work_type.lower()
                if work_type
                else "çalışan"
            )

            day_label = self.format_workforce_day_label(start, today)

            if scope == "department":
                return (
                    f"{day_label} <b>{department}</b> departmanında <b>"
                    f"{label} olan toplam <b>{count}</b> kişi bulunuyor."
                )

            elif scope == "company":
                return (
                    f"{day_label} şirkette <b>"
                    f"{label} olan toplam <b>{count}</b> kişi bulunuyor."
                )

            else:
                return (
                    f"{day_label} {label} olan toplam <b>{count}</b> kişi bulunuyor."
                )

        # LIST
        if query_type == "list":

            is_department_only = (scope == "department")

            sql = "SELECT u.username, u.name, u.surname, u.departman " + base_sql + " ORDER BY u.departman, u.name"
            cursor.execute(sql, params)
            self.auto_log_sql(sql, params)
            rows = self._apply_directory_overrides_to_rows(cursor.fetchall())
            cursor.close()
            conn.close()

            if scope == "department":
                rows = [
                    r for r in rows
                    if self._canonical_department_label(r.get("departman")) == department
                ]

            if not rows:
                return "Bu kriterlere uygun çalışan bulunamadı."

            grouped = {}

            if is_department_only:
                grouped[department] = [
                    r.get("display_name") or f"{r['name']} {r['surname']}" for r in rows
                ]
            else:
                for r in rows:
                    dept = (
                        r.get("display_department")
                        or (r["departman"] or "Departman belirtilmemiş").strip()
                    )
                    name_text = r.get("display_name") or f"{r['name']} {r['surname']}"
                    grouped.setdefault(dept, []).append(name_text)

            lines = []

            for dept, people in grouped.items():
                people_str = ", ".join(people)
                lines.append(f"<b>{dept}</b> ({len(people)}): {people_str}")

            day_label = self.format_workforce_day_label(start, today)

            if work_type == "Home":
                emoji = "🏠"
                label = "evden çalışanlar"
            elif work_type == "Bayi":
                emoji = "🚗"
                label = "sahada olanlar"
            elif work_type == "Izin":
                emoji = "🌴"
                label = "izinli olanlar"
            elif work_type == "Ofis":
                emoji = "🏢"
                label = "ofiste olanlar"
            else:
                emoji = "👥"
                label = "çalışanlar"

            if scope == "department":
                title = f"{emoji} {day_label} {department} – {label}"
            else:
                title = f"{emoji} {day_label} {label}"

            return "<br>".join([title, ""] + lines)

        # STATUS (tek kişi)
        sql = "SELECT og.ofis_bayi " + base_sql + " LIMIT 1"
        cursor.execute(sql, params)
        self.auto_log_sql(sql, params)
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        person_name = persons[0] if persons else "Kişi"
        day_label = self.format_workforce_day_label(start, today)

        if not row:
            return f"{person_name} için {day_label} herhangi bir çalışma durumu işaretlenmemiş."

        wt = self.normalize_db_work_type(row["ofis_bayi"])

        is_explicit_izin_mark = any(k in msg for k in [
            "izin isaretle","izinliydi","izinli","izin yapmış","izinlimiydi","izinmi almış",
            "izin almış","izinlimiydi","izinmi aldı","izin işaretlemiş",
            "izin işaretle", "izinli işaretleyenle","izin yazanlar",
            "izin giren","izin girenler","izin almis"
        ])

        if wt == "Izin":
            if is_explicit_izin_mark:
                wt_text = "izin işaretlemiş olarak gözükmekte."
            else:
                wt_text = "izinli"
        elif wt == "Ofis":
            wt_text = "ofisten çalışıyor olarak gözükmekte."
        elif wt == "Home":
            wt_text = "evden çalışıyor olarak gözükmekte."
        elif wt == "Bayi":
            wt_text = "bayide çalışıyor olarak gözükmekte."
        else:
            wt_text = "çalışma tipi tanımsız olarak gözükmekte."

        return f"{person_name}, {day_label} {wt_text}."

    def is_simple_smalltalk(self, message: str) -> bool:
        if not message:
            return False

        msg = self.normalize_people_text(message)

        SMALLTALK_KEYS = [
            "merhaba", "selam", "slm", "hello", "hi",
            "naber", "nbr", "nasilsin", "nasılsın",
            "iyi misin", "ne haber",
            "tesekkur", "teşekkür", "sagol", "sağol"
        ]
        return any(k in msg for k in SMALLTALK_KEYS)

    def is_password_guide_question(self, msg: str) -> bool:
        text = self.normalize_people_text(msg)

        PASSWORD_GUIDE_KEYWORDS = [
            "nasil",
            "nasıl",
            "degistir",
            "değiştir",
            "guncelle",
            "güncelle",
            "sifremi degistir",
            "sifremi değistir",
            "sifre nasil",
            "sifre degistirme",
            "reset",
            "yenile"
        ]

        PASSWORD_CONTEXT = [
            "sifre","sifrem","pc sifrem","pc sifremi","bilgisayar sifremi","pc sfre","pc sifresi",
            "sifrem",
            "sifremi",
            "password",
            "pc sifre",
            "pc sifrem",
            "bilgisayar sifre",
            "bilgisayar sifrem"
        ]
        return (
            any(k in text for k in PASSWORD_GUIDE_KEYWORDS)
            and any(p in text for p in PASSWORD_CONTEXT)
        )


    def get_trend_analysis(self):

        conn = _mysql_conn()
        cursor = conn.cursor(dictionary=True)

        query = """
        SELECT
            DATE_FORMAT(selected_date, '%Y-%m') AS ay,
            COUNT(*) AS toplam,
            SUM(ofis_bayi = 'Ofis') AS ofis,
            SUM(ofis_bayi = 'Home') AS home,
            SUM(ofis_bayi = 'Bayi') AS bayi,
            SUM(ofis_bayi IS NULL) AS izin
        FROM ofis_gunleri
        WHERE selected_date >= DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 1 MONTH), '%Y-%m-01')
        GROUP BY ay
        ORDER BY ay;
        """

        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        if len(rows) < 2:
            return {"error": "Yeterli veri yok"}

        def calc_ratio(val, total):
            if total == 0:
                return 0
            return round((val / total) * 100, 1)

        last = rows[0]
        this = rows[1]

        return {
            "this_month": {
                "ay": this["ay"],
                "Ofis": calc_ratio(this["ofis"], this["toplam"]),
                "Home": calc_ratio(this["home"], this["toplam"]),
                "Bayi": calc_ratio(this["bayi"], this["toplam"]),
                "Izin": calc_ratio(this["izin"], this["toplam"]),
            },
            "last_month": {
                "ay": last["ay"],
                "Ofis": calc_ratio(last["ofis"], last["toplam"]),
                "Home": calc_ratio(last["home"], last["toplam"]),
                "Bayi": calc_ratio(last["bayi"], last["toplam"]),
                "Izin": calc_ratio(last["izin"], last["toplam"]),
            }
        }
    
    def format_trend_response(self, data):

        if "error" in data:
            return "Aylık kıyas raporu için yeterli veri bulunamadı."

        this_m = data["this_month"]
        last_m = data["last_month"]

        prompt = f"""
        Aşağıdaki iki aylık çalışma dağılım verisini değerlendir.

        {this_m['ay']}:
        Ofis: %{this_m['Ofis']}
        Home: %{this_m['Home']}
        Bayi: %{this_m['Bayi']}
        Izin: %{this_m['Izin']}

        {last_m['ay']}:
        Ofis: %{last_m['Ofis']}
        Home: %{last_m['Home']}
        Bayi: %{last_m['Bayi']}
        Izin: %{last_m['Izin']}

        Kurumsal ve sade bir değerlendirme yaz.
        'kanal', 'artış', 'düşüş' kelimelerini kullanma.
        Yeni sayı üretme. en fazla 3 cümle net kısa olsun.
        """

        response = self.client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[
                {"role": "system", "content": "Kurumsal ve net bir yönetici değerlendirmesi yaz."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2
        )

        comment = response.choices[0].message.content[:300]

        return f"""
    📊 <b>Aylık Çalışma Kıyas Raporu</b><br><br>

    <b>{this_m['ay']}:</b><br>
    🏢 Ofis: %{this_m['Ofis']}<br>
    🏠 Home: %{this_m['Home']}<br>
    🏢 Bayi: %{this_m['Bayi']}<br>
    🌴 İzin: %{this_m['Izin']}<br><br>

    <b>{last_m['ay']}:</b><br>
    🏢 Ofis: %{last_m['Ofis']}<br>
    🏠 Home: %{last_m['Home']}<br>
    🏢 Bayi: %{last_m['Bayi']}<br>
    🌴 İzin: %{last_m['Izin']}<br><br>

    ⚡ <b>Analiz sonucu:</b><br>
    {comment}
    """
    
    all_cities = {
        "adana","adiyaman","afyonkarahisar","agri","amasya","ankara","antalya",
        "artvin","aydin","balikesir","bilecik","bingol","bitlis","bolu","burdur",
        "bursa","canakkale","cankiri","corum","denizli","diyarbakir","edirne",
        "elazig","erzincan","erzurum","eskisehir","gaziantep","giresun","gumushane",
        "hakkari","hatay","isparta","mersin","istanbul","izmir","kars","kastamonu",
        "kayseri","kirklareli","kirsehir","kocaeli","konya","kutahya","malatya",
        "manisa","kahramanmaras","mardin","mugla","mus","nevsehir","nigde",
        "ordu","rize","sakarya","samsun","siirt","sinop","sivas","tekirdag",
        "tokat","trabzon","tunceli","sanliurfa","usak","van","yozgat","zonguldak",
        "aksaray","bayburt","karaman","kirikkale","batman","sirnak","bartin",
        "ardahan","igdir","yalova","karabuk","kilis","osmaniye","duzce"
    }


    def detect_city_from_text(self, msg_norm):

        print("CITY FUNC INPUT:", msg_norm)
        words = re.findall(r"\w+", msg_norm)
        print("ALL CITIES SAMPLE:", list(self.all_cities)[:10])
        print("MSG NORM:", msg_norm)
        print("ankara in all_cities:", "ankara" in self.all_cities)
        for word in words:

            for city in self.all_cities:

                # direkt eşleşme
                if word == city:
                    return city

                # ekli versiyon (ankara + da/de)
                if word.startswith(city):
                    return city

                # fuzzy eşleşme (ankarda → ankara)
                if difflib.SequenceMatcher(None, word, city).ratio() > 0.85:
                    return city

        return None
    
    def get_all_dealers(self, force_refresh=False):

        # 10 dakika cache
        if (
            not force_refresh
            and self._dealer_cache
            and self._dealer_cache_time
            and datetime.now() - self._dealer_cache_time < timedelta(minutes=10)
        ):
            return self._dealer_cache

        conn = self._oracle_conn()
        if not conn:
            return []

        QUERY = """
        SELECT
            YS.FK_EDUCATION_FIRM_ID AS FIRM_ID,
            YS.FIRM_CODE,
            MAX(CASE WHEN DR.REGION_TYPE = 'S' THEN DRC.REGION_CHIEF_FULL_NAME END),
            MAX(CASE WHEN DR.REGION_TYPE = 'L' THEN DRC.REGION_CHIEF_FULL_NAME END),
            CC.CITY_NAME
        FROM
            YUCE_DM.TBL_DW_GEN_DIM_DEALER YS
            INNER JOIN YUCE_DM.TBL_DW_GEN_DIM_DEALER_REGION DR
                ON YS.DEALER_ID = DR.FK_DEALER_ID
            AND DR.REGION_TYPE IN ('S','L')
            INNER JOIN YUCE_DM.TBL_DW_GEN_DIM_DEALER_REGION_CHIEF DRC
                ON DRC.FK_REGION_DEALER_ID = DR.REGION_DEALER_ID
            INNER JOIN YUCE_DM.TBL_DW_GEN_DIM_FIRM GEN_FIRM
                ON GEN_FIRM.EDUCATION_FIRM_ID = YS.FK_EDUCATION_FIRM_ID
            LEFT JOIN YUCE_DM.TBL_DW_GEN_DIM_CITY_COUNTY CC
                ON CC.CODE = YS.FK_CITY_CODE
        WHERE
            YS.STATUS_EXPLANATION = 'Yürürlükte'
            AND YS.FIRM_CODE NOT IN ('YÜCE AUTO')
        GROUP BY
            YS.FK_EDUCATION_FIRM_ID, YS.FIRM_CODE, CC.CITY_NAME
        ORDER BY
            YS.FIRM_CODE ASC
        """

        try:
            cursor = conn.cursor()
            cursor.execute(QUERY)
            rows = cursor.fetchall()

            dealers = []
            for r in rows:
                dealers.append({
                    "firm_id": r[0],
                    "firm_code": r[1],
                    "sales_manager": r[2],
                    "ssh_manager": r[3],
                    "city": r[4]
                })

            # cache'e yaz
            self._dealer_cache = dealers
            self._dealer_cache_time = datetime.now()

            return dealers

        except Exception as e:
            print("❌ Dealer sorgu hatası:", e)
            return []

        finally:
            cursor.close()
            conn.close()

    def get_satis_count_by_year(self, year: int):
        conn = self._oracle_conn()
        if not conn:
            return None

        cursor = conn.cursor()

        sql = """
            SELECT COUNT(DISTINCT VEHICLE_ID)
            FROM YUCE_DM.TBL_DW_VEH_RPT_VEHICLE_STATUS_UP_TO_DATE D
            WHERE EXTRACT(YEAR FROM D.AD_INVOICE_DATE) = :year
        """

        cursor.execute(sql, {"year": year})
        result = cursor.fetchone()[0]

        cursor.close()
        conn.close()

        return result

    def handle_satis_count(self, msg: str):

        current_year = datetime.now().year
        years = []

        # Yıl yakalama (2025, 2026 vs.)
        found_years = re.findall(r"20\d{2}", msg)

        if found_years:
            years = [int(y) for y in found_years]

        elif "gecen yil" in msg or "geçen yıl" in msg:
            years = [current_year - 1]

        else:
            # Direkt "kaç satış yaptık" → bu yıl + geçen yıl
            years = [current_year, current_year - 1]

        responses = []

        for year in years:
            count = self.get_satis_count_by_year(year)

            if count is not None:
                responses.append(
                    f"🚗 {year} yılına baktığımızda, toplamda {count:,} adet araç satışı gerçekleştirmişiz. "
                    f"Oldukça güçlü bir performans görünüyor."
                )
            else:
                responses.append(
                    f"🚗 {year} yılı için satış verisine şu anda ulaşılamadı."
                )

        return "<br><br>".join(responses)

    def hanlde_bayi_about(self, user_message):

        msg = self.normalize_tr(user_message)

        COUNT_KEYS = [
            "kaç", "kac",
            "adet", "tane",
            "toplam", "total",
            "sayisi", "sayısı","adeti","adeti","syısı"
        ]

        # 1️⃣ Önce COUNT kontrol edilir
        if any(k in msg for k in COUNT_KEYS):
            return self._handle_bayi_count(msg)

        # 2️⃣ Sonra liste intenti
        if "bayi" in msg:
            return self._handle_bayi_list_with_count(msg)

        return "Bayi ile ilgili isteğini biraz daha net yazabilir misin?"
    
    def normalize_tr(self, text: str) -> str:
        if not text:
            return ""

        text = text.lower()

        replacements = {
            "ı": "i",
            "i̇": "i",
            "ş": "s",
            "ç": "c",
            "ğ": "g",
            "ü": "u",
            "ö": "o",
        }

        for k, v in replacements.items():
            text = text.replace(k, v)

        return text.strip()
    
    def _handle_bayi_list_with_count(self, msg):

        data = self.get_all_dealers()
        if not data:
            return "Bayi verisi alınamadı."

        msg_norm = self.normalize_tr(msg)

        city_found = None

        for d in data:
            city_raw = d.get("city")
            if not city_raw:
                continue

            city_norm = self.normalize_tr(city_raw)

            pattern = rf"\b{re.escape(city_norm)}(da|de|daki|deki)?\b"

            if re.search(pattern, msg_norm):
                city_found = city_raw
                break

        if not city_found:
            return "Hangi şehir için bayi listesi istediğini belirtir misin?"

        filtered = [
            d for d in data
            if d["city"] == city_found
        ]

        count = len(filtered)

        if count == 0:
            return f"📍 {city_found} ilinde aktif bayi bulunmamaktadır."

        firm_list = sorted([
            d["firm_code"] for d in filtered
            if d["firm_code"]
        ])

        response = f"📍 {city_found} ilinde {count} adet aktif bayi bulunmaktadır.\n\n"

        for firm in firm_list:
            response += f"- {firm}\n"

        return response
    
    def _handle_bayi_count(self, msg):

        data = self.get_all_dealers()
        if not data:
            return "Bayi verisi alınamadı."

        msg_norm = self.normalize_tr(msg)

        # 81 il üzerinden şehir yakala
        detected_city = self.detect_city_from_text(msg_norm)
        print("DETECTED CITY:", detected_city)
        # DB şehir seti
        db_cities = {
            self.normalize_tr(d["city"]).strip()
            for d in data if d["city"]
        }

        # -------------------------
        # ŞEHİR YAZILMIŞ
        # -------------------------
        if detected_city:

            if self.normalize_tr(detected_city) not in db_cities:
                return (
                    f"📍 {detected_city.title()} ilinde aktif bayimiz bulunmamaktadır.\n\n"
                    "Yetkili satıcı ve servis noktalarımızı aşağıdaki bağlantıdan kontrol edebilirsiniz:\n"
                    "🔗 https://www.skoda.com.tr/yetkili-satici-ve-servisler"
                )

            count = len([
                d for d in data
                if self.normalize_tr(d["city"]) == detected_city.strip()
            ])

            if count == 1:
                return f"📍 {detected_city.title()} ilinde 1 adet aktif bayi bulunmaktadır."

            return f"📍 {detected_city.title()} ilinde {count} adet aktif bayi bulunmaktadır."

        # -------------------------
        # GENEL COUNT
        # -------------------------
        total = len(data)
        return f"🇹🇷 Türkiye genelinde toplam {total} adet aktif bayi bulunmaktadır."
    
    #   BEYNİN temel fonksiyon - fonksiyonlara yönlendirir intent tespiti sonrası çalışır
    @log_process_message
    def process_message(self, user_message: str):

        print("🟨 [PROCESS] Gelen Mesaj:", user_message)

        signals = self.extract_message_signals(user_message)
        print("🧠 [SIGNALS]")

        # Kullanıcı bilgilerini session’dan çek
        user_info = {
            "id": session.get("ID"),
            "username": session.get("username"),
            "fullname": f"{session.get('name', '')} {session.get('surname', '')}".strip(),
            "plaka": session.get("plaka"),
            "plaka2": session.get("plaka2"),
            "departman": session.get("departman"),
            "email": session.get("email"),
            "arac_marka": session.get("arac_marka"),
        }

        #makinenin rahat anlayacağı hale getiriyor ",ğ->g ,A->a, ı->i
        msg_norm = self.normalize_people_text(user_message)

        if (
            re.search(r"\b(adin|adinın|adı|adi)\s+(ne(y|ydi)?|yugii)\b", msg_norm)
            or "ismin ne" in msg_norm or "ismin" in msg_norm
        ):
            return "Ben Yugii’yim 🙂 Sana nasıl yardımcı olabilirim?"
    
        company_only = self.is_company_only(user_message)

        if company_only:
            if "skoda" in self.normalize_people_text(user_message):
                return self.handle_skoda_reputation()
            else:
                return self.handle_yuce_auto_reputation()
            
        #  1) SİNYALLERİ TOPLA
        
        items = list(signals.items())
        for i in range(0, len(items), 3):
            chunk = items[i:i+3]
            print(" | ".join(f"{k}={v}" for k, v in chunk))


        #  2) DOMAIN KARARI VER
        intent = self.detect_business_intent(user_message, signals)
        print("🟪 [DOMAIN]:", intent)

        #  3) domaşn karar sonrası intent fonk seçme
        if intent is None:
            intent = self.detect_main_intent(user_message, signals )

        print("🧠 [FINAL INTENT]:", intent)
            
        # 🔹 3) SADECE İSİM VARSA (AMA SORU DEĞİLSE!)
        if (
            intent is None
            and not signals["has_question"]
            and not any(k in msg_norm for k in [
                "yuce portal",
                "yuceportal",
                "yuce auto","yuce autodaki","yuce autonun",
                "yuce oto",
                "sirket","skoda",
                "portal","yüce portal"
            ])
        ):
            prescan = self.detect_person_prescan(user_message)
            if prescan:
                intent = "company_people"

        
        if intent == "assistant_capabilities":
            return STATIC_CAPABILITY_TEXT

        if any(k in msg_norm for k in KIM_YAPTI_SORUSU):
            return "Beni Yüce Auto’nun IT ekibinden Sude Bayhan geliştirdi 🙂 Teknik sorularınız için Sude Bayhan ile iletişime geçebilirsiniz."
            
        try:
            if not user_info["plaka"] and not user_info["plaka2"]:
                conn = pyodbc.connect(YA_2El_AracSatis)
                cur = conn.cursor()
                cur.execute("""
                    SELECT TOP 1 plaka, plaka2
                    FROM [Yuce_Portal].[dbo].[users1]
                    WHERE LOWER(username) = LOWER(?)
                """, (user_info["username"],))
                row = cur.fetchone()
                conn.close()
                if row:
                    # Önce plaka, yoksa plaka2
                    user_info["plaka"] = row[0] or row[1]
                    print("🧩 users1 fallback → plaka bulundu:", user_info["plaka"])
        except Exception as e:
            print("⚠️ users1 plaka kontrol hatası:", e)

        print("👤 USER INFO:", user_info)
        
        if self.is_password_guide_question(user_message):
            return self.handle_password_guide()
        
        if self.has_password_request(user_message):

            other_fields_requested = False
            norm_msg = self.normalize_people_text(user_message)

            for keywords in RAW_FIELD_KEYWORDS.values():
                for k in keywords:
                    if self.normalize_people_text(k) in norm_msg:
                        other_fields_requested = True
                        break
            if not other_fields_requested :
                return (
                    "🔒 Şifre gibi kişisel ve gizli bilgiler paylaşılamaz.<br>"
                    "Güvenliğiniz için bu tür taleplerde bulunmamanızı rica ederim."
                )

            sanitized = user_message
            for k in self.SENSITIVE_PASSWORD_KEYWORDS:
                sanitized = sanitized.replace(k, "")

            user_message = sanitized
            msg_norm = self.normalize_people_text(user_message)

        if intent == "today_day":
            return self.handle_today_question(msg_norm)
        
        if self.detect_food_intent(user_message):
            print(" [PROCESS] Yemek menüsü isteği algılandı.")
            target_date = self.resolve_food_date(user_message)
            print(" [PROCESS] Menü tarihi:", target_date)

            intro = self.naturalize_food_intro()
            outro = self.naturalize_food_outro()

            
            if target_date is None:
                body = "🍽 Şu anda sadece <b>bugün</b>, <b>yarın</b> ve <b>sonraki gün</b> için yemek menüsü görüntülenebiliyor."
                return f"{intro}<br>{body}"   

            
            if target_date.weekday() >= 5:
                body = "🍽 Hafta sonları yemek menüsü bulunmamaktadır."
                return f"{intro}<br>{body}"   

            row = self.get_food_menu_for_date(target_date)
            readable = target_date.strftime("%d.%m.%Y")

            if not row:
                body = f"🍽 <b>{readable}</b> için menü bilgisi henüz girilmemiş."
                return f"{intro}<br>{body}"   

            
            body = f"🍽 <b>{readable}</b> menüsü:<br>"
            for i in range(1, 7):
                yemek = row[i-1]
                fiyat = row[6 + i - 1]
                if yemek:
                    line = f"{i}. {yemek}"
                    if fiyat:
                        line += f" — {fiyat} ₺"
                    body += line + "<br>"
            if row[-1]:
                body += f"<br>📝 Not: {row[-1]}"

            return f"{intro}<br>{body}<br>{outro}"
 
        if self.is_plate_add_question(user_message):
            return (
                "🚗 Araç plakanı şu an için profil sayfası üzerinden ekleyebilirsin 😊<br><br>"
                "Profil sayfanda araç bilgileri bölümünden ekleme veya güncelleme yapabilirsin.<br>"
                "Yugii üzerinden plaka ekleme için geliştirmeler ilerleyen dönemde planlanıyor."
            )
        
        if intent == "charge_reservation":
            return self.handle_charge_reservation_redirect()
        
        if intent == "weather":
            return self.handle_weather_redirect(user_message)
    
        if intent == "portal_info":
            return self.handle_portal_info()
        
        if intent == "yuce_auto_reputation":
            return self.handle_yuce_auto_reputation()
         
        if intent == "fun_request":
            return self.handle_fun_intent(user_message)
 
        if intent == "currency_rate":
            return self.handle_currency_rate()
        
        if intent == "currency_calculation":
            return self.handle_currency_calculation(user_message)

        if intent == "department_count":
            return self.handle_department_count()
        
        if intent == "department_list":
            return self.handle_department_list()

        is_link = False

        if (
            intent is None
            and not self.contains_department_word(user_message)
            and not self.looks_like_company_people(user_message)
        ):
            is_link = self.detect_link_intent(user_message)
            print("🧪 [TEST] detect_link_intent:", is_link)

        if intent == "help":
            parsed_help = self.help_user_with_parse(user_message)
            return self.llm_with_handle_help(parsed_help, user_message)
        
        if intent == "people_count_only":
            return self.handle_people_count_only()
        
        if intent == "people_export":
            return self.handle_people_export(user_message, user_info, msg_norm)
        
        if intent == "portal_link":
            return self.find_page_link(user_message)
        
        if intent == "otopark_create":
            return self.handle_otopark_create(user_message, user_info)

        if intent == "otopark_cancel":
            return self.handle_otopark_cancel(user_message, user_info)

        if intent == "otopark_status":
            return self.handle_otopark_status(user_message, user_info)

        if intent == "otopark_status_user":
            return self.handle_otopark_status_user(user_message, user_info)
        
        
        if intent == "haftalik_takvim":
            return self.handle_haftalik_takvim(user_message, user_info)
        
        if intent == "work_history":
            return self.control_work_history(user_message, user_info)

        if intent == "company_people":
            return self.handle_company_people(user_message)
        
        if intent == "cancel_without_context":
            return self.handle_cancel_without_context()
        
        if intent == "monthly_trend_analysis":
            data = self.get_trend_analysis()
            return self.format_trend_response(data)
        
        if intent == "bayi_about":
            return self.hanlde_bayi_about(user_message)

        if intent == "satis_sayisi":
            return self.handle_satis_count(msg_norm)
        
        if intent == "workforce_status":
            result = self.handle_workforce_status(user_message, user_info)
            if result:
                return result
            
        # İşlem niyeti var ama intent net değilse → statik netleştirme
        if (
            intent is None
            and self.detect_haftalik_takvim_intent(user_message)
        ):
            return (
                "Takvimde işlem yapabilmem için gün ve çalışma tipini biraz daha net yazar mısın?<br><br>"
                "Örnekler:<br>"
                "• “Haftaya tüm günler ofis olarak işaretle”<br>"
                "• “Gelecek hafta pazartesi–cuma evden çalışma işaretle”"
            )
        
        if intent is None:
            pop_category = self.detect_pop_culture_category(user_message)
            if pop_category:
                return self.generate_pop_culture_response(pop_category)

        #SMALLTALK → ask_gpt (KONTROLLÜ)
        if self.is_simple_smalltalk(user_message):
            print("🟦 [PROCESS] Smalltalk → ask_gpt")
            return self.ask_gpt(user_message)

        # Genel bilgi asistanı else
        if not self.general_assistant.is_blocked(user_message):

            if self.general_assistant.is_profane(user_message):
                return (
                    "Bu şekilde yardımcı olamam. Sormak istediğiniz farklı bir konu varsa yardımcı olmak isterim 🙂. "
                )

            answer = self.general_assistant.ask(user_message)
            return self.general_assistant.sanitize(answer)
            

        print("hiçbir yere girmiyor return metin.")
        return "Sormak istediğin konuyu biraz daha detaylandırabilir misin? Şuan için ne demek istediğini anlamadım veya bu konuda yetkim yok diyebilirim.🙂"

try:
    BrainAIYugii.get_instance()
    print("Yugii Brain ready")
except Exception as e:
    print("⚠️ Yugii Brain preload hatası:", e)

