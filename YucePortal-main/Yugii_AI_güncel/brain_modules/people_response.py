from .shared import *


class PeopleResponseMixin:
    def find_company_people_data(self, parsed: dict):
        """
        Company People DB Fetch (MySQL)
        - users1 tablosundan SADECE whitelist alanları alır
        - Parse yapmaz
        """
        conn = _mysql_conn()
        if conn is None:
            print("❌ MySQL bağlantısı kurulamadı")
            return []

        cursor = conn.cursor(dictionary=True)

        fields_sql = ", ".join(PEOPLE_DB_FIELDS)
        rows = []

        if parsed["target"] in ("person", "people_list") and parsed.get("persons"):

            persons = parsed["persons"]
            conditions = []
            params = []

            print("🟦 [DB QUERY] target:", parsed.get("target"))
            print("🟦 [DB QUERY] persons:", persons)
            print("🟦 [DB QUERY] SELECT FIELDS:", PEOPLE_DB_FIELDS)

            YugiiLogger.get_instance().write_plain(
                "\n[DB QUERY PERSONS]\n"
                f"TARGET: {parsed.get('target')}\n"
                f"PERSONS: {persons}\n"
                + "-" * 50 + "\n"
            )
            
            #  TAM İSİM GELDİYSE → SADECE TAM EŞLEŞME
            if parsed["target"] == "person" and len(persons) == 1 and " " in persons[0]:
                full = persons[0].lower()

                sql = f"""
                    SELECT {fields_sql}
                    FROM users1
                    WHERE is_active = 1
                    AND LOWER(CONCAT(name,' ',surname)) = %s
                """
                cursor.execute(sql, (full,))
                rows = cursor.fetchall()
                return rows
            else:
                for p in persons:
                    conditions.append("LOWER(CONCAT(name,' ',surname)) LIKE %s")
                    params.append(f"%{p.lower()}%")

            conditions_sql = " OR ".join(conditions)

            sql = f"""
                SELECT {fields_sql}
                FROM users1
                WHERE is_active = 1
                AND ({conditions_sql})
            """

            cursor.execute(sql, params)
            self.auto_log_sql(sql, params)
            rows = cursor.fetchall()

        # İSİM + SOYİSİM AKILLI AYRIMI
        if (
            parsed.get("target") == "people_list"
            and len(parsed.get("persons", [])) == 2
            and rows
        ):
            p1, p2 = parsed["persons"]

            exact = [
                r for r in rows
                if r["name"].lower() == p1.lower()
                and r["surname"].lower() == p2.lower()
            ]

            if len(exact) == 1:
                print("🟢 [DB FIX] İsim + Soyisim tek kişi olarak algılandı")

                parsed["target"] = "person"
                parsed["persons"] = [f"{p1} {p2}"]
                rows = exact

            print("🟩 [DB RESULT RAW]:")
            for r in rows:
                print("   ", r)

        # 🏢 DEPARTMAN BAZLI (listeleme)
        elif parsed["target"] == "department" and parsed.get("department"):
            sql = f"""
                SELECT {fields_sql}
                FROM users1
                WHERE is_active = 1
                AND departman = %s
                ORDER BY name, surname
            """

            cursor.execute(sql, (parsed["department"],))
            rows = cursor.fetchall()

        cursor.close()
        conn.close()

        return rows or []

    def extract_department_name(self, msg: str):
        DEPARTMENT_ALIASES = {
            "yönetim kurulu": [
                "yönetim kurulu", "yonetim kurulu", "yk"
            ],
            "stratejik planlama ve mali işler": [
                "stratejik planlama ve mali işler",
                "stratejik planlama",
                "mali işler", "mali isler","muhasebe","maliisler","maliisler"

            ],
            "stratejik planlama ve proje yönetimi": [
                "stratejik planlama ve proje yönetimi",
                "proje yönetimi", "proje yonetimi","proje yönetim","proje yonetim"
            ],
            "insan kaynakları": [
                "insan kaynakları", "insan kaynaklari",
                "ik", "ik ekibi", "hr", "hr ekibi","Insan kaynakları","IK","ık","ınsan kaynakları","ıkda","IK"
            ],
            "satış": [
                "satış", "satis", "satış ekibi", "satis ekibi", "sales"
            ],
            "bayi geliştirme": [
                "bayi geliştirme", "bayi gelistirme","bayi gliştirme"
            ],
            "filo satış": [
                "filo satış", "filo satis", "filo"
            ],
            "ürün": [
                "ürün", "urun", "ürün ekibi", "product","üründeki",
            ],
            "pazarlama": [
                "pazarlama", "marketing"
            ],
            "ssh": [
                "ssh", "satış sonrası", "satis sonrasi","ssh'daki",
                "servis", "teknik servis","sshda"
            ],
            "dijital dönüşüm ve müşteri deneyim yönetimi": [
                "dijital dönüşüm ve müşteri deneyim yönetimi",
                "dijital dönüşüm", "dijital donusum",
                "müşteri deneyimi", "musteri deneyimi"
            ],
            "bilgi teknolojileri": [
                "bilgi teknolojileri", "bilgi islem", "bilgi işlem","bilgi işlemdeki",
                "it", "itciler", "it ekibi", "it departmanı", "information technology","IT","itdeki","bilgi teknolojilerinde","bilgi teknolojileri","bilgi teknolojilerin","İT"
            ],
            "yeni iş modelleri": [
                "yeni iş modelleri", "yeni is modelleri"
            ],
            "iç denetim": [
                "iç denetim", "ic denetim", "denetim"
            ]
        }

        msg_norm = self.normalize_people_text(msg)
        msg_norm = self.normalize_department_suffixes(msg_norm) #da,de,dan için
        
        for dept, aliases in DEPARTMENT_ALIASES.items():
            for a in aliases:
                a_norm = self.normalize_people_text(a)

                # 🔥 SADECE KELİME OLARAK EŞLEŞSİN
                pattern = rf"\b{re.escape(a_norm)}(?:ler|lar|de|da|den|dan|te|ta|nde|nda|inde|ında|indaki|ndaki|dakiler|indekiler)?\b"

                if re.search(pattern, msg_norm):
                    return dept.title()

        return None
    
    #profil link verme company people için
    def enrich_people_with_profile_links(self, data: list):
        """
        data: find_company_people_data çıktısı
        """
        count = len(data)

        # SADECE 1–3 kişi
        if not (1 <= count <= 3):
            for d in data:
                d["profile_url"] = None
            return data

        for d in data:
            username = d.get("username")
            if username:
                d["profile_url"] = (
                    f"https://yuceportal.skoda.com.tr/user_profile_view/{username}"
                )
            else:
                d["profile_url"] = None

        return data

    def compose_company_people_response(self, question: str, parsed: dict, data: list):
        """
        Company People Response Composer
        - PEOPLE_DB_FIELDS içindeki HER alan gösterilebilir
        - Kullanıcının istediği alanlar birebir gösterilir
        - Alan yoksa default set kullanılır
        - LLM sadece anlatır (halüsinasyon yok)
        """
        print("\n🧪 [COMPOSE llm fonksiyonunda:]")
        print("🟨 [COMPOSE] requested fields:", parsed.get("fields"))
        print("🟨 [COMPOSE] available row keys:", list(data[0].keys()) if data else [])
        
        MONTH_NAMES = {
            1: "Ocak",
            2: "Şubat",
            3: "Mart",
            4: "Nisan",
            5: "Mayıs",
            6: "Haziran",
            7: "Temmuz",
            8: "Ağustos",
            9: "Eylül",
            10: "Ekim",
            11: "Kasım",
            12: "Aralık",
        }
        if len(data) > 1:
            responses = []

            for row in data:
                name = f"{row['name']} {row['surname']}"

                for field in parsed.get("fields", []):
                    value = row.get(field)

                    if value:
                        responses.append(
                            f"{name} için {field} bilgisi {value} olarak kayıtlıdır."
                        )
                    else:
                        responses.append(
                            f"{name} için {field} bilgisi bulunmamaktadır."
                        )

            return "<br>".join(responses)
        
        requested_fields = {
            self.normalize_field_name(f)
            for f in parsed.get("fields", [])
        }

        # DEFAULT FIELD SET (hiç alan belirtilmemişse)
        DEFAULT_FIELDS = {
            "name",
            "surname",
            "title",
            "departman",
            "email",
            "TelNo"
        }

        if not requested_fields or "identity" in requested_fields:
            show_fields = list(DEFAULT_FIELDS)
        else:
            # SADECE whitelist + istenenler
            show_fields = [
                f for f in PEOPLE_DB_FIELDS
                if f in requested_fields or f in {"name", "surname"}
            ]
        if parsed["target"] == "department" and parsed.get("fields") == {"name", "surname"}:
            names = [f"{r['name']} {r['surname']}" for r in data]

            if not names:
                return "Bu departmanda aktif çalışan bulunamadı."

            if len(names) == 1:
                people_str = names[0]
            elif len(names) == 2:
                people_str = " ve ".join(names)
            else:
                people_str = ", ".join(names[:-1]) + " ve " + names[-1]

            return f"{parsed['department']} ekibinde yer alan kişiler {people_str}."

        lines = []

        for row in data:
            parts = []
            if "dogum_gun" in show_fields or "dogum_ay" in show_fields:
                gun = row.get("dogum_gun")
                ay = row.get("dogum_ay")

                if gun and ay:
                    ay_adi = MONTH_NAMES.get(int(ay))
                    if ay_adi:
                        parts.append(f"Doğum günü: {gun} {ay_adi}")
                else:
                    if {"dogum_gun", "dogum_ay"} & set(parsed.get("fields", [])):
                        parts.append("Doğum günü: bu bilgi kayıtlarımızda bulunmuyor")

            for f in show_fields:
                if f in {"dogum_gun", "dogum_ay"}:
                    continue

                value = row.get(f)
                if f == "Kidem_Tarihi":
                    if value:
                        try:
                            dt = value if isinstance(value, date) else value.date()
                            formatted = f"{dt.day} {MONTH_NAMES[dt.month]} {dt.year}"
                            parts.append(f"Kıdem tarihi: {formatted}")
                        except Exception:
                            parts.append(f"Kıdem tarihi: {value}")
                    else:
                        if f in requested_fields:
                            parts.append("Kıdem tarihi: bu bilgi kayıtlarımızda bulunmuyor")
                    continue

                if value is not None:
                    parts.append(f"{f}: {value}")
                else:
                    if f in requested_fields:
                        parts.append(f"{f}: bu bilgi kayıtlarımızda bulunmuyor")

            if parts:
                lines.append(" | ".join(parts))

        if not lines:
            # isim yok ama alan sorulmuşsa → kullanıcıdan isim iste
            if not parsed.get("persons"):
                return (
                    "Kimden bahsettiğini tam anlayamadım🤔\n"
                    "İsimle birlikte tekrar sorabilir misin? Sana yardımcı olayım."
                )

            return "Bu bilgiye ait paylaşılabilir bir veri bulunamadı."

        context_text = "\n".join(lines)

        prompt = f"""
        Aşağıda şirket içi rehberden alınmış GERÇEK bilgiler yer almaktadır.

        KESİN VE DEĞİŞMEZ KURALLAR:
        - SADECE aşağıdaki bilgiler kullanılır.
        - ASLA yeni bilgi eklenmez.
        - ASLA yorum, tahmin veya çıkarım yapılmaz.
        - ASLA birinci tekil şahıs kullanılmaz.
        - ASLA "ben", "bana", "benim", "çalışıyorum", "görev yapıyorum" gibi ifadeler yazılmaz.
        - Metin ÜÇÜNCÜ TEKİL ŞAHIS ile yazılmalıdır.
        - Anlatıcı, sistem dışı tarafsız bir anlatıcıdır.
        - Selamlama, hitap veya kapanış yapılmaz.
        - Kullanıcıya doğrudan konuşulmaz.
        - Tek paragraf yazılır.
        - Listeleme, madde işareti veya başlık kullanılmaz.
        - Dil kurumsal, sade ve nettir.

        Bilgiler:
        {context_text}

        Kullanıcı sorusu:
        "{question}"

        Yukarıdaki kurallara UYGUN TEK paragraf cevap yaz.
        """

        try:
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[{"role": "system", "content": prompt}],
                max_tokens=120,
                temperature=0.0
            )
            reply = response.choices[0].message.content.strip()

            profile_lines = []
            if data and data[0].get("profile_url"):
                reply += (
                    f'<br>🔗 <a href="{data[0]["profile_url"]}" target="_blank">'
                    f'Profil sayfası</a>'
                )
            if profile_lines:
                reply = reply + "".join(profile_lines)

            #  Persona guard — Yugii ASLA kişi gibi konuşamaz
            if any(x in reply.lower() for x in ["ben ", "bana ", "benim ", "çalışıyorum"]):
                raise RuntimeError("Persona ihlali: 1. tekil şahıs tespit edildi")

            return reply

        except Exception as e:
            print("❌ compose_company_people_response LLM hata:", e)
            return "Bilgiyi şu anda düzgün şekilde sunamıyorum."
   

    def try_resolve_pending_person(self, text, pending_persons):
        text_norm = self.normalize_people_text(text)
        tokens = text_norm.split()


        for p in pending_persons:
            p_norm = self.normalize_people_text(p)
            p_tokens = p_norm.split()  # ["doga", "ozkutuk"]

            # 🔑 İsim veya soyisimden biri geçiyorsa yeterli
            if any(t in p_tokens for t in tokens):
                return p

        return None

    def handle_people_export(self, user_message, user_info, msg_norm):
        print("📄 [HANDLE] people_export çalışıyor")

        msg = msg_norm
        #  Kullanıcının istediği alanları yakala (mail + telefon vb.)
        requested_fields = set()

        for word, col in FIELD_TO_COLUMN.items():
            if word in msg:
                requested_fields.add(col)

        # 1️⃣ Preset / Alan seçimi
        if requested_fields:
            # mail / telefon gibi alanlar açıkça istenmişse
            selected_columns = list(requested_fields)

        elif any(k in msg for k in ["tüm", "hepsi", "full", "tüm bilgiler"]):
            selected_columns = list(EXPORT_PRESETS["all_allowed"])

        elif any(k in msg for k in [
            "çalışan", "çalışanlar",
            "çalışan listesi", "şirket çalışanları","iş yerindekiler","şirkettekiler","insanlar","çlşanlar","calısanlar"
        ]):
            selected_columns = EXPORT_PRESETS["company_list"]

        else:
            selected_columns = EXPORT_PRESETS["basic_list"]

        for col in CORE_REQUIRED_COLUMNS:
            if col not in selected_columns:
                selected_columns.insert(0, col)

        final_columns = [c for c in selected_columns if c in ALLOWED_COLUMNS]
        print("📄 [HANDLE] Final export columns:", final_columns)

        conn = _mysql_conn()
        cursor = conn.cursor(dictionary=True)

        columns_sql = ", ".join(final_columns)

        dept = self.extract_department_name(user_message)

        query = f"""
            SELECT {columns_sql}
            FROM users1
            WHERE is_active = 1
        """
        params = []

        if dept:
            query += " AND departman = %s"
            params.append(dept)

        cursor.execute(query, params)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        if not rows:
            return "⚠️ Seçilen kriterlere uygun aktif çalışan bulunamadı."

        df = pd.DataFrame(rows, columns=final_columns)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"company_people_{ts}.xlsx"
        export_dir = os.path.join("static", "exports")
        os.makedirs(export_dir, exist_ok=True)
        file_path = os.path.join(export_dir, filename)

        df.to_excel(file_path, index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)
        ws = wb.active   # ← GARANTİLİ

        ws.freeze_panes = "C2"

        alt_fill = PatternFill(start_color="F6F7F8", end_color="F6F7F8", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = alt_fill

        header_fill = PatternFill(start_color="0E3A2F", end_color="0E3A2F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        COLUMN_WIDTHS = {
            "name": 18,
            "surname": 20,
            "username": 22,
            "email": 38,
            "departman": 26,
            "title": 28,
            "TelNo": 22,
            "plaka": 18,
            "arac_plaka2": 18,
            "arac_marka": 22,
            "dogum_gun": 14,
            "dogum_ay": 14,
            "Kidem_Tarihi": 18,
            "linkedin": 40,
        }

        # width
        for idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(col_name, 25)

        # filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(file_path)

        #  Chat cevabı + indirme linki
        download_url = f"/static/exports/{filename}"

        return f"""
        📄 <b>İstediğiniz çalışan listesi hazır.</b><br><br>

        Toplam <b>{len(df)}</b> aktif çalışan için,
        <b>{', '.join(final_columns)}</b> alanlarını içeren bir Excel dosyası oluşturdum.<br><br>

        <div style="
        display:flex;
        align-items:center;
        justify-content:space-between;
        padding:2px 0;
        ">

        <div style="display:flex;align-items:center;gap:6px;">
            📊 <span style="font-weight:500;">Excel dosyası</span>
            <span style="font-size:12px;color:#6b7280;">({len(df)} çalışan)</span>
        </div>

        <a href="{download_url}" download
            style="
            width:17px;
            height:17px;
            border-radius:8px;
            border:1.5px solid #0E3A2F;
            display:flex;
            align-items:center;
            justify-content:center;
            text-decoration:none;
            "
            title="Excel’i indir">

            <svg width="16" height="16" viewBox="0 0 24 24"
                fill="none" stroke="#0E3A2F" stroke-width="2.3"
                stroke-linecap="round" stroke-linejoin="round">
            <path d="M12 3v12"></path>
            <path d="M7 10l5 5 5-5"></path>
            <path d="M5 21h14"></path>
            </svg>

        </a>
        </div>

        <br>
        Başka nasıl yardımcı olabilirim?
        """

